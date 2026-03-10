#!/usr/bin/env python3
"""
Semantic Model Usage Analyzer

Deterministic tool that cross-references TMDL semantic model definitions
against PBIR report definitions to identify measure and column usage.

Usage:
    python scripts/analyze_model_usage.py [search_path] [--format full|unused|json|xlsx]
    python scripts/analyze_model_usage.py --models-path <path> [<path> ...] --reports-path <path> [<path> ...]
    python scripts/analyze_model_usage.py [search_path] --interactive
    python scripts/analyze_model_usage.py --format xlsx -o report.xlsx

Output: Markdown report, JSON, or Excel (.xlsx) showing all measures/columns and their usage status.
The analyzer expects exactly one semantic model and one or more reports.
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Optional


# ── Data Classes ──────────────────────────────────────────────────────────────


@dataclass
class ModelItem:
    item_type: str  # "Measure", "Column", "Calculated Column"
    table: str
    name: str
    display_folder: str = ""
    dax_body: str = ""
    is_hidden: bool = False
    is_key: bool = False
    is_inferred: bool = False
    sort_by_column: str = ""

    @property
    def key(self) -> tuple:
        return (self.table, self.name)


@dataclass
class UsageRef:
    table: str
    name: str
    ref_type: str  # "Measure", "Column", "metadata", "HierarchyLevel"
    report: str = ""
    page: str = ""
    visual_type: str = ""
    visual_title: str = ""
    context: str = ""


@dataclass
class HierarchyInfo:
    table: str
    name: str
    columns: list  # column names referenced by hierarchy levels


# ── Discovery ─────────────────────────────────────────────────────────────────


def _unique_sorted_paths(paths: list[Path]) -> list[Path]:
    unique: dict[Path, Path] = {}
    for p in paths:
        resolved = p.resolve()
        unique[resolved] = resolved
    return sorted(unique.values(), key=lambda p: (p.name.casefold(), str(p).casefold()))


def _discover_artifact_dirs(search_roots: list[Path], suffix: str, conventional_dir: str) -> list[Path]:
    discovered = []

    for root in search_roots:
        if not root.exists():
            continue

        # Direct artifact path support, e.g. /x/Foo.SemanticModel
        if root.is_dir() and root.name.endswith(suffix):
            discovered.append(root)
            continue

        if not root.is_dir():
            continue

        # Fast path for common workspace layout.
        conventional = root / conventional_dir
        if conventional.exists() and conventional.is_dir():
            conventional_hits = [
                p for p in conventional.iterdir()
                if p.is_dir() and p.name.endswith(suffix)
            ]
            discovered.extend(conventional_hits)
            if conventional_hits:
                # Prefer explicit workspace layout and avoid over-scanning.
                continue

        # Fallback recursive discovery for non-standard layouts.
        discovered.extend(
            p for p in root.rglob(f"*{suffix}")
            if p.is_dir()
        )

    return _unique_sorted_paths(discovered)


def discover_models(search_roots: list[Path]) -> list[Path]:
    return _discover_artifact_dirs(search_roots, ".SemanticModel", "Models")


def discover_reports(search_roots: list[Path]) -> list[Path]:
    return _discover_artifact_dirs(search_roots, ".Report", "Reports")


def report_display_name(report_path: Path) -> str:
    return report_path.name.replace(".Report", "")


def filter_models(models: list[Path], model_filters: Optional[list[str]]) -> list[Path]:
    if not model_filters:
        return models
    return [
        m for m in models
        if any(f.lower() in m.name.lower() for f in model_filters)
    ]


def filter_reports(reports: list[Path], report_filters: Optional[list[str]]) -> list[Path]:
    if not report_filters:
        return reports
    return [
        r for r in reports
        if any(f.lower() in report_display_name(r).lower() for f in report_filters)
    ]


def _parse_selection_spec(spec: str, max_index: int) -> list[int]:
    raw = spec.strip().lower()
    if raw in ("", "all", "*"):
        return list(range(1, max_index + 1))

    selected = set()
    for part in raw.split(","):
        token = part.strip()
        if not token:
            continue
        if "-" in token:
            start_s, end_s = token.split("-", 1)
            start = int(start_s)
            end = int(end_s)
            if start > end:
                start, end = end, start
            for idx in range(start, end + 1):
                if 1 <= idx <= max_index:
                    selected.add(idx)
        else:
            idx = int(token)
            if 1 <= idx <= max_index:
                selected.add(idx)

    return sorted(selected)


def select_paths_interactively(label: str, paths: list[Path], formatter) -> list[Path]:
    if not paths:
        return []

    print(f"\nSelect {label} (comma list, range like 1-3, or 'all'):")
    for idx, path in enumerate(paths, start=1):
        print(f"{idx:>3}. {formatter(path)}")

    while True:
        raw = input(f"{label} selection [all]: ").strip()
        try:
            picks = _parse_selection_spec(raw, len(paths))
        except ValueError:
            print("Invalid selection. Use e.g. 1,3-5 or all.")
            continue
        if picks:
            return [paths[i - 1] for i in picks]
        print("No valid items selected. Try again.")


def normalize_key(table: str, name: str) -> tuple[str, str]:
    return (table.casefold(), name.casefold())


def format_item_ref(key: tuple[str, str]) -> str:
    return f"{key[0]}[{key[1]}]"


# ── TMDL Parsing ─────────────────────────────────────────────────────────────


def parse_model_items(model_path: Path) -> list[ModelItem]:
    tables_dir = model_path / "definition" / "tables"
    if not tables_dir.exists():
        return []
    items = []
    for f in sorted(tables_dir.glob("*.tmdl")):
        items.extend(_parse_tmdl_file(f))
    return items


def parse_hierarchies(model_path: Path) -> list[HierarchyInfo]:
    """Extract hierarchy definitions and their backing columns from TMDL."""
    tables_dir = model_path / "definition" / "tables"
    if not tables_dir.exists():
        return []
    hierarchies = []
    for f in sorted(tables_dir.glob("*.tmdl")):
        hierarchies.extend(_parse_tmdl_hierarchies(f))
    return hierarchies


def _parse_tmdl_hierarchies(filepath: Path) -> list[HierarchyInfo]:
    lines = filepath.read_text(encoding="utf-8").splitlines()
    hierarchies = []
    current_table = None
    i = 0

    while i < len(lines):
        line = lines[i]

        if not line.startswith("\t") and line.startswith("table "):
            current_table = line[6:].strip().strip("'\"")
            i += 1
            continue

        if not current_table:
            i += 1
            continue

        # Hierarchy declaration at 1-tab
        h = re.match(r"^\thierarchy\s+'([^']+)'\s*$", line) or \
            re.match(r"^\thierarchy\s+(\S.+?)\s*$", line)
        if h:
            hier_name = h.group(1).strip().strip("'")
            columns = []
            i += 1

            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    i += 1
                    continue
                if inner.startswith("\t\t") and not inner.startswith("\t\t\t"):
                    prop = inner.strip()
                    # level declaration
                    if prop.startswith("level "):
                        i += 1
                        # Look for column: inside the level
                        while i < len(lines):
                            level_inner = lines[i]
                            if level_inner.startswith("\t\t\t"):
                                level_prop = level_inner.strip()
                                if level_prop.startswith("column:"):
                                    col_name = level_prop.split(":", 1)[1].strip().strip("'\"")
                                    columns.append(col_name)
                                i += 1
                                continue
                            break
                        continue
                    i += 1
                    continue
                if inner.startswith("\t") and not inner.startswith("\t\t"):
                    break
                if not inner.startswith("\t"):
                    break
                i += 1

            hierarchies.append(HierarchyInfo(
                table=current_table,
                name=hier_name,
                columns=columns,
            ))
            continue

        i += 1

    return hierarchies


def _parse_tmdl_file(filepath: Path) -> list[ModelItem]:
    lines = filepath.read_text(encoding="utf-8").splitlines()
    items = []
    current_table = None
    i = 0

    while i < len(lines):
        line = lines[i]

        # Table declaration at 0-indent
        if not line.startswith("\t") and line.startswith("table "):
            current_table = line[6:].strip().strip("'\"")
            i += 1
            continue

        if not current_table:
            i += 1
            continue

        # Measure declaration at 1-tab
        m = re.match(r"^\tmeasure\s+'([^']+)'\s*=(.*)", line) or \
            re.match(r"^\tmeasure\s+(.+?)\s*=(.*)", line)
        if m:
            name = m.group(1).strip().strip("'")
            first_dax = m.group(2).strip()
            if first_dax.startswith("```"):
                first_dax = first_dax[3:].strip()

            dax_lines = [first_dax] if first_dax and first_dax != "```" else []
            display_folder = ""
            is_hidden = False
            i += 1

            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    i += 1
                    continue
                if inner.startswith("\t\t\t"):
                    cleaned = inner.strip()
                    if cleaned != "```":
                        dax_lines.append(cleaned)
                    i += 1
                    continue
                if inner.startswith("\t\t") and not inner.startswith("\t\t\t"):
                    prop = inner.strip()
                    if prop.startswith("displayFolder:"):
                        display_folder = prop.split(":", 1)[1].strip()
                    if prop.startswith("hidden:") or prop.startswith("isHidden:"):
                        is_hidden = prop.split(":", 1)[1].strip().lower() == "true"
                    elif prop in ("hidden", "isHidden"):
                        is_hidden = True
                    # formatStringDefinition may contain DAX column refs
                    if prop.startswith("formatStringDefinition"):
                        expr = prop.split("=", 1)
                        if len(expr) > 1:
                            dax_lines.append(expr[1].strip())
                    i += 1
                    continue
                if inner.startswith("\t") and not inner.startswith("\t\t"):
                    break
                if not inner.startswith("\t"):
                    break
                i += 1

            items.append(ModelItem(
                item_type="Measure",
                table=current_table,
                name=name,
                display_folder=display_folder,
                dax_body="\n".join(dax_lines),
                is_hidden=is_hidden,
            ))
            continue

        # Column declaration at 1-tab
        c = re.match(r"^\tcolumn\s+'([^']+)'\s*$", line) or \
            re.match(r"^\tcolumn\s+(.+?)\s*$", line)
        if c:
            name = c.group(1).strip().strip("'")
            is_calculated = False
            display_folder = ""
            is_hidden = False
            is_key = False
            is_inferred = False
            sort_by_column = ""
            dax_lines = []
            i += 1

            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    i += 1
                    continue
                if inner.startswith("\t\t\t"):
                    dax_lines.append(inner.strip())
                    i += 1
                    continue
                if inner.startswith("\t\t") and not inner.startswith("\t\t\t"):
                    prop = inner.strip()
                    if "expression" in prop and "=" in prop and not prop.startswith("formatString"):
                        is_calculated = True
                        expr_part = prop.split("=", 1)[1].strip()
                        if expr_part:
                            dax_lines.append(expr_part)
                    if prop.startswith("displayFolder:"):
                        display_folder = prop.split(":", 1)[1].strip()
                    if prop.startswith("hidden:") or prop.startswith("isHidden:"):
                        is_hidden = prop.split(":", 1)[1].strip().lower() == "true"
                    elif prop in ("hidden", "isHidden"):
                        is_hidden = True
                    if prop.startswith("isKey:"):
                        is_key = prop.split(":", 1)[1].strip().lower() == "true"
                    elif prop == "isKey":
                        is_key = True
                    if prop.startswith("isNameInferred:") or prop.startswith("isDataTypeInferred:"):
                        if prop.split(":", 1)[1].strip().lower() == "true":
                            is_inferred = True
                    elif prop in ("isNameInferred", "isDataTypeInferred"):
                        is_inferred = True
                    if prop.startswith("sortByColumn:"):
                        sort_by_column = prop.split(":", 1)[1].strip().strip("'\"")
                    i += 1
                    continue
                if inner.startswith("\t") and not inner.startswith("\t\t"):
                    break
                if not inner.startswith("\t"):
                    break
                i += 1

            items.append(ModelItem(
                item_type="Calculated Column" if is_calculated else "Column",
                table=current_table,
                name=name,
                display_folder=display_folder,
                dax_body="\n".join(dax_lines),
                is_hidden=is_hidden,
                is_key=is_key,
                is_inferred=is_inferred,
                sort_by_column=sort_by_column,
            ))
            continue

        i += 1

    return items


def parse_relationships(model_path: Path) -> set[tuple[str, str]]:
    rel_file = model_path / "definition" / "relationships.tmdl"
    if not rel_file.exists():
        return set()

    text = rel_file.read_text(encoding="utf-8")
    refs = set()

    # Match: fromColumn/toColumn: 'Table Name'.column  or  Table.column  or  Table.'Column'
    pattern = r"(?:fromColumn|toColumn):\s+(?:'([^']+)'|([^\s.]+))\.(?:'([^']+)'|(\S+))"
    for m in re.finditer(pattern, text):
        table = m.group(1) or m.group(2)
        column = m.group(3) or m.group(4)
        if table and column:
            refs.add((table, column))

    return refs


def parse_rls_roles(model_path: Path) -> list[tuple[str, str, str]]:
    """Returns list of (role_name, table, column) for RLS filter references."""
    results = []
    definition_dir = model_path / "definition"

    files_to_check = []
    model_file = definition_dir / "model.tmdl"
    if model_file.exists():
        files_to_check.append(model_file)
    roles_dir = definition_dir / "roles"
    if roles_dir.exists():
        files_to_check.extend(roles_dir.glob("*.tmdl"))

    for f in files_to_check:
        lines = f.read_text(encoding="utf-8").splitlines()
        current_role = None
        i = 0

        while i < len(lines):
            line = lines[i]

            if line.startswith("role ") and not line.startswith("\t"):
                current_role = line[5:].strip().strip("'\"")
                i += 1
                continue

            if current_role and "filterExpression" in line:
                expr_lines = []
                if "=" in line:
                    expr_lines.append(line.split("=", 1)[1].strip())

                i += 1
                while i < len(lines):
                    inner = lines[i]
                    if inner.startswith("\t\t"):
                        cleaned = inner.strip().strip("`")
                        if cleaned:
                            expr_lines.append(cleaned)
                        i += 1
                        continue
                    break

                expr = " ".join(expr_lines)
                for cm in re.finditer(r"'([^']+)'\[([^\]]+)\]", expr):
                    results.append((current_role, cm.group(1), cm.group(2)))
                for cm in re.finditer(r"(?<!')\b(\w+)\[([^\]]+)\]", expr):
                    results.append((current_role, cm.group(1), cm.group(2)))
                continue

            i += 1

    return results


# ── JSON Reference Scanning ──────────────────────────────────────────────────


def _find_json_refs(obj, path_parts: Optional[list] = None) -> list[dict]:
    """Recursively find all Entity/Property measure/column references in JSON."""
    if path_parts is None:
        path_parts = []
    refs = []

    if isinstance(obj, dict):
        # Measure or Column reference object
        for ref_type in ("Measure", "Column"):
            if ref_type in obj and isinstance(obj[ref_type], dict):
                inner = obj[ref_type]
                if "Property" in inner and "Expression" in inner:
                    expr = inner.get("Expression", {})
                    src = expr.get("SourceRef", {}) if isinstance(expr, dict) else {}
                    if isinstance(src, dict) and "Entity" in src:
                        refs.append({
                            "table": src["Entity"],
                            "name": inner["Property"],
                            "ref_type": ref_type,
                            "path": ".".join(path_parts + [ref_type]),
                        })

        # Aggregation wrapping a Column/Measure
        if "Aggregation" in obj and isinstance(obj["Aggregation"], dict):
            agg_expr = obj["Aggregation"].get("Expression", {})
            if isinstance(agg_expr, dict):
                for ref_type in ("Column", "Measure"):
                    if ref_type in agg_expr:
                        inner = agg_expr[ref_type]
                        if isinstance(inner, dict) and "Property" in inner:
                            src = inner.get("Expression", {}).get("SourceRef", {})
                            if isinstance(src, dict) and "Entity" in src:
                                refs.append({
                                    "table": src["Entity"],
                                    "name": inner["Property"],
                                    "ref_type": ref_type,
                                    "path": ".".join(path_parts + ["Aggregation"]),
                                })

        # Metadata selector: flat "Table.Name" string
        if "metadata" in obj and isinstance(obj["metadata"], str):
            meta = obj["metadata"]
            if "." in meta:
                parts = meta.split(".", 1)
                refs.append({
                    "table": parts[0],
                    "name": parts[1],
                    "ref_type": "metadata",
                    "path": ".".join(path_parts + ["metadata"]),
                })

        # HierarchyLevel reference object
        if "HierarchyLevel" in obj and isinstance(obj["HierarchyLevel"], dict):
            hl = obj["HierarchyLevel"]
            level_name = hl.get("Level", "")
            hier_expr = hl.get("Expression", {})
            if isinstance(hier_expr, dict) and "Hierarchy" in hier_expr:
                hier = hier_expr["Hierarchy"]
                if isinstance(hier, dict):
                    hier_name = hier.get("Hierarchy", "")
                    hier_src = hier.get("Expression", {}).get("SourceRef", {})
                    if isinstance(hier_src, dict) and "Entity" in hier_src:
                        refs.append({
                            "table": hier_src["Entity"],
                            "name": hier_name,
                            "ref_type": "HierarchyLevel",
                            "level": level_name,
                            "path": ".".join(path_parts + ["HierarchyLevel"]),
                        })

        # Recurse
        for key, value in obj.items():
            if key in ("$schema",):
                continue
            refs.extend(_find_json_refs(value, path_parts + [key]))

    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            refs.extend(_find_json_refs(item, path_parts + [f"[{i}]"]))

    return refs


def _determine_context(path: str) -> str:
    if "queryState" in path:
        m = re.search(r"queryState\.(\w+)", path)
        if m:
            role_map = {
                "Category": "Category",
                "Y": "Y-axis",
                "Y2": "Y2-axis",
                "Values": "Values",
                "Rows": "Rows",
                "Columns": "Columns",
                "Data": "Data",
                "Tooltips": "Tooltips",
                "Series": "Series",
            }
            return role_map.get(m.group(1), m.group(1))
    if "sortDefinition" in path:
        return "Sort"
    if "drillthrough" in path.lower():
        return "Drillthrough"
    if "filterConfig" in path:
        return "Filter"
    if "filter" in path.lower() and "filterConfig" not in path:
        return "Filter"
    if "Aggregation" in path:
        return "Aggregation"
    if "objects" in path or "metadata" in path:
        return "Formatting"
    return "Other"


def _get_visual_info(data: dict) -> tuple[str, str]:
    visual = data.get("visual", {})
    visual_type = visual.get("visualType", "unknown")
    title = ""

    # Try visualContainerObjects.title first, then objects.general
    for container_key in ("visualContainerObjects", "objects"):
        container = visual.get(container_key, {})
        for key in ("title", "general"):
            items = container.get(key, [])
            for obj in items:
                props = obj.get("properties", {})
                title_prop = props.get("title", {})
                if isinstance(title_prop, dict):
                    text = title_prop.get("expr", {})
                    if isinstance(text, dict) and "Literal" in text:
                        val = text["Literal"].get("Value", "")
                        title = val.strip("'\"")
                        if title:
                            return visual_type, title

    return visual_type, title


def _get_page_name(page_dir: Path) -> str:
    page_json = page_dir / "page.json"
    if page_json.exists():
        try:
            data = json.loads(page_json.read_text(encoding="utf-8"))
            return data.get("displayName", page_dir.name)
        except (json.JSONDecodeError, KeyError):
            pass
    return page_dir.name


# ── Report Scanning ───────────────────────────────────────────────────────────


def scan_report_visuals(report_path: Path) -> tuple[list[UsageRef], dict]:
    """Returns (usages, visual_meta) where visual_meta maps visual_id -> info."""
    rpt_name = report_display_name(report_path)
    pages_dir = report_path / "definition" / "pages"
    usages = []
    visual_meta = {}

    if not pages_dir.exists():
        return usages, visual_meta

    for page_dir in sorted(pages_dir.iterdir()):
        if not page_dir.is_dir():
            continue

        page_name = _get_page_name(page_dir)
        visuals_dir = page_dir / "visuals"
        if not visuals_dir.exists():
            continue

        for visual_dir in sorted(visuals_dir.iterdir()):
            if not visual_dir.is_dir():
                continue

            visual_json = visual_dir / "visual.json"
            if not visual_json.exists():
                continue

            try:
                data = json.loads(visual_json.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                continue

            visual_type, visual_title = _get_visual_info(data)
            visual_id = visual_dir.name
            visual_refs_local = []

            refs = _find_json_refs(data)
            for ref in refs:
                context = _determine_context(ref["path"])
                u = UsageRef(
                    table=ref["table"],
                    name=ref["name"],
                    ref_type=ref["ref_type"],
                    report=rpt_name,
                    page=page_name,
                    visual_type=visual_type,
                    visual_title=visual_title,
                    context=context,
                )
                usages.append(u)
                visual_refs_local.append((ref["table"], ref["name"]))

            visual_meta[visual_id] = {
                "type": visual_type,
                "title": visual_title,
                "page": page_name,
                "refs": visual_refs_local,
            }

    return usages, visual_meta


def scan_visual_interactions(report_path: Path, visual_meta: dict) -> list[UsageRef]:
    rpt_name = report_display_name(report_path)
    pages_dir = report_path / "definition" / "pages"
    usages = []

    if not pages_dir.exists():
        return usages

    for page_dir in sorted(pages_dir.iterdir()):
        if not page_dir.is_dir():
            continue

        page_json = page_dir / "page.json"
        if not page_json.exists():
            continue

        try:
            data = json.loads(page_json.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue

        page_name = data.get("displayName", page_dir.name)

        # ── Visual interactions (slicer targets) ──
        interactions = data.get("visualInteractions", [])

        for inter in interactions:
            if inter.get("type") != "DataFilter":
                continue

            source_id = inter.get("source", "")
            target_id = inter.get("target", "")
            source_meta = visual_meta.get(source_id, {})
            target_meta = visual_meta.get(target_id, {})

            if source_meta.get("type") == "slicer":
                target_label = target_meta.get("type", "unknown")
                if target_meta.get("title"):
                    target_label += f' "{target_meta["title"]}"'

                for table, name in source_meta.get("refs", []):
                    usages.append(UsageRef(
                        table=table,
                        name=name,
                        ref_type="Column",
                        report=rpt_name,
                        page=page_name,
                        visual_type=f"slicer \u2192 {target_label}",
                        visual_title=source_meta.get("title", ""),
                        context="Slicer Interaction",
                    ))

        # ── Page-level filters ──
        page_filters = data.get("filters", [])
        if isinstance(page_filters, list):
            for filt in page_filters:
                if not isinstance(filt, dict):
                    continue
                refs = _find_json_refs(filt)
                for ref in refs:
                    usages.append(UsageRef(
                        table=ref["table"],
                        name=ref["name"],
                        ref_type=ref["ref_type"],
                        report=rpt_name,
                        page=page_name,
                        visual_type="Page Filter",
                        visual_title="",
                        context="Page Filter",
                    ))

        # ── Drillthrough target fields ──
        drillthrough = data.get("drillthrough", data.get("drillthroughFilters", []))
        if isinstance(drillthrough, list):
            for dt in drillthrough:
                if not isinstance(dt, dict):
                    continue
                refs = _find_json_refs(dt)
                for ref in refs:
                    usages.append(UsageRef(
                        table=ref["table"],
                        name=ref["name"],
                        ref_type=ref["ref_type"],
                        report=rpt_name,
                        page=page_name,
                        visual_type="Drillthrough",
                        visual_title="",
                        context="Drillthrough",
                    ))
        elif isinstance(drillthrough, dict):
            refs = _find_json_refs(drillthrough)
            for ref in refs:
                usages.append(UsageRef(
                    table=ref["table"],
                    name=ref["name"],
                    ref_type=ref["ref_type"],
                    report=rpt_name,
                    page=page_name,
                    visual_type="Drillthrough",
                    visual_title="",
                    context="Drillthrough",
                ))

    return usages


def scan_bookmarks(report_path: Path) -> list[UsageRef]:
    rpt_name = report_display_name(report_path)
    bookmarks_dir = report_path / "definition" / "bookmarks"
    usages = []

    if not bookmarks_dir.exists():
        return usages

    for bm_file in sorted(bookmarks_dir.glob("*.bookmark.json")):
        try:
            data = json.loads(bm_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue

        bm_name = data.get("displayName", bm_file.stem)
        refs = _find_json_refs(data)

        for ref in refs:
            usages.append(UsageRef(
                table=ref["table"],
                name=ref["name"],
                ref_type=ref["ref_type"],
                report=rpt_name,
                page="\u2014",
                visual_type=f"Bookmark: {bm_name}",
                visual_title="",
                context="Bookmark",
            ))

    return usages


def scan_additional_definition_json(
    report_path: Path,
    excluded_files: set[Path],
) -> list[UsageRef]:
    """Fallback scan for model refs in report JSON not covered by visual/page/bookmark scanners."""
    rpt_name = report_display_name(report_path)
    definition_dir = report_path / "definition"
    usages = []

    if not definition_dir.exists():
        return usages

    for json_file in sorted(definition_dir.rglob("*.json")):
        if json_file in excluded_files:
            continue

        try:
            data = json.loads(json_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue

        refs = _find_json_refs(data)
        if not refs:
            continue

        rel_path = json_file.relative_to(definition_dir).as_posix()
        for ref in refs:
            usages.append(UsageRef(
                table=ref["table"],
                name=ref["name"],
                ref_type=ref["ref_type"],
                report=rpt_name,
                page="\u2014",
                visual_type=f"Definition JSON: {rel_path}",
                visual_title="",
                context="Definition Metadata",
            ))

    return usages


# ── DAX Dependency Analysis ──────────────────────────────────────────────────


def _extract_dax_qualified_refs(dax_body: str) -> set[tuple[str, str]]:
    """Extract qualified refs: 'Table'[Name] and Table[Name]."""
    refs = set()
    for m in re.finditer(r"'([^']+)'\[([^\]]+)\]", dax_body):
        refs.add((m.group(1), m.group(2)))
    for m in re.finditer(r"(?<!')\b([A-Za-z_]\w*)\[([^\]]+)\]", dax_body):
        refs.add((m.group(1), m.group(2)))
    return refs


def _extract_dax_unqualified_refs(dax_body: str) -> set[str]:
    """Extract unqualified [Name] references."""
    cleaned = re.sub(r"'[^']+'\[[^\]]+\]", " ", dax_body)
    cleaned = re.sub(r"(?<!')\b[A-Za-z_]\w*\[[^\]]+\]", " ", cleaned)

    refs = set()
    for m in re.finditer(r"\[([^\]]+)\]", cleaned):
        name = m.group(1)
        if not name.startswith("@"):
            refs.add(name)
    return refs


def build_dax_dependency_graph(items: list[ModelItem]) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Build item_key -> {referenced measure keys} graph."""
    measure_keys = {item.key for item in items if item.item_type == "Measure"}
    measure_name_index: dict[str, set[tuple[str, str]]] = defaultdict(set)
    measure_key_index = {normalize_key(*k): k for k in measure_keys}
    for key in measure_keys:
        measure_name_index[key[1].casefold()].add(key)

    deps: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for item in items:
        if not item.dax_body or item.item_type not in ("Measure", "Calculated Column"):
            continue

        refs: set[tuple[str, str]] = set()
        for tbl, name in _extract_dax_qualified_refs(item.dax_body):
            key = measure_key_index.get(normalize_key(tbl, name))
            if key:
                refs.add(key)

        for name in _extract_dax_unqualified_refs(item.dax_body):
            refs |= measure_name_index.get(name.casefold(), set())

        deps[item.key] = refs

    return deps


def build_dax_column_deps(items: list[ModelItem]) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Build item_key -> {(table, column)} graph for column refs in DAX."""
    column_keys = {
        item.key for item in items
        if item.item_type in ("Column", "Calculated Column")
    }
    column_key_index = {normalize_key(*k): k for k in column_keys}

    deps: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for item in items:
        if not item.dax_body or item.item_type not in ("Measure", "Calculated Column"):
            continue

        refs: set[tuple[str, str]] = set()
        for tbl, name in _extract_dax_qualified_refs(item.dax_body):
            key = column_key_index.get(normalize_key(tbl, name))
            if key:
                refs.add(key)

        # In calculated columns, [Name] often refers to another same-table column.
        if item.item_type == "Calculated Column":
            for name in _extract_dax_unqualified_refs(item.dax_body):
                key = column_key_index.get(normalize_key(item.table, name))
                if key:
                    refs.add(key)

        deps[item.key] = refs

    return deps


def resolve_indirect_measures(
    all_items: list[ModelItem],
    directly_used_measures: set[tuple[str, str]],
    dax_deps: dict[tuple[str, str], set[tuple[str, str]]],
) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Transitive closure: find all measures needed by directly-used ones.
    Returns {indirect_measure_key: {measures that depend on it}}."""
    # Forward pass: find ALL measures reachable from directly-used
    all_needed = set(directly_used_measures)
    changed = True
    while changed:
        changed = False
        for measure_key in list(all_needed):
            if measure_key in dax_deps:
                for dep in dax_deps[measure_key]:
                    if dep not in all_needed:
                        all_needed.add(dep)
                        changed = True

    # Build "via" mapping for indirect ones
    indirectly_used = {}
    for item in all_items:
        if item.item_type == "Measure" and item.key in all_needed and item.key not in directly_used_measures:
            via = set()
            for dep_key, refs in dax_deps.items():
                if item.key in refs and dep_key in all_needed:
                    via.add(dep_key)
            indirectly_used[item.key] = via

    return indirectly_used


def resolve_indirect_columns(
    all_items: list[ModelItem],
    directly_used_keys: set[tuple[str, str]],
    relationship_keys: set[tuple[str, str]],
    rls_keys: set[tuple[str, str]],
    all_needed_measures: set[tuple[str, str]],
    dax_col_deps: dict[tuple[str, str], set[tuple[str, str]]],
) -> dict[tuple[str, str], set[str]]:
    """Find columns only used via DAX of needed measures.
    Returns {(table, col): {measure names that reference it}}."""
    already_used = directly_used_keys | relationship_keys | rls_keys
    indirect_cols = {}

    for measure_key in all_needed_measures:
        for col_key in dax_col_deps.get(measure_key, set()):
            col_nkey = normalize_key(*col_key)
            if col_nkey not in already_used:
                if col_nkey not in indirect_cols:
                    indirect_cols[col_nkey] = set()
                indirect_cols[col_nkey].add(format_item_ref(measure_key))

    return indirect_cols


# ── Main Analysis ─────────────────────────────────────────────────────────────


def analyze(
    workspace: Path,
    model_filters: list[str] | None = None,
    report_filters: list[str] | None = None,
    model_paths: list[Path] | None = None,
    report_paths: list[Path] | None = None,
    model_search_roots: list[Path] | None = None,
    report_search_roots: list[Path] | None = None,
) -> dict:
    if model_paths is not None:
        models = _unique_sorted_paths(model_paths)
    else:
        model_roots = model_search_roots or [workspace]
        models = discover_models(model_roots)
        models = filter_models(models, model_filters)

    if report_paths is not None:
        reports = _unique_sorted_paths(report_paths)
    else:
        report_roots = report_search_roots or [workspace]
        reports = discover_reports(report_roots)
        reports = filter_reports(reports, report_filters)

    if not models:
        roots = model_search_roots or [workspace]
        roots_display = ", ".join(str(p) for p in roots)
        print(f"Error: No matching *.SemanticModel found under: {roots_display}", file=sys.stderr)
        sys.exit(1)
    if not reports:
        roots = report_search_roots or [workspace]
        roots_display = ", ".join(str(p) for p in roots)
        print(f"Error: No matching *.Report found under: {roots_display}", file=sys.stderr)
        sys.exit(1)

    # ── Parse model ──
    all_items = []
    all_relationship_cols = set()
    all_rls_refs = []
    all_hierarchies = []

    for model_path in models:
        all_items.extend(parse_model_items(model_path))
        all_relationship_cols |= parse_relationships(model_path)
        all_rls_refs.extend(parse_rls_roles(model_path))
        all_hierarchies.extend(parse_hierarchies(model_path))

    # ── Scan reports ──
    all_visual_usages = []
    all_interaction_usages = []
    all_bookmark_usages = []
    all_definition_meta_usages = []

    for report_path in reports:
        visual_usages, visual_meta = scan_report_visuals(report_path)
        all_visual_usages.extend(visual_usages)
        all_interaction_usages.extend(scan_visual_interactions(report_path, visual_meta))
        all_bookmark_usages.extend(scan_bookmarks(report_path))

        excluded_files = set(report_path.glob("definition/pages/*/visuals/*/visual.json"))
        excluded_files |= set(report_path.glob("definition/pages/*/page.json"))
        excluded_files |= set(report_path.glob("definition/bookmarks/*.bookmark.json"))
        all_definition_meta_usages.extend(
            scan_additional_definition_json(report_path, excluded_files)
        )

    all_usages = (
        all_visual_usages
        + all_interaction_usages
        + all_bookmark_usages
        + all_definition_meta_usages
    )

    # ── Build indices ──
    usage_index: dict[tuple[str, str], list[UsageRef]] = defaultdict(list)
    for u in all_usages:
        usage_index[normalize_key(u.table, u.name)].append(u)

    relationship_keys = {normalize_key(t, c) for t, c in all_relationship_cols}
    rls_keys = {normalize_key(r[1], r[2]) for r in all_rls_refs}
    rls_role_map: dict[tuple[str, str], list[str]] = defaultdict(list)
    for role, tbl, col in all_rls_refs:
        rls_role_map[normalize_key(tbl, col)].append(role)

    # ── DAX dependencies ──
    dax_deps = build_dax_dependency_graph(all_items)
    dax_col_deps = build_dax_column_deps(all_items)

    directly_used_measures = set()
    directly_used_keys = set()
    for item in all_items:
        nkey = normalize_key(*item.key)
        if nkey in usage_index:
            directly_used_keys.add(nkey)
            if item.item_type == "Measure":
                directly_used_measures.add(item.key)

    # Also count relationship/RLS as directly used for transitive analysis
    for item in all_items:
        nkey = normalize_key(*item.key)
        if nkey in relationship_keys or nkey in rls_keys:
            directly_used_keys.add(nkey)

    indirect_measures = resolve_indirect_measures(all_items, directly_used_measures, dax_deps)
    all_needed_measures = directly_used_measures | set(indirect_measures.keys())

    indirect_columns = resolve_indirect_columns(
        all_items, directly_used_keys, relationship_keys, rls_keys,
        all_needed_measures, dax_col_deps,
    )

    # ── Hierarchy index ──
    # Map hierarchy (table, name) -> [column names]
    hierarchy_map: dict[tuple[str, str], list[str]] = {}
    for h in all_hierarchies:
        hierarchy_map[(h.table, h.name)] = h.columns

    # Find which hierarchies are referenced in reports (via HierarchyLevel refs)
    used_hierarchies: set[tuple[str, str]] = set()
    for u in all_usages:
        if u.ref_type == "HierarchyLevel":
            used_hierarchies.add(normalize_key(u.table, u.name))

    # Build hierarchy-column usage: columns backing used hierarchies
    hierarchy_col_keys: dict[tuple[str, str], str] = {}  # nkey -> hierarchy name
    for (tbl, hier_name), columns in hierarchy_map.items():
        hier_nkey = normalize_key(tbl, hier_name)
        if hier_nkey in used_hierarchies:
            for col in columns:
                col_nkey = normalize_key(tbl, col)
                hierarchy_col_keys[col_nkey] = hier_name

    # ── sortByColumn index ──
    sort_by_map: dict[tuple[str, str], str] = {}  # target nkey -> source column name
    for item in all_items:
        if item.sort_by_column and item.item_type in ("Column", "Calculated Column"):
            target_nkey = normalize_key(item.table, item.sort_by_column)
            sort_by_map[target_nkey] = item.name

    # ── isKey index ──
    key_col_keys = {normalize_key(*item.key) for item in all_items if item.is_key}

    # ── Classify each item ──
    results = []
    for item in all_items:
        nkey = normalize_key(*item.key)
        usages = usage_index.get(nkey, [])
        is_relationship = nkey in relationship_keys
        is_rls = nkey in rls_keys
        is_indirect_measure = item.key in indirect_measures and item.item_type == "Measure"
        is_indirect_column = nkey in indirect_columns and item.item_type in ("Column", "Calculated Column")
        is_key_col = nkey in key_col_keys
        is_hierarchy_col = nkey in hierarchy_col_keys
        is_sort_target = nkey in sort_by_map

        if usages:
            status = "USED"
        elif is_relationship:
            status = "USED (Relationship)"
        elif is_rls:
            roles = rls_role_map.get(nkey, [])
            status = f"USED (RLS: {', '.join(roles)})"
        elif is_key_col:
            status = "USED (Key Column)"
        elif is_hierarchy_col:
            status = f"USED (Hierarchy: {hierarchy_col_keys[nkey]})"
        elif is_sort_target:
            # Only mark as sort-used if the column being sorted is itself used
            source_nkey = normalize_key(item.table, sort_by_map[nkey])
            source_used = source_nkey in usage_index or source_nkey in relationship_keys
            if source_used:
                status = f"USED (Sort Column for: {sort_by_map[nkey]})"
            elif is_indirect_measure:
                via = ", ".join(sorted(format_item_ref(k) for k in indirect_measures[item.key]))
                status = f"INDIRECT (via: {via})"
            elif is_indirect_column:
                via = ", ".join(sorted(indirect_columns[nkey]))
                status = f"INDIRECT (via: {via})"
            else:
                status = "NOT USED"
        elif is_indirect_measure:
            via = ", ".join(sorted(format_item_ref(k) for k in indirect_measures[item.key]))
            status = f"INDIRECT (via: {via})"
        elif is_indirect_column:
            via = ", ".join(sorted(indirect_columns[nkey]))
            status = f"INDIRECT (via: {via})"
        else:
            status = "NOT USED"

        # ── Removal risk ──
        if status != "NOT USED":
            removal_risk = ""
        elif item.is_inferred:
            removal_risk = "Do not remove"
        else:
            # Check if any other item's DAX references this item
            has_dax_dependents = False
            for dep_key, refs in dax_deps.items():
                if item.key in refs:
                    has_dax_dependents = True
                    break
            if not has_dax_dependents:
                for dep_key, refs in dax_col_deps.items():
                    if item.key in refs:
                        has_dax_dependents = True
                        break
            if has_dax_dependents:
                removal_risk = "Caution"
            elif item.is_hidden or item.is_key:
                removal_risk = "Review"
            else:
                removal_risk = "Safe"

        results.append({
            "item": item,
            "status": status,
            "usages": usages,
            "removal_risk": removal_risk,
        })

    # ── Table-level summary ──
    table_stats: dict[str, dict] = defaultdict(lambda: {
        "total": 0, "used": 0, "unused": 0,
        "measures": 0, "measures_used": 0,
        "columns": 0, "columns_used": 0,
    })
    for r in results:
        tbl = r["item"].table
        is_measure = r["item"].item_type == "Measure"
        is_used = r["status"] != "NOT USED"
        table_stats[tbl]["total"] += 1
        if is_used:
            table_stats[tbl]["used"] += 1
        else:
            table_stats[tbl]["unused"] += 1
        if is_measure:
            table_stats[tbl]["measures"] += 1
            if is_used:
                table_stats[tbl]["measures_used"] += 1
        else:
            table_stats[tbl]["columns"] += 1
            if is_used:
                table_stats[tbl]["columns_used"] += 1

    # ── Summary stats ──
    summary = {
        "total_measures": sum(1 for i in all_items if i.item_type == "Measure"),
        "total_columns": sum(1 for i in all_items if i.item_type in ("Column", "Calculated Column")),
        "total_calc_columns": sum(1 for i in all_items if i.item_type == "Calculated Column"),
        "used_in_visuals": sum(1 for r in results if r["status"] == "USED"),
        "used_relationship": sum(1 for r in results if "Relationship" in r["status"]),
        "used_rls": sum(1 for r in results if "RLS" in r["status"]),
        "used_key_column": sum(1 for r in results if "Key Column" in r["status"]),
        "used_hierarchy": sum(1 for r in results if "Hierarchy" in r["status"]),
        "used_sort_column": sum(1 for r in results if "Sort Column" in r["status"]),
        "indirect": sum(1 for r in results if r["status"].startswith("INDIRECT")),
        "not_used": sum(1 for r in results if r["status"] == "NOT USED"),
        "total_usage_refs": len(all_usages),
        "models": [m.name for m in models],
        "reports": [report_display_name(r) for r in reports],
        "tables": dict(table_stats),
    }

    return {"items": results, "summary": summary}


# ── Output Formatters ─────────────────────────────────────────────────────────


def _escape_pipe(s: str) -> str:
    return s.replace("|", "\\|")


def format_full(results: dict) -> str:
    lines = []
    s = results["summary"]

    lines.append("# Semantic Model Usage Analysis\n")
    lines.append(f"**Models**: {', '.join(s['models'])}")
    lines.append(f"**Reports**: {', '.join(s['reports'])}\n")

    lines.append("## Summary\n")
    lines.append("| Metric | Count |")
    lines.append("|--------|------:|")
    lines.append(f"| Total measures | {s['total_measures']} |")
    lines.append(f"| Total columns | {s['total_columns']} |")
    lines.append(f"| Calculated columns | {s['total_calc_columns']} |")
    lines.append(f"| Used in visuals | {s['used_in_visuals']} |")
    lines.append(f"| Used in relationships only | {s['used_relationship']} |")
    lines.append(f"| Used in RLS only | {s['used_rls']} |")
    lines.append(f"| Used as key column | {s['used_key_column']} |")
    lines.append(f"| Used in hierarchy | {s['used_hierarchy']} |")
    lines.append(f"| Used as sort column | {s['used_sort_column']} |")
    lines.append(f"| Indirectly used (via DAX) | {s['indirect']} |")
    lines.append(f"| **Not used** | **{s['not_used']}** |")
    lines.append(f"| Total usage references | {s['total_usage_refs']} |")
    lines.append("")

    # ── Table-level summary ──
    table_stats = s.get("tables", {})
    if table_stats:
        lines.append("## Table Summary\n")
        lines.append("| Table | Total | Used | Unused | % Used |")
        lines.append("|-------|------:|-----:|-------:|-------:|")
        for tbl in sorted(table_stats.keys(), key=str.casefold):
            ts = table_stats[tbl]
            total = ts["total"]
            used = ts["used"]
            unused = ts["unused"]
            pct = round(used / total * 100) if total else 0
            marker = " \u2717" if pct == 0 else ""
            lines.append(f"| {_escape_pipe(tbl)} | {total} | {used} | {unused} | {pct}%{marker} |")
        lines.append("")

    lines.append("## Detailed Usage\n")
    lines.append("| Type | Table | Name | Hidden | Display Folder | Report | Page | Visual | Context | Status | Risk |")
    lines.append("|------|-------|------|--------|----------------|--------|------|--------|---------|--------|------|")

    def sort_key(r):
        st = r["status"]
        if st.startswith("INDIRECT"):
            order = 1
        elif st == "NOT USED":
            order = 0
        else:
            order = 2
        return (order, r["item"].table, r["item"].name)

    for r in sorted(results["items"], key=sort_key):
        item = r["item"]
        usages = r["usages"]
        status = r["status"]
        hidden = "\u2713" if item.is_hidden else "\u2014"
        risk = r.get("removal_risk", "")

        if usages:
            for u in usages:
                title_str = f' "{_escape_pipe(u.visual_title)}"' if u.visual_title else ""
                lines.append(
                    f"| {item.item_type} | {_escape_pipe(item.table)} | {_escape_pipe(item.name)} | "
                    f"{hidden} | {_escape_pipe(item.display_folder) or '\u2014'} | {_escape_pipe(u.report)} | "
                    f"{_escape_pipe(u.page)} | {_escape_pipe(u.visual_type)}{title_str} | "
                    f"{_escape_pipe(u.context)} | {_escape_pipe(status)} | {risk} |"
                )
        else:
            lines.append(
                f"| {item.item_type} | {_escape_pipe(item.table)} | {_escape_pipe(item.name)} | "
                f"{hidden} | {_escape_pipe(item.display_folder) or '\u2014'} | \u2014 | \u2014 | \u2014 | \u2014 | "
                f"{_escape_pipe(status)} | {risk} |"
            )

    return "\n".join(lines)


def format_unused(results: dict) -> str:
    lines = []
    s = results["summary"]
    total = s["total_measures"] + s["total_columns"]

    lines.append("# Unused Measures & Columns\n")
    lines.append(f"**Models**: {', '.join(s['models'])} | **Reports**: {', '.join(s['reports'])}\n")
    lines.append(f"**{s['not_used']}** unused out of **{total}** total items\n")

    unused = [r for r in results["items"] if r["status"] == "NOT USED"]

    if not unused:
        lines.append("All measures and columns are used. \u2714")
        return "\n".join(lines)

    # ── Table-level summary (unused tables only) ──
    table_stats = s.get("tables", {})
    fully_unused = [t for t, ts in table_stats.items() if ts["used"] == 0]
    if fully_unused:
        lines.append("## Fully Unused Tables\n")
        for tbl in sorted(fully_unused, key=str.casefold):
            lines.append(f"- **{tbl}** ({table_stats[tbl]['total']} items)")
        lines.append("")

    lines.append("## Unused Items\n")
    lines.append("| Type | Table | Name | Hidden | Display Folder | Risk |")
    lines.append("|------|-------|------|--------|----------------|------|")

    risk_order = {"Safe": 0, "Review": 1, "Caution": 2, "Do not remove": 3}

    for r in sorted(unused, key=lambda x: (
        risk_order.get(x.get("removal_risk", ""), 9),
        x["item"].item_type,
        x["item"].table,
        x["item"].name,
    )):
        item = r["item"]
        hidden = "\u2713" if item.is_hidden else "\u2014"
        risk = r.get("removal_risk", "")
        lines.append(
            f"| {item.item_type} | {_escape_pipe(item.table)} | "
            f"{_escape_pipe(item.name)} | {hidden} | {_escape_pipe(item.display_folder) or '\u2014'} | {risk} |"
        )

    return "\n".join(lines)


def format_json_output(results: dict) -> str:
    output = {
        "summary": results["summary"],
        "items": [],
    }
    for r in results["items"]:
        output["items"].append({
            "type": r["item"].item_type,
            "table": r["item"].table,
            "name": r["item"].name,
            "displayFolder": r["item"].display_folder,
            "isHidden": r["item"].is_hidden,
            "isKey": r["item"].is_key,
            "isInferred": r["item"].is_inferred,
            "sortByColumn": r["item"].sort_by_column or None,
            "status": r["status"],
            "removalRisk": r.get("removal_risk", "") or None,
            "usages": [
                {
                    "report": u.report,
                    "page": u.page,
                    "visualType": u.visual_type,
                    "visualTitle": u.visual_title,
                    "context": u.context,
                }
                for u in r["usages"]
            ],
        })
    return json.dumps(output, indent=2, ensure_ascii=False)


def format_xlsx(results: dict, output_path: str) -> None:
    """Write results to an Excel workbook with Summary, Details, and Unused sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl is required for xlsx output. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"

    def _auto_width(ws, min_width=10, max_width=50):
        for col in ws.columns:
            length = min_width
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = max(length, min(len(str(cell.value)) + 2, max_width))
            ws.column_dimensions[col_letter].width = length

    s = results["summary"]
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    label_font = Font(bold=True)
    light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    warn_font = Font(bold=True, color="FF0000")

    def _section_banner(ws, row, text, cols=2):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = section_fill
            cell.font = section_font
            cell.border = thin_border
        ws.cell(row=row, column=1, value=text)

    def _kv_row(ws, row, label, value, bold_value=False):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=value)
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="right")
        if bold_value:
            c2.font = Font(bold=True)

    # ── Sheet 1: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "4472C4"

    # Title
    row = 1
    ws_sum.cell(row=row, column=1, value="Semantic Model Usage Analysis").font = Font(bold=True, size=16)
    ws_sum.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    # Model / Report info
    row = 3
    _section_banner(ws_sum, row, "Source", 2)
    row += 1
    _kv_row(ws_sum, row, "Models", ", ".join(s["models"]))
    row += 1
    _kv_row(ws_sum, row, "Reports", ", ".join(s["reports"]))

    # ── Overall Metrics ──
    row += 2
    _section_banner(ws_sum, row, "Overall Metrics", 2)
    total_items = s["total_measures"] + s["total_columns"]
    total_used = total_items - s["not_used"]
    overall_metrics = [
        ("Total items", total_items),
        ("Total used", total_used),
        ("Total unused", s["not_used"]),
        ("Usage references", s["total_usage_refs"]),
    ]
    for label, val in overall_metrics:
        row += 1
        _kv_row(ws_sum, row, label, val, bold_value=(label == "Total unused"))

    # ── Measures ──
    row += 2
    _section_banner(ws_sum, row, "Measures", 2)
    measures_used = sum(
        1 for r in results["items"]
        if r["item"].item_type == "Measure" and r["status"] != "NOT USED"
    )
    measures_unused = s["total_measures"] - measures_used
    measure_metrics = [
        ("Total measures", s["total_measures"]),
        ("Used", measures_used),
        ("Unused", measures_unused),
        ("Indirectly used (via DAX)", s["indirect"]),
    ]
    for label, val in measure_metrics:
        row += 1
        _kv_row(ws_sum, row, label, val, bold_value=(label == "Unused"))

    # ── Columns ──
    row += 2
    _section_banner(ws_sum, row, "Columns", 2)
    columns_used = sum(
        1 for r in results["items"]
        if r["item"].item_type in ("Column", "Calculated Column") and r["status"] != "NOT USED"
    )
    columns_unused = s["total_columns"] - columns_used
    column_metrics = [
        ("Total columns", s["total_columns"]),
        ("Calculated columns", s["total_calc_columns"]),
        ("Used", columns_used),
        ("Unused", columns_unused),
        ("Used in visuals", s["used_in_visuals"]),
        ("Used in relationships only", s["used_relationship"]),
        ("Used in RLS only", s["used_rls"]),
        ("Used as key column", s["used_key_column"]),
        ("Used in hierarchy", s["used_hierarchy"]),
        ("Used as sort column", s["used_sort_column"]),
    ]
    for label, val in column_metrics:
        row += 1
        _kv_row(ws_sum, row, label, val, bold_value=(label == "Unused"))

    # ── Table Summary ──
    table_stats = s.get("tables", {})
    if table_stats:
        row += 2
        tbl_headers = ["Table", "Measures", "Measures Used", "Columns", "Columns Used",
                        "Total", "Used", "Unused", "% Used"]
        _section_banner(ws_sum, row, "Table Summary", len(tbl_headers))
        row += 1
        for i, h in enumerate(tbl_headers, 1):
            cell = ws_sum.cell(row=row, column=i, value=h)
            cell.font = label_font
            cell.fill = light_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        for tbl in sorted(table_stats.keys(), key=str.casefold):
            row += 1
            ts = table_stats[tbl]
            total = ts["total"]
            used = ts["used"]
            unused = ts["unused"]
            pct = round(used / total * 100) if total else 0
            values = [
                tbl,
                ts["measures"], ts["measures_used"],
                ts["columns"], ts["columns_used"],
                total, used, unused, f"{pct}%",
            ]
            for i, val in enumerate(values, 1):
                cell = ws_sum.cell(row=row, column=i, value=val)
                cell.border = thin_border
                if i >= 2:
                    cell.alignment = Alignment(horizontal="center")
            if pct == 0:
                ws_sum.cell(row=row, column=9).font = warn_font

    # Column widths
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 45
    for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws_sum.column_dimensions[col_letter].width = 15

    # ── Sheet 2: Details (one row per item, deduplicated) ──
    ws_detail = wb.create_sheet("Details")
    detail_headers = ["Type", "Table", "Name", "Hidden", "Display Folder",
                      "Status", "Removal Risk", "Pages Used", "Visual Types", "Contexts"]
    _write_header(ws_detail, detail_headers)

    risk_order = {"Safe": 0, "Review": 1, "Caution": 2, "Do not remove": 3}
    not_used_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    caution_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    sorted_items = sorted(results["items"], key=lambda x: (
        risk_order.get(x.get("removal_risk", ""), 9),
        x["item"].item_type,
        x["item"].table,
        x["item"].name,
    ))

    for row_idx, r in enumerate(sorted_items, 2):
        item = r["item"]
        usages = r["usages"]
        pages = sorted({u.page for u in usages if u.page and u.page != "\u2014"})
        visual_types = sorted({u.visual_type for u in usages if u.visual_type and u.visual_type != "\u2014"})
        contexts = sorted({u.context for u in usages if u.context and u.context != "\u2014"})

        row_data = [
            item.item_type,
            item.table,
            item.name,
            "Yes" if item.is_hidden else "No",
            item.display_folder or "",
            r["status"],
            r.get("removal_risk", ""),
            ", ".join(pages),
            ", ".join(visual_types),
            ", ".join(contexts),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_detail.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
        if r["status"] == "NOT USED":
            risk = r.get("removal_risk", "")
            fill = caution_fill if risk in ("Caution", "Do not remove") else not_used_fill
            for col in range(1, len(row_data) + 1):
                ws_detail.cell(row=row_idx, column=col).fill = fill

    _auto_width(ws_detail)

    # ── Sheet 3: All References (one row per usage ref) ──
    ws_refs = wb.create_sheet("All References")
    ref_headers = ["Type", "Table", "Name", "Hidden", "Display Folder",
                   "Report", "Page", "Visual Type", "Visual Title", "Context",
                   "Status", "Removal Risk"]
    _write_header(ws_refs, ref_headers)

    row_idx = 2
    for r in sorted_items:
        item = r["item"]
        hidden = "Yes" if item.is_hidden else "No"
        status = r["status"]
        risk = r.get("removal_risk", "")
        if r["usages"]:
            for u in r["usages"]:
                row_data = [
                    item.item_type, item.table, item.name, hidden,
                    item.display_folder or "",
                    u.report, u.page, u.visual_type, u.visual_title or "",
                    u.context, status, risk,
                ]
                for col, val in enumerate(row_data, 1):
                    ws_refs.cell(row=row_idx, column=col, value=val).border = thin_border
                row_idx += 1
        else:
            row_data = [
                item.item_type, item.table, item.name, hidden,
                item.display_folder or "",
                "", "", "", "", "", status, risk,
            ]
            for col, val in enumerate(row_data, 1):
                ws_refs.cell(row=row_idx, column=col, value=val).border = thin_border
            if status == "NOT USED":
                fill = caution_fill if risk in ("Caution", "Do not remove") else not_used_fill
                for col in range(1, len(row_data) + 1):
                    ws_refs.cell(row=row_idx, column=col).fill = fill
            row_idx += 1

    _auto_width(ws_refs)

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


# ── Entry Point ───────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Analyze one TMDL semantic model against one or more PBIR reports",
    )
    parser.add_argument(
        "workspace",
        nargs="?",
        default=".",
        help="Base search path (default: .). Supports non-standard layouts via recursive discovery.",
    )
    parser.add_argument(
        "--format",
        choices=["full", "unused", "json", "xlsx"],
        default="full",
        help="Output format: full (detailed matrix), unused (only unused items), json, xlsx (Excel)",
    )
    parser.add_argument(
        "-o", "--output",
        help="Output file path. Required for xlsx format. For other formats, writes to file instead of stdout.",
    )
    parser.add_argument(
        "--models-path",
        nargs="+",
        help="One or more paths to search for the single .SemanticModel folder to analyze",
    )
    parser.add_argument(
        "--reports-path",
        nargs="+",
        help="One or more paths to search for .Report folders (or direct .Report paths)",
    )
    parser.add_argument(
        "--model",
        nargs="+",
        help="Filter semantic models by name (substring match, case-insensitive)",
    )
    parser.add_argument(
        "--report",
        nargs="+",
        help="Filter reports by name (substring match, case-insensitive)",
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Interactive selector for models and reports",
    )
    args = parser.parse_args()

    workspace = Path(args.workspace).resolve()
    if not workspace.exists() and not args.models_path and not args.reports_path:
        print(f"Error: workspace path does not exist: {workspace}", file=sys.stderr)
        sys.exit(1)

    model_roots = [Path(p).resolve() for p in args.models_path] if args.models_path else [workspace]
    report_roots = [Path(p).resolve() for p in args.reports_path] if args.reports_path else [workspace]

    for p in model_roots + report_roots:
        if not p.exists():
            print(f"Error: search path does not exist: {p}", file=sys.stderr)
            sys.exit(1)

    models = filter_models(discover_models(model_roots), args.model)
    reports = filter_reports(discover_reports(report_roots), args.report)

    if args.interactive:
        if not sys.stdin.isatty():
            print("Error: --interactive requires a TTY terminal.", file=sys.stderr)
            sys.exit(1)
        models = select_paths_interactively(
            "models",
            models,
            lambda p: f"{p.name} ({p})",
        )
        reports = select_paths_interactively(
            "reports",
            reports,
            lambda p: f"{report_display_name(p)} ({p})",
        )

    if len(models) != 1:
        if not models:
            print("Error: No matching semantic model found.", file=sys.stderr)
        else:
            found = ", ".join(str(m) for m in models)
            print(
                "Error: exactly one semantic model must be selected. "
                f"Found {len(models)}: {found}",
                file=sys.stderr,
            )
        sys.exit(1)

    results = analyze(
        workspace,
        model_paths=models,
        report_paths=reports,
        model_search_roots=model_roots,
        report_search_roots=report_roots,
    )

    if args.format == "xlsx":
        output_path = args.output
        if not output_path:
            model_name = models[0].name.replace(".SemanticModel", "") if models else "model"
            output_path = f"{model_name}_usage_analysis.xlsx"
        format_xlsx(results, output_path)
    else:
        if args.format == "full":
            text = format_full(results)
        elif args.format == "unused":
            text = format_unused(results)
        elif args.format == "json":
            text = format_json_output(results)
        if args.output:
            Path(args.output).write_text(text, encoding="utf-8")
            print(f"Report saved to: {args.output}")
        else:
            print(text)


if __name__ == "__main__":
    main()
