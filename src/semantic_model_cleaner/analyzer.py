#!/usr/bin/env python3
"""
Semantic Model Usage Analyzer

Deterministic tool that cross-references TMDL semantic model definitions
against PBIR report definitions to identify measure and column usage.

Usage:
    semantic-model-cleaner [search_path] [--format full|unused|json|xlsx]
    semantic-model-cleaner --models-path <path> [<path> ...] --reports-path <path> [<path> ...]
    semantic-model-cleaner [search_path] --interactive
    semantic-model-cleaner --format xlsx -o report.xlsx

Output: Markdown report, JSON, or Excel (.xlsx) showing all measures/columns and their usage status.
The analyzer expects exactly one semantic model and one or more reports.
"""

import argparse
import json
import re
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

from semantic_model_cleaner.tmdl_identifiers import (
    parse_tmdl_dotted_ref,
    read_single_quoted_name,
    split_tmdl_name_and_expression,
    unquote_tmdl_name,
)


# ── Data Classes ──────────────────────────────────────────────────────────────


@dataclass
class ModelItem:
    item_type: str  # "Measure", "Column", "Calculated Column"
    table: str
    name: str
    display_folder: str = ""
    dax_body: str = ""
    is_hidden: bool = False
    is_key: bool = False
    is_inferred: bool = False
    sort_by_column: str = ""
    source_kind: str = "model"
    source_artifact: str = ""
    source_file: str = ""
    format_string: str = ""
    explicit_measure_refs: tuple[tuple[str, str], ...] = ()

    @property
    def key(self) -> tuple:
        return (self.table, self.name)


@dataclass
class UsageRef:
    table: str
    name: str
    ref_type: str  # "Measure", "Column", "metadata", "HierarchyLevel"
    report: str = ""
    page: str = ""
    visual_type: str = ""
    visual_title: str = ""
    visual_id: str = ""
    context: str = ""
    source_path: str = ""
    artifact_kind: str = ""
    artifact_path: str = ""
    selector_value: str = ""
    is_stale: bool = False
    stale_kind: str = ""
    visual_hidden: bool = False
    page_hidden: bool = False
    visual_x: float | int | None = None
    visual_y: float | int | None = None
    visual_width: float | int | None = None
    visual_height: float | int | None = None


@dataclass
class HierarchyInfo:
    table: str
    name: str
    columns: list  # column names referenced by hierarchy levels


@dataclass
class FieldParameterInfo:
    table: str
    source_file: Path
    targets: list[tuple[str, str]]


@dataclass
class RelationshipInfo:
    name: str
    from_table: str
    from_column: str
    to_table: str
    to_column: str
    from_cardinality: str = ""
    to_cardinality: str = ""
    is_active: bool = True


@dataclass
class AnalyzerWarning:
    code: str
    severity: str
    message: str
    model: str | None = None
    table: str | None = None
    source_file: str | None = None


@dataclass
class ReportIssue:
    severity: str
    issue_type: str
    report: str
    message: str
    page: str = ""
    visual_type: str = ""
    visual_title: str = ""
    visual_id: str = ""
    context: str = ""
    table: str = ""
    name: str = ""
    ref_type: str = ""
    artifact_kind: str = ""
    artifact_path: str = ""
    source_path: str = ""
    selector_value: str = ""
    stale_kind: str = ""
    visual_hidden: bool = False
    page_hidden: bool = False
    visual_x: float | int | None = None
    visual_y: float | int | None = None
    visual_width: float | int | None = None
    visual_height: float | int | None = None
    suggestions: list[dict] = field(default_factory=list)


@dataclass
class UnsupportedMetadataRef:
    area: str
    item_keys: set[tuple[str, str]]
    source_file: str = ""
    possible_hidden_dependency: str = ""
    user_harm: str = ""


REPORT_EXTENSION_PRIMITIVE_TYPES = {
    "Binary",
    "Boolean",
    "Date",
    "DateTime",
    "DateTimeZone",
    "Decimal",
    "Double",
    "Duration",
    "Integer",
    "Json",
    "None",
    "Null",
    "Text",
    "Time",
    "Variant",
}


# ── Discovery ─────────────────────────────────────────────────────────────────


def _unique_sorted_paths(paths: list[Path]) -> list[Path]:
    unique: dict[Path, Path] = {}
    for p in paths:
        resolved = p.resolve()
        unique[resolved] = resolved
    return sorted(unique.values(), key=lambda p: (p.name.casefold(), str(p).casefold()))


def _is_semantic_model_cleaner_source_root(root: Path) -> bool:
    return (
        (root / "pyproject.toml").exists()
        and (root / "src" / "semantic_model_cleaner").is_dir()
    )


def _skip_discovery_path(path: Path, root: Path, *, source_root: bool) -> bool:
    try:
        parts = path.relative_to(root).parts
    except ValueError:
        parts = path.parts
    if any(part.startswith(".") for part in parts):
        return True
    return bool(source_root and parts and parts[0] in {"tests", "examples"})


def _discover_artifact_dirs(search_roots: list[Path], suffix: str, conventional_dir: str) -> list[Path]:
    discovered = []

    for root in search_roots:
        root = root.resolve()
        if not root.exists():
            continue
        source_root = _is_semantic_model_cleaner_source_root(root)

        # Direct artifact path support, e.g. /x/Foo.SemanticModel
        if root.is_dir() and root.name.endswith(suffix):
            discovered.append(root)
            continue

        if not root.is_dir():
            continue

        # Fast path for common workspace layout.
        conventional = root / conventional_dir
        if conventional.exists() and conventional.is_dir():
            conventional_hits = [
                p for p in conventional.iterdir()
                if p.is_dir() and p.name.endswith(suffix) and not _skip_discovery_path(p, root, source_root=source_root)
            ]
            discovered.extend(conventional_hits)
            if conventional_hits:
                # Prefer explicit workspace layout and avoid over-scanning.
                continue

        # Fallback recursive discovery for non-standard layouts.
        discovered.extend(
            p for p in root.rglob(f"*{suffix}")
            if p.is_dir() and not _skip_discovery_path(p, root, source_root=source_root)
        )

    return _unique_sorted_paths(discovered)


def discover_models(search_roots: list[Path]) -> list[Path]:
    return _discover_artifact_dirs(search_roots, ".SemanticModel", "Models")


def discover_reports(search_roots: list[Path]) -> list[Path]:
    return _discover_artifact_dirs(search_roots, ".Report", "Reports")


def report_display_name(report_path: Path) -> str:
    return report_path.name.replace(".Report", "")


def filter_models(models: list[Path], model_filters: Optional[list[str]]) -> list[Path]:
    if not model_filters:
        return models
    return [
        m for m in models
        if any(f.lower() in m.name.lower() for f in model_filters)
    ]


def filter_reports(reports: list[Path], report_filters: Optional[list[str]]) -> list[Path]:
    if not report_filters:
        return reports
    return [
        r for r in reports
        if any(f.lower() in report_display_name(r).lower() for f in report_filters)
    ]


def _parse_selection_spec(spec: str, max_index: int) -> list[int]:
    raw = spec.strip().lower()
    if raw in ("", "all", "*"):
        return list(range(1, max_index + 1))

    selected = set()
    for part in raw.split(","):
        token = part.strip()
        if not token:
            continue
        if "-" in token:
            start_s, end_s = token.split("-", 1)
            start = int(start_s)
            end = int(end_s)
            if start > end:
                start, end = end, start
            for idx in range(start, end + 1):
                if 1 <= idx <= max_index:
                    selected.add(idx)
        else:
            idx = int(token)
            if 1 <= idx <= max_index:
                selected.add(idx)

    return sorted(selected)


def select_paths_interactively(label: str, paths: list[Path], formatter) -> list[Path]:
    if not paths:
        return []

    print(f"\nSelect {label} (comma list, range like 1-3, or 'all'):")
    for idx, path in enumerate(paths, start=1):
        print(f"{idx:>3}. {formatter(path)}")

    while True:
        raw = input(f"{label} selection [all]: ").strip()
        try:
            picks = _parse_selection_spec(raw, len(paths))
        except ValueError:
            print("Invalid selection. Use e.g. 1,3-5 or all.")
            continue
        if picks:
            return [paths[i - 1] for i in picks]
        print("No valid items selected. Try again.")


def normalize_key(table: str, name: str) -> tuple[str, str]:
    return (table.casefold(), name.casefold())


def format_item_ref(key: tuple[str, str]) -> str:
    return f"{key[0]}[{key[1]}]"


def _serialize_warning(warning: AnalyzerWarning) -> dict:
    return {
        "code": warning.code,
        "severity": warning.severity,
        "message": warning.message,
        "model": warning.model,
        "table": warning.table,
        "source_file": warning.source_file,
    }


def _serialize_report_issue(issue: ReportIssue) -> dict:
    return {
        "severity": issue.severity,
        "issueType": issue.issue_type,
        "report": issue.report,
        "page": issue.page,
        "visualType": issue.visual_type,
        "visualTitle": issue.visual_title,
        "visualId": issue.visual_id,
        "context": issue.context,
        "table": issue.table,
        "name": issue.name,
        "refType": issue.ref_type,
        "artifactKind": issue.artifact_kind,
        "artifactPath": issue.artifact_path,
        "sourcePath": issue.source_path,
        "selectorValue": issue.selector_value,
        "staleKind": issue.stale_kind,
        "visualHidden": issue.visual_hidden,
        "pageHidden": issue.page_hidden,
        "visualX": issue.visual_x,
        "visualY": issue.visual_y,
        "visualWidth": issue.visual_width,
        "visualHeight": issue.visual_height,
        "message": issue.message,
        "suggestions": issue.suggestions,
    }


def _add_warning(
    warnings: list[AnalyzerWarning],
    code: str,
    message: str,
    model: str | None = None,
    table: str | None = None,
    source_file: Path | str | None = None,
) -> None:
    warnings.append(AnalyzerWarning(
        code=code,
        severity="warning",
        message=message,
        model=model,
        table=table,
        source_file=str(source_file) if source_file else None,
    ))


# ── TMDL Parsing ─────────────────────────────────────────────────────────────


def _parse_tmdl_keyword_declaration(line: str, keyword: str) -> tuple[str, str, bool] | None:
    stripped = line.strip()
    prefix = f"{keyword} "
    if not stripped.startswith(prefix):
        return None
    name, expression, has_expression = split_tmdl_name_and_expression(stripped[len(prefix):])
    if not name:
        return None
    return name, expression, has_expression


def parse_model_items(model_path: Path) -> list[ModelItem]:
    tables_dir = model_path / "definition" / "tables"
    if not tables_dir.exists():
        return []
    items = []
    for f in sorted(tables_dir.glob("*.tmdl")):
        items.extend(_parse_tmdl_file(f))
    return items


def parse_unsupported_metadata_refs(model_path: Path) -> list[UnsupportedMetadataRef]:
    definition_dir = model_path / "definition"
    perspectives_dir = definition_dir / "perspectives"
    refs: list[UnsupportedMetadataRef] = []

    refs.extend(_parse_unsupported_tmdl_metadata_refs(model_path))
    refs.extend(_parse_culture_translation_refs(model_path))

    if perspectives_dir.exists():
        for filepath in sorted(perspectives_dir.glob("*.tmdl")):
            item_keys = _parse_perspective_item_refs(filepath)
            if item_keys:
                refs.append(UnsupportedMetadataRef(
                    area="Perspectives",
                    item_keys=item_keys,
                    source_file=filepath.relative_to(model_path).as_posix(),
                    possible_hidden_dependency=(
                        "perspective membership may keep this field available outside scanned report visuals"
                    ),
                    user_harm=(
                        "deleting it could break curated perspective views, Excel connections, "
                        "or downstream tools that rely on the perspective"
                    ),
                ))

    return refs


def _unsupported_metadata_info(area: str) -> tuple[str, str]:
    default_dependency = "metadata outside scanned report visuals may reference this field"
    default_harm = "deleting it could break semantic model behavior that the analyzer cannot fully inspect"
    info = {
        "Calculation Groups": (
            "calculation group or calculation item expressions may reference this field",
            "deleting it could break dynamic measure logic or calculation item results",
        ),
        "KPI expressions": (
            "KPI status, target, or trend expressions may reference this field",
            "deleting it could break KPI indicators or scorecard semantics",
        ),
        "Detail rows": (
            "detail row expressions may reference this field",
            "deleting it could break drillthrough/detail-row result sets",
        ),
        "Format string definitions": (
            "dynamic format string expressions may reference this field",
            "deleting it could change or break value formatting in reports and clients",
        ),
        "Data coverage definitions": (
            "data coverage expressions may reference this field",
            "deleting it could break data freshness or coverage metadata used by clients",
        ),
        "Secondary expressions": (
            "secondary expressions may reference this field",
            "deleting it could break alternate semantic expressions used by clients",
        ),
        "Cultures/translations": (
            "culture or translation metadata may reference this field",
            "deleting it could break localized names, descriptions, or translated metadata",
        ),
    }
    return info.get(area, (default_dependency, default_harm))


def _unsupported_ref(
    area: str,
    item_keys: set[tuple[str, str]],
    source_file: Path,
    model_path: Path,
) -> UnsupportedMetadataRef | None:
    if not item_keys:
        return None
    hidden_dependency, user_harm = _unsupported_metadata_info(area)
    return UnsupportedMetadataRef(
        area=area,
        item_keys=item_keys,
        source_file=source_file.relative_to(model_path).as_posix(),
        possible_hidden_dependency=hidden_dependency,
        user_harm=user_harm,
    )


def _extract_item_keys_from_metadata_text(text: str) -> set[tuple[str, str]]:
    item_keys = set(_extract_dax_qualified_refs_from_text(text))
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return item_keys

    for ref in _find_json_refs(data):
        table = str(ref.get("table", "")).strip()
        name = str(ref.get("name", "")).strip()
        if table and name:
            item_keys.add((table, name))
    return item_keys


def _tmdl_indent_depth(line: str) -> int:
    return len(line) - len(line.lstrip("\t"))


def _extract_tmdl_metadata_blocks(text: str, markers: tuple[str, ...]) -> list[str]:
    lines = text.splitlines()
    marker_terms = tuple(marker.casefold() for marker in markers)
    blocks: list[str] = []

    for index, line in enumerate(lines):
        if not any(marker in line.casefold() for marker in marker_terms):
            continue

        base_indent = _tmdl_indent_depth(line)
        block = [line]
        next_index = index + 1
        while next_index < len(lines):
            next_line = lines[next_index]
            if not next_line.strip():
                lookahead = next_index + 1
                while lookahead < len(lines) and not lines[lookahead].strip():
                    lookahead += 1
                if lookahead < len(lines) and _tmdl_indent_depth(lines[lookahead]) > base_indent:
                    block.append(next_line)
                    next_index += 1
                    continue
                break
            if _tmdl_indent_depth(next_line) <= base_indent:
                break
            block.append(next_line)
            next_index += 1

        blocks.append("\n".join(block))

    return blocks


def _parse_unsupported_tmdl_metadata_refs(model_path: Path) -> list[UnsupportedMetadataRef]:
    tables_dir = model_path / "definition" / "tables"
    if not tables_dir.exists():
        return []

    marker_areas = [
        ("Calculation Groups", ("calculationGroup", "calculationItem")),
        ("KPI expressions", ("kpi", "targetExpression", "statusExpression", "trendExpression")),
        ("Detail rows", ("detailRowsDefinition", "defaultDetailRowsExpression")),
        ("Format string definitions", ("formatStringDefinition",)),
        ("Data coverage definitions", ("dataCoverageDefinition",)),
        ("Secondary expressions", ("secondaryExpression", "secondaryExpressions")),
    ]
    refs: list[UnsupportedMetadataRef] = []

    for filepath in sorted(tables_dir.glob("*.tmdl")):
        text = filepath.read_text(encoding="utf-8")
        for area, markers in marker_areas:
            metadata_blocks = _extract_tmdl_metadata_blocks(text, markers)
            if not metadata_blocks:
                continue
            ref = _unsupported_ref(
                area,
                _extract_item_keys_from_metadata_text("\n".join(metadata_blocks)),
                filepath,
                model_path,
            )
            if ref:
                refs.append(ref)

    return refs


def _parse_culture_translation_refs(model_path: Path) -> list[UnsupportedMetadataRef]:
    definition_dir = model_path / "definition"
    candidate_dirs = [
        definition_dir / "cultures",
        definition_dir / "translations",
    ]
    refs: list[UnsupportedMetadataRef] = []

    for candidate_dir in candidate_dirs:
        if not candidate_dir.exists():
            continue
        for filepath in sorted(candidate_dir.rglob("*")):
            if not filepath.is_file() or filepath.suffix.casefold() not in (".tmdl", ".json"):
                continue
            ref = _unsupported_ref(
                "Cultures/translations",
                _extract_item_keys_from_metadata_text(filepath.read_text(encoding="utf-8")),
                filepath,
                model_path,
            )
            if ref:
                refs.append(ref)

    return refs


def parse_hierarchies(model_path: Path) -> list[HierarchyInfo]:
    """Extract hierarchy definitions and their backing columns from TMDL."""
    tables_dir = model_path / "definition" / "tables"
    if not tables_dir.exists():
        return []
    hierarchies = []
    for f in sorted(tables_dir.glob("*.tmdl")):
        hierarchies.extend(_parse_tmdl_hierarchies(f))
    return hierarchies


def _parse_perspective_item_refs(filepath: Path) -> set[tuple[str, str]]:
    lines = filepath.read_text(encoding="utf-8").splitlines()
    item_keys: set[tuple[str, str]] = set()
    current_table = ""

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("perspectiveTable "):
            current_table = unquote_tmdl_name(stripped[len("perspectiveTable "):])
            continue
        if not current_table:
            continue
        for keyword in ("perspectiveMeasure", "perspectiveColumn"):
            prefix = f"{keyword} "
            if stripped.startswith(prefix):
                name = unquote_tmdl_name(stripped[len(prefix):])
                if name:
                    item_keys.add((current_table, name))

    return item_keys


def _unsupported_metadata_review_trigger(
    item: ModelItem,
    ref: UnsupportedMetadataRef,
) -> str:
    source = f" in {ref.source_file}" if ref.source_file else ""
    return (
        f"Unsupported Metadata: {ref.area}{source} can reference {format_item_ref(item.key)}. "
        f"Hidden dependency: {ref.possible_hidden_dependency}. "
        f"User harm: {ref.user_harm}."
    )


def _extract_nameof_targets(text: str) -> list[tuple[str, str]]:
    targets = []
    active_text, _ = _split_dax_comments(text)
    for match in re.finditer(r"NAMEOF\s*\(", active_text, flags=re.IGNORECASE):
        close = active_text.find(")", match.end())
        if close == -1:
            continue
        for table, name, _, _ in _scan_dax_qualified_refs(active_text[match.end():close]):
            targets.append((table, name))
    return targets


def _iter_tmdl_table_sections(lines: list[str]):
    current_table = None
    i = 0

    while i < len(lines):
        line = lines[i]

        if not line.startswith("\t") and line.startswith("table "):
            current_table = unquote_tmdl_name(line[6:])
            i += 1
            continue

        if not current_table or not line.startswith("\t") or line.startswith("\t\t"):
            i += 1
            continue

        block_lines = [line]
        header = line.strip()
        i += 1

        while i < len(lines):
            inner = lines[i]
            if inner.startswith("\t") and not inner.startswith("\t\t"):
                break
            if not inner.startswith("\t") and inner.strip():
                break
            block_lines.append(inner)
            i += 1

        yield current_table, header, "\n".join(block_lines)


def parse_field_parameters(
    model_path: Path,
    warnings: list[AnalyzerWarning] | None = None,
) -> list[FieldParameterInfo]:
    tables_dir = model_path / "definition" / "tables"
    if not tables_dir.exists():
        return []

    field_parameters = []
    model_name = model_path.name

    for filepath in sorted(tables_dir.glob("*.tmdl")):
        text = filepath.read_text(encoding="utf-8")
        if "NAMEOF(" not in text.upper():
            continue

        valid_targets = []
        invalid_nameof = False
        lines = text.splitlines()

        for current_table, header, block_text in _iter_tmdl_table_sections(lines):
            if "NAMEOF(" not in block_text.upper():
                continue

            header_lc = header.casefold()
            is_supported_table_block = (
                header_lc.startswith("partition ")
                or header_lc.startswith("source")
                or header_lc.startswith("expression")
            )

            if is_supported_table_block:
                valid_targets.extend(_extract_nameof_targets(block_text))
            else:
                invalid_nameof = True
                if warnings is not None:
                    _add_warning(
                        warnings,
                        "NAMEOF_PATTERN_NOT_IN_FIELD_PARAMETER_TABLE",
                        (
                            "Found NAMEOF(...) outside a table-level calculated-table "
                            f"source/expression block in {filepath.name}; skipping that pattern."
                        ),
                        model=model_name,
                        table=current_table,
                        source_file=filepath,
                    )

        if valid_targets:
            table_name = next(
                (table for table, _, block_text in _iter_tmdl_table_sections(lines) if "NAMEOF(" in block_text.upper()),
                filepath.stem,
            )
            field_parameters.append(FieldParameterInfo(
                table=table_name,
                source_file=filepath,
                targets=valid_targets,
            ))
        elif "NAMEOF(" in text.upper() and not invalid_nameof and warnings is not None:
            _add_warning(
                warnings,
                "NAMEOF_PATTERN_NOT_IN_FIELD_PARAMETER_TABLE",
                (
                    "Found NAMEOF(...) in "
                    f"{filepath.name}, but not inside a supported field-parameter table source block."
                ),
                model=model_name,
                source_file=filepath,
            )

    return field_parameters


def resolve_field_parameter_targets(
    field_parameters: list[FieldParameterInfo],
    items: list[ModelItem],
    model_name: str,
    warnings: list[AnalyzerWarning],
) -> list[tuple[FieldParameterInfo, list[ModelItem]]]:
    item_index: dict[tuple[str, str], list[ModelItem]] = defaultdict(list)
    for item in items:
        item_index[normalize_key(*item.key)].append(item)

    resolved = []
    for info in field_parameters:
        resolved_targets = []
        seen_keys = set()
        unresolved_targets = set()

        for target_table, target_name in info.targets:
            matches = item_index.get(normalize_key(target_table, target_name), [])
            if len(matches) == 1:
                match = matches[0]
                if match.key not in seen_keys:
                    resolved_targets.append(match)
                    seen_keys.add(match.key)
            elif len(matches) == 0:
                unresolved_key = normalize_key(target_table, target_name)
                if unresolved_key not in unresolved_targets:
                    unresolved_targets.add(unresolved_key)
                    _add_warning(
                        warnings,
                        "UNRESOLVED_NAMEOF_TARGET",
                        (
                            f"NAMEOF target {target_table}[{target_name}] from field parameter table "
                            f"{info.table} does not resolve to a model item."
                        ),
                        model=model_name,
                        table=info.table,
                        source_file=info.source_file,
                    )
                continue
            else:
                match_types = ", ".join(sorted(item.item_type for item in matches))
                _add_warning(
                    warnings,
                    "AMBIGUOUS_NAMEOF_TARGET",
                    (
                        f"NAMEOF target {target_table}[{target_name}] from field parameter table "
                        f"{info.table} matched multiple items ({match_types})."
                    ),
                    model=model_name,
                    table=info.table,
                    source_file=info.source_file,
                )

        resolved.append((info, resolved_targets))

    return resolved


def promote_field_parameter_usages(
    direct_usages: list[UsageRef],
    resolved_field_parameters: list[tuple[FieldParameterInfo, list[ModelItem]]],
) -> tuple[list[UsageRef], dict[tuple[str, str], set[str]]]:
    usages_by_table: dict[str, list[UsageRef]] = defaultdict(list)
    for usage in direct_usages:
        usages_by_table[usage.table.casefold()].append(usage)

    synthetic_usages = []
    promoted_tables: dict[tuple[str, str], set[str]] = defaultdict(set)

    for info, targets in resolved_field_parameters:
        table_usages = usages_by_table.get(info.table.casefold(), [])
        if not table_usages:
            continue

        for target in targets:
            target_nkey = normalize_key(*target.key)
            promoted_tables[target_nkey].add(info.table)
            ref_type = "Measure" if target.item_type == "Measure" else "Column"

            for origin in table_usages:
                synthetic_usages.append(UsageRef(
                    table=target.table,
                    name=target.name,
                    ref_type=ref_type,
                    report=origin.report,
                    page=origin.page,
                    visual_type=origin.visual_type,
                    visual_title=origin.visual_title,
                    context="Field Parameter",
                ))

    return synthetic_usages, promoted_tables


def _parse_tmdl_hierarchies(filepath: Path) -> list[HierarchyInfo]:
    lines = filepath.read_text(encoding="utf-8").splitlines()
    hierarchies = []
    current_table = None
    i = 0

    while i < len(lines):
        line = lines[i]

        if not line.startswith("\t") and line.startswith("table "):
            current_table = unquote_tmdl_name(line[6:])
            i += 1
            continue

        if not current_table:
            i += 1
            continue

        # Hierarchy declaration at 1-tab
        h = _parse_tmdl_keyword_declaration(line, "hierarchy")
        if h:
            hier_name = h[0]
            columns = []
            i += 1

            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    i += 1
                    continue
                if inner.startswith("\t\t") and not inner.startswith("\t\t\t"):
                    prop = inner.strip()
                    # level declaration
                    if prop.startswith("level "):
                        i += 1
                        # Look for column: inside the level
                        while i < len(lines):
                            level_inner = lines[i]
                            if level_inner.startswith("\t\t\t"):
                                level_prop = level_inner.strip()
                                if level_prop.startswith("column:"):
                                    col_name = unquote_tmdl_name(level_prop.split(":", 1)[1])
                                    columns.append(col_name)
                                i += 1
                                continue
                            break
                        continue
                    i += 1
                    continue
                if inner.startswith("\t") and not inner.startswith("\t\t"):
                    break
                if not inner.startswith("\t"):
                    break
                i += 1

            hierarchies.append(HierarchyInfo(
                table=current_table,
                name=hier_name,
                columns=columns,
            ))
            continue

        i += 1

    return hierarchies


def _parse_tmdl_file(filepath: Path) -> list[ModelItem]:
    lines = filepath.read_text(encoding="utf-8").splitlines()
    items = []
    current_table = None
    i = 0

    while i < len(lines):
        line = lines[i]

        # Table declaration at 0-indent
        if not line.startswith("\t") and line.startswith("table "):
            current_table = unquote_tmdl_name(line[6:])
            i += 1
            continue

        if not current_table:
            i += 1
            continue

        # Measure declaration at 1-tab
        m = _parse_tmdl_keyword_declaration(line, "measure")
        if m:
            name, first_dax, _ = m
            if first_dax.startswith("```"):
                first_dax = first_dax[3:].strip()

            dax_lines = [first_dax] if first_dax and first_dax != "```" else []
            display_folder = ""
            is_hidden = False
            i += 1

            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    i += 1
                    continue
                if inner.startswith("\t\t\t"):
                    cleaned = inner.strip()
                    if cleaned != "```":
                        dax_lines.append(cleaned)
                    i += 1
                    continue
                if inner.startswith("\t\t") and not inner.startswith("\t\t\t"):
                    prop = inner.strip()
                    if prop.startswith("displayFolder:"):
                        display_folder = prop.split(":", 1)[1].strip()
                    if prop.startswith("hidden:") or prop.startswith("isHidden:"):
                        is_hidden = prop.split(":", 1)[1].strip().lower() == "true"
                    elif prop in ("hidden", "isHidden"):
                        is_hidden = True
                    # formatStringDefinition may contain DAX column refs
                    if prop.startswith("formatStringDefinition"):
                        expr = prop.split("=", 1)
                        if len(expr) > 1:
                            dax_lines.append(expr[1].strip())
                    i += 1
                    continue
                if inner.startswith("\t") and not inner.startswith("\t\t"):
                    break
                if not inner.startswith("\t"):
                    break
                i += 1

            items.append(ModelItem(
                item_type="Measure",
                table=current_table,
                name=name,
                display_folder=display_folder,
                dax_body="\n".join(dax_lines),
                is_hidden=is_hidden,
                source_file=str(filepath),
            ))
            continue

        # Column declaration at 1-tab
        c = _parse_tmdl_keyword_declaration(line, "column")
        if c:
            name, first_dax, _ = c
            is_calculated = bool(first_dax)
            display_folder = ""
            is_hidden = False
            is_key = False
            is_inferred = False
            sort_by_column = ""
            dax_lines = [first_dax] if first_dax else []
            i += 1

            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    i += 1
                    continue
                if inner.startswith("\t\t\t"):
                    dax_lines.append(inner.strip())
                    i += 1
                    continue
                if inner.startswith("\t\t") and not inner.startswith("\t\t\t"):
                    prop = inner.strip()
                    if "expression" in prop and "=" in prop and not prop.startswith("formatString"):
                        is_calculated = True
                        expr_part = prop.split("=", 1)[1].strip()
                        if expr_part:
                            dax_lines.append(expr_part)
                    if prop.startswith("displayFolder:"):
                        display_folder = prop.split(":", 1)[1].strip()
                    if prop.startswith("hidden:") or prop.startswith("isHidden:"):
                        is_hidden = prop.split(":", 1)[1].strip().lower() == "true"
                    elif prop in ("hidden", "isHidden"):
                        is_hidden = True
                    if prop.startswith("isKey:"):
                        is_key = prop.split(":", 1)[1].strip().lower() == "true"
                    elif prop == "isKey":
                        is_key = True
                    if prop.startswith("isNameInferred:") or prop.startswith("isDataTypeInferred:"):
                        if prop.split(":", 1)[1].strip().lower() == "true":
                            is_inferred = True
                    elif prop in ("isNameInferred", "isDataTypeInferred"):
                        is_inferred = True
                    if prop.startswith("sortByColumn:"):
                        sort_by_column = unquote_tmdl_name(prop.split(":", 1)[1])
                    i += 1
                    continue
                if inner.startswith("\t") and not inner.startswith("\t\t"):
                    break
                if not inner.startswith("\t"):
                    break
                i += 1

            items.append(ModelItem(
                item_type="Calculated Column" if is_calculated else "Column",
                table=current_table,
                name=name,
                display_folder=display_folder,
                dax_body="\n".join(dax_lines),
                is_hidden=is_hidden,
                is_key=is_key,
                is_inferred=is_inferred,
                sort_by_column=sort_by_column,
                source_file=str(filepath),
            ))
            continue

        i += 1

    return items


def parse_report_extension_measures(
    report_path: Path,
    warnings: Optional[list[AnalyzerWarning]] = None,
) -> list[ModelItem]:
    report_extensions = report_path / "definition" / "reportExtensions.json"
    if not report_extensions.exists():
        return []

    try:
        data = json.loads(report_extensions.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        if warnings is not None:
            _add_warning(
                warnings,
                code="INVALID_REPORT_EXTENSION_JSON",
                message=f"Could not parse reportExtensions.json in {report_display_name(report_path)}: {exc}",
                source_file=report_extensions,
            )
        return []

    items = []
    for entity in data.get("entities", []):
        if not isinstance(entity, dict):
            continue
        entity_name = str(entity.get("name", "")).strip()
        if not entity_name:
            continue
        for measure in entity.get("measures", []):
            if not isinstance(measure, dict):
                continue
            measure_name = str(measure.get("name", "")).strip()
            expression = str(measure.get("expression", "") or "")
            if not measure_name or not expression:
                continue
            explicit_refs = []
            references = measure.get("references", {})
            if isinstance(references, dict):
                for ref in references.get("measures", []):
                    if not isinstance(ref, dict):
                        continue
                    ref_entity = str(ref.get("entity", "")).strip()
                    ref_name = str(ref.get("name", "")).strip()
                    if ref_entity and ref_name:
                        explicit_refs.append((ref_entity, ref_name))

            items.append(ModelItem(
                item_type="Measure",
                table=entity_name,
                name=measure_name,
                display_folder=str(measure.get("displayFolder", "") or ""),
                dax_body=expression,
                is_hidden=bool(measure.get("hidden", False)),
                source_kind="report",
                source_artifact=report_display_name(report_path),
                source_file=str(report_extensions),
                format_string=str(measure.get("formatString", "") or ""),
                explicit_measure_refs=tuple(explicit_refs),
            ))

    return items


def _report_extension_item_ref(entity_name: str, measure: dict) -> str:
    measure_name = str(measure.get("name", "")).strip()
    if entity_name and measure_name:
        return f"'{entity_name}[{measure_name}]'"
    if entity_name:
        return f"in entity '{entity_name}'"
    if measure_name:
        return f"'{measure_name}'"
    return ""


def _report_extension_measure_subject(entity_name: str, measure: dict) -> str:
    item_ref = _report_extension_item_ref(entity_name, measure)
    if item_ref:
        return f"Report extension measure {item_ref}"
    return "Report extension measure"


def _report_extension_issue(
    *,
    report_path: Path,
    issue_type: str,
    message: str,
) -> ReportIssue:
    report_extensions = report_path / "definition" / "reportExtensions.json"
    return ReportIssue(
        severity="warning",
        issue_type=issue_type,
        report=report_display_name(report_path),
        artifact_kind="Report Extension",
        artifact_path=_artifact_rel_path(report_path, report_extensions),
        message=message,
    )


def scan_report_extension_issues(report_path: Path) -> list[ReportIssue]:
    report_extensions = report_path / "definition" / "reportExtensions.json"
    if not report_extensions.exists():
        return []

    try:
        data = json.loads(report_extensions.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []

    issues: list[ReportIssue] = []
    if not isinstance(data, dict):
        return [
            _report_extension_issue(
                report_path=report_path,
                issue_type="invalid_report_extension",
                message="Report extension payload must be an object.",
            )
        ]

    for field in ("$schema", "name"):
        value = data.get(field)
        if not isinstance(value, str) or not value.strip():
            issues.append(_report_extension_issue(
                report_path=report_path,
                issue_type="invalid_report_extension",
                message=f"Report extension is missing required field '{field}'.",
            ))

    entities = data.get("entities", [])
    if entities is None:
        entities = []
    if not isinstance(entities, list):
        issues.append(_report_extension_issue(
            report_path=report_path,
            issue_type="invalid_report_extension_entity",
            message="Report extension field 'entities' must be an array.",
        ))
        return issues

    for entity_idx, entity in enumerate(entities):
        if not isinstance(entity, dict):
            issues.append(_report_extension_issue(
                report_path=report_path,
                issue_type="invalid_report_extension_entity",
                message=f"Report extension entity at entities[{entity_idx}] must be an object.",
            ))
            continue

        entity_name = str(entity.get("name", "")).strip()
        if not entity_name:
            issues.append(_report_extension_issue(
                report_path=report_path,
                issue_type="invalid_report_extension_entity",
                message=f"Report extension entity at entities[{entity_idx}] is missing required field 'name'.",
            ))

        measures = entity.get("measures", [])
        if measures is None:
            measures = []
        if not isinstance(measures, list):
            issues.append(_report_extension_issue(
                report_path=report_path,
                issue_type="invalid_report_extension_measure",
                message=f"Report extension measures for entity '{entity_name or entity_idx}' must be an array.",
            ))
            continue

        for measure_idx, measure in enumerate(measures):
            if not isinstance(measure, dict):
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="invalid_report_extension_measure",
                    message=(
                        "Report extension measure at "
                        f"entities[{entity_idx}].measures[{measure_idx}] must be an object."
                    ),
                ))
                continue

            subject = _report_extension_measure_subject(entity_name, measure)
            for field in ("dataType", "expression", "name"):
                value = measure.get(field)
                if not isinstance(value, str) or not value.strip():
                    issues.append(_report_extension_issue(
                        report_path=report_path,
                        issue_type="invalid_report_extension_measure",
                        message=f"{subject} is missing required field '{field}'.",
                    ))

            data_type = measure.get("dataType")
            if (
                isinstance(data_type, str)
                and data_type.strip()
                and data_type not in REPORT_EXTENSION_PRIMITIVE_TYPES
            ):
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="invalid_report_extension_measure",
                    message=f"{subject} has unsupported dataType '{data_type}'.",
                ))

            hidden = measure.get("hidden")
            if hidden is not None and not isinstance(hidden, bool):
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="invalid_report_extension_measure",
                    message=f"{subject} field 'hidden' must be a boolean.",
                ))

            annotations = measure.get("annotations")
            if annotations is not None and not isinstance(annotations, list):
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="invalid_report_extension_measure",
                    message=f"{subject} field 'annotations' must be an array.",
                ))
            elif isinstance(annotations, list):
                for annotation_idx, annotation in enumerate(annotations):
                    if not isinstance(annotation, dict):
                        issues.append(_report_extension_issue(
                            report_path=report_path,
                            issue_type="invalid_report_extension_measure",
                            message=(
                                f"{subject} annotation at annotations[{annotation_idx}] "
                                "must be an object."
                            ),
                        ))
                        continue
                    for field in ("name", "value"):
                        value = annotation.get(field)
                        if not isinstance(value, str) or not value.strip():
                            issues.append(_report_extension_issue(
                                report_path=report_path,
                                issue_type="invalid_report_extension_measure",
                                message=(
                                    f"{subject} annotation at annotations[{annotation_idx}] "
                                    f"is missing required field '{field}'."
                                ),
                            ))

            references = measure.get("references")
            if references is None:
                continue
            if not isinstance(references, dict):
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="invalid_report_extension_measure",
                    message=f"{subject} field 'references' must be an object.",
                ))
                continue

            unrecognized = references.get("unrecognizedReferences")
            if unrecognized is True:
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="report_extension_unrecognized_reference",
                    message=f"{subject} contains unrecognized references.",
                ))
            elif unrecognized not in (None, False):
                issues.append(_report_extension_issue(
                    report_path=report_path,
                    issue_type="invalid_report_extension_measure",
                    message=(
                        f"{subject} field 'references.unrecognizedReferences' "
                        "must be a boolean."
                    ),
                ))

    return issues


def _dedupe_items_with_warnings(
    items: list[ModelItem],
    warnings: list[AnalyzerWarning],
) -> list[ModelItem]:
    seen: dict[tuple[str, str, str], ModelItem] = {}
    duplicates: dict[tuple[str, str, str], list[ModelItem]] = defaultdict(list)

    for item in items:
        nkey = normalize_key(*item.key) + (item.item_type.casefold(),)
        if nkey in seen:
            duplicates[nkey].append(item)
            continue
        seen[nkey] = item

    for nkey, dup_items in duplicates.items():
        first = seen[nkey]
        dup_labels = [f"{first.source_kind}:{first.source_artifact or '-'}"]
        dup_labels.extend(f"{item.source_kind}:{item.source_artifact or '-'}" for item in dup_items)
        _add_warning(
            warnings,
            code="DUPLICATE_ITEM_KEY",
            message=(
                f"Duplicate item key {format_item_ref(first.key)} was found across sources "
                f"({', '.join(dup_labels)}). Keeping the first occurrence only."
            ),
            table=first.table,
        )

    return list(seen.values())


def _parse_relationship_ref(value: str) -> tuple[str, str] | None:
    return parse_tmdl_dotted_ref(value)


def _iter_relationship_blocks(lines: list[str]):
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("relationship ") or line.startswith("relationship\t"):
            name = unquote_tmdl_name(line.split(None, 1)[1]) if len(line.split(None, 1)) > 1 else ""
            block_lines = []
            i += 1
            while i < len(lines):
                inner = lines[i]
                if inner == "" or inner.strip() == "":
                    j = i + 1
                    while j < len(lines) and (lines[j] == "" or lines[j].strip() == ""):
                        j += 1
                    if j < len(lines) and lines[j].startswith("\t"):
                        block_lines.append(inner)
                        i = j
                        continue
                    break
                if inner.startswith("\t"):
                    block_lines.append(inner)
                    i += 1
                    continue
                break
            while i < len(lines) and (lines[i] == "" or lines[i].strip() == ""):
                i += 1
            yield name, block_lines
            continue
        i += 1


def parse_relationship_details(model_path: Path) -> list[RelationshipInfo]:
    rel_file = model_path / "definition" / "relationships.tmdl"
    if not rel_file.exists():
        return []

    lines = rel_file.read_text(encoding="utf-8").splitlines()
    relationships: list[RelationshipInfo] = []

    for name, block_lines in _iter_relationship_blocks(lines):
        props: dict[str, str] = {}
        for line in block_lines:
            stripped = line.strip()
            if ":" in stripped:
                key, value = stripped.split(":", 1)
                props[key.strip()] = value.strip()
            elif stripped:
                props[stripped] = "true"

        from_ref = _parse_relationship_ref(props.get("fromColumn", ""))
        to_ref = _parse_relationship_ref(props.get("toColumn", ""))
        if not from_ref or not to_ref:
            continue

        is_active_raw = unquote_tmdl_name(props.get("isActive", "true")).strip().lower()
        relationships.append(RelationshipInfo(
            name=name,
            from_table=from_ref[0],
            from_column=from_ref[1],
            to_table=to_ref[0],
            to_column=to_ref[1],
            from_cardinality=unquote_tmdl_name(props.get("fromCardinality", "")).strip().lower(),
            to_cardinality=unquote_tmdl_name(props.get("toCardinality", "")).strip().lower(),
            is_active=is_active_raw not in ("false", "0", "no"),
        ))

    return relationships


def parse_relationships(model_path: Path) -> set[tuple[str, str]]:
    refs = set()
    for rel in parse_relationship_details(model_path):
        refs.add((rel.from_table, rel.from_column))
        refs.add((rel.to_table, rel.to_column))
    return refs


def parse_rls_roles(model_path: Path) -> list[tuple[str, str, str]]:
    """Returns list of (role_name, table, column) for RLS filter references."""
    results = []
    seen = set()

    def add_expr_refs(role_name: str, expr: str, default_table: str | None = None) -> None:
        for cm in re.finditer(r"'([^']+)'\[([^\]]+)\]", expr):
            ref = (role_name, cm.group(1), cm.group(2))
            if ref not in seen:
                seen.add(ref)
                results.append(ref)
        for cm in re.finditer(r"(?<!')\b(\w+)\[([^\]]+)\]", expr):
            ref = (role_name, cm.group(1), cm.group(2))
            if ref not in seen:
                seen.add(ref)
                results.append(ref)
        if default_table:
            for cm in re.finditer(r"(?<![\w'])\[([^\]]+)\]", expr):
                ref = (role_name, default_table, cm.group(1))
                if ref not in seen:
                    seen.add(ref)
                    results.append(ref)

    definition_dir = model_path / "definition"

    files_to_check = []
    model_file = definition_dir / "model.tmdl"
    if model_file.exists():
        files_to_check.append(model_file)
    roles_dir = definition_dir / "roles"
    if roles_dir.exists():
        files_to_check.extend(roles_dir.glob("*.tmdl"))

    for f in files_to_check:
        lines = f.read_text(encoding="utf-8").splitlines()
        current_role = None
        i = 0

        while i < len(lines):
            line = lines[i]

            if line.startswith("role ") and not line.startswith("\t"):
                current_role = line[5:].strip().strip("'\"")
                i += 1
                continue

            if current_role and "filterExpression" in line:
                expr_lines = []
                if "=" in line:
                    expr_lines.append(line.split("=", 1)[1].strip())

                i += 1
                while i < len(lines):
                    inner = lines[i]
                    if inner.startswith("\t\t"):
                        cleaned = inner.strip().strip("`")
                        if cleaned:
                            expr_lines.append(cleaned)
                        i += 1
                        continue
                    break

                add_expr_refs(current_role, " ".join(expr_lines))
                continue

            stripped = line.strip()
            if current_role and stripped.startswith("tablePermission "):
                permission = stripped[len("tablePermission "):].strip()
                permission_table = ""
                expr_lines = []
                if "=" in permission:
                    permission_table, expr = permission.split("=", 1)
                    expr_lines.append(expr.strip())
                else:
                    permission_table = permission

                permission_table = permission_table.strip().strip("'\"")

                i += 1
                while i < len(lines):
                    inner = lines[i]
                    if inner.startswith("\t\t"):
                        cleaned = inner.strip().strip("`")
                        if cleaned:
                            expr_lines.append(cleaned)
                        i += 1
                        continue
                    break

                add_expr_refs(current_role, " ".join(expr_lines), permission_table or None)
                continue

            i += 1

    return results


# ── JSON Reference Scanning ──────────────────────────────────────────────────


def _split_top_level_query_ref(value: str) -> tuple[str, str] | None:
    depth_paren = 0
    depth_bracket = 0
    for idx, ch in enumerate(value):
        if ch == "(":
            depth_paren += 1
        elif ch == ")":
            depth_paren = max(0, depth_paren - 1)
        elif ch == "[":
            depth_bracket += 1
        elif ch == "]":
            depth_bracket = max(0, depth_bracket - 1)
        elif ch == "." and depth_paren == 0 and depth_bracket == 0:
            left = value[:idx].strip()
            right = value[idx + 1 :].strip()
            if left and right:
                return left, right
            return None
    return None


def _source_aliases_from_obj(obj) -> dict[str, str]:
    aliases: dict[str, str] = {}
    if not isinstance(obj, dict):
        return aliases
    sources = obj.get("From")
    if not isinstance(sources, list):
        return aliases
    for source in sources:
        if not isinstance(source, dict):
            continue
        name = source.get("Name")
        entity = source.get("Entity")
        if isinstance(name, str) and isinstance(entity, str) and name and entity:
            aliases[name] = entity
    return aliases


def _source_aliases_from_tree(obj) -> dict[str, str]:
    aliases = _source_aliases_from_obj(obj)
    if isinstance(obj, dict):
        for value in obj.values():
            aliases.update(_source_aliases_from_tree(value))
    elif isinstance(obj, list):
        for item in obj:
            aliases.update(_source_aliases_from_tree(item))
    return aliases


def _source_ref_entity(source_ref: dict, aliases: dict[str, str]) -> str | None:
    if not isinstance(source_ref, dict):
        return None
    entity = source_ref.get("Entity")
    if isinstance(entity, str) and entity:
        return entity
    source = source_ref.get("Source")
    if isinstance(source, str):
        return aliases.get(source)
    return None


def _find_json_refs(obj, path_parts: Optional[list] = None, aliases: Optional[dict[str, str]] = None) -> list[dict]:
    """Recursively find all Entity/Property measure/column references in JSON."""
    if path_parts is None:
        path_parts = []
    if aliases is None:
        aliases = _source_aliases_from_tree(obj)
    refs = []

    if isinstance(obj, dict):
        local_aliases = {**aliases, **_source_aliases_from_obj(obj)}

        # Measure or Column reference object
        for ref_type in ("Measure", "Column"):
            if ref_type in obj and isinstance(obj[ref_type], dict):
                inner = obj[ref_type]
                if "Property" in inner and "Expression" in inner:
                    expr = inner.get("Expression", {})
                    src = expr.get("SourceRef", {}) if isinstance(expr, dict) else {}
                    entity = _source_ref_entity(src, local_aliases) if isinstance(src, dict) else None
                    if entity:
                        refs.append({
                            "table": entity,
                            "name": inner["Property"],
                            "ref_type": ref_type,
                            "path": ".".join(path_parts + [ref_type]),
                        })

        # Aggregation wrapping a Column/Measure
        if "Aggregation" in obj and isinstance(obj["Aggregation"], dict):
            agg_expr = obj["Aggregation"].get("Expression", {})
            if isinstance(agg_expr, dict):
                for ref_type in ("Column", "Measure"):
                    if ref_type in agg_expr:
                        inner = agg_expr[ref_type]
                        if isinstance(inner, dict) and "Property" in inner:
                            src = inner.get("Expression", {}).get("SourceRef", {})
                            entity = _source_ref_entity(src, local_aliases) if isinstance(src, dict) else None
                            if entity:
                                refs.append({
                                    "table": entity,
                                    "name": inner["Property"],
                                    "ref_type": ref_type,
                                    "path": ".".join(path_parts + ["Aggregation"]),
                                })

        # Metadata selector: flat "Table.Name" string
        if "metadata" in obj and isinstance(obj["metadata"], str):
            meta = obj["metadata"]
            parts = _split_top_level_query_ref(meta)
            if parts:
                refs.append({
                    "table": parts[0],
                    "name": parts[1],
                    "ref_type": "metadata",
                    "path": ".".join(path_parts + ["metadata"]),
                    "raw_value": meta,
                })

        # HierarchyLevel reference object
        if "HierarchyLevel" in obj and isinstance(obj["HierarchyLevel"], dict):
            hl = obj["HierarchyLevel"]
            level_name = hl.get("Level", "")
            hier_expr = hl.get("Expression", {})
            if isinstance(hier_expr, dict) and "Hierarchy" in hier_expr:
                hier = hier_expr["Hierarchy"]
                if isinstance(hier, dict):
                    hier_name = hier.get("Hierarchy", "")
                    hier_src = hier.get("Expression", {}).get("SourceRef", {})
                    entity = _source_ref_entity(hier_src, local_aliases) if isinstance(hier_src, dict) else None
                    if entity:
                        refs.append({
                            "table": entity,
                            "name": hier_name,
                            "ref_type": "HierarchyLevel",
                            "level": level_name,
                            "path": ".".join(path_parts + ["HierarchyLevel"]),
                        })

        # Recurse
        for key, value in obj.items():
            if key in ("$schema",):
                continue
            refs.extend(_find_json_refs(value, path_parts + [key], local_aliases))

    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            refs.extend(_find_json_refs(item, path_parts + [f"[{i}]"], aliases))

    return refs


def _extract_query_refs(obj) -> set[str]:
    refs: set[str] = set()
    if isinstance(obj, dict):
        query_ref = obj.get("queryRef")
        if isinstance(query_ref, str) and query_ref:
            refs.add(query_ref)
        query_refs = obj.get("queryRefs")
        if isinstance(query_refs, list):
            refs |= {value for value in query_refs if isinstance(value, str) and value}
        for value in obj.values():
            refs |= _extract_query_refs(value)
    elif isinstance(obj, list):
        for item in obj:
            refs |= _extract_query_refs(item)
    return refs


def _extract_field_parameter_buckets(query_state: dict) -> set[str]:
    buckets: set[str] = set()
    if not isinstance(query_state, dict):
        return buckets
    for bucket, config in query_state.items():
        if isinstance(config, dict) and isinstance(config.get("fieldParameters"), list) and config.get("fieldParameters"):
            buckets.add(bucket)
    return buckets


def _is_selector_metadata_path(path: str) -> bool:
    return ".selector.metadata" in path or path.endswith("selector.metadata")


def _selector_entry_prefix(path: str) -> str:
    if _is_selector_metadata_path(path):
        return path.rsplit(".selector.metadata", 1)[0]
    return path


def _formatting_selector_entry_prefix(path: str) -> str:
    return path.split(".selector.", 1)[0] if ".selector." in path else ""


def _formatting_rule_entry_prefix(path: str) -> str:
    match = re.search(r"(.*\.objects\.[^.]+\.\[\d+\])\.properties\.", path)
    return match.group(1) if match else ""


def _determine_context(path: str) -> str:
    if "queryState" in path:
        m = re.search(r"queryState\.(\w+)", path)
        if m:
            role_map = {
                "Category": "Category",
                "Y": "Y-axis",
                "Y2": "Y2-axis",
                "Values": "Values",
                "Rows": "Rows",
                "Columns": "Columns",
                "Data": "Data",
                "Tooltips": "Tooltips",
                "Series": "Series",
            }
            return role_map.get(m.group(1), m.group(1))
    if "sortDefinition" in path:
        return "Sort"
    if "drillthrough" in path.lower():
        return "Drillthrough"
    if "filterConfig" in path:
        return "Filters pane (visual)"
    if "filter" in path.lower() and "filterConfig" not in path:
        return "Filter"
    if "Aggregation" in path:
        return "Aggregation"
    if "objects" in path or "metadata" in path:
        return "Formatting"
    return "Other"


def _filter_pane_refs(filter_obj: dict, path_parts: list[str]) -> list[dict]:
    active_refs = []
    for key in ("filter", "filterDefinition", "where", "conditions", "expression", "values"):
        value = filter_obj.get(key)
        if isinstance(value, (dict, list)):
            active_refs.extend(_find_json_refs(value, path_parts + [key]))
    if active_refs:
        return active_refs

    field = filter_obj.get("field")
    if isinstance(field, (dict, list)):
        refs = _find_json_refs(field, path_parts + ["field"])
        if refs:
            return refs
    return _find_json_refs(filter_obj, path_parts)


def _inactive_filter_field_prefixes(data: dict) -> set[str]:
    prefixes: set[str] = set()
    filter_configs = [
        ("filterConfig", data.get("filterConfig")),
        ("visual.filterConfig", data.get("visual", {}).get("filterConfig") if isinstance(data.get("visual"), dict) else None),
    ]
    for path_prefix, filter_config in filter_configs:
        filters = filter_config.get("filters", []) if isinstance(filter_config, dict) else []
        if not isinstance(filters, list):
            continue
        for idx, filter_obj in enumerate(filters):
            if not isinstance(filter_obj, dict) or "field" not in filter_obj:
                continue
            has_active_state = any(
                key in filter_obj
                for key in ("filter", "filterDefinition", "where", "conditions", "expression", "values")
            )
            if not has_active_state:
                prefixes.add(f"{path_prefix}.filters.[{idx}].field")
    return prefixes


def _get_visual_info(data: dict) -> tuple[str, str]:
    visual = data.get("visual", {})
    visual_type = visual.get("visualType", "unknown")
    title = ""

    # Try visualContainerObjects.title first, then objects.general
    for container_key in ("visualContainerObjects", "objects"):
        container = visual.get(container_key, {})
        for key in ("title", "general"):
            items = container.get(key, [])
            for obj in items:
                props = obj.get("properties", {})
                for prop_name in ("text", "title"):
                    title_prop = props.get(prop_name, {})
                    if not isinstance(title_prop, dict):
                        continue
                    text = title_prop.get("expr", {})
                    if isinstance(text, dict) and "Literal" in text:
                        val = text["Literal"].get("Value", "")
                        title = val.strip("'\"")
                        if title:
                            return visual_type, title

    return visual_type, title


def _get_page_name(page_dir: Path) -> str:
    page_json = page_dir / "page.json"
    if page_json.exists():
        try:
            data = json.loads(page_json.read_text(encoding="utf-8"))
            return data.get("displayName", page_dir.name)
        except (json.JSONDecodeError, KeyError):
            pass
    return page_dir.name


def _page_is_hidden(data: dict) -> bool:
    visibility = str(data.get("visibility", "")).casefold()
    return bool(data.get("isHidden") is True or data.get("hidden") is True or visibility == "hiddeninviewmode")


def _get_page_info(page_dir: Path) -> tuple[str, bool]:
    page_json = page_dir / "page.json"
    if page_json.exists():
        try:
            data = json.loads(page_json.read_text(encoding="utf-8"))
            return data.get("displayName", page_dir.name), _page_is_hidden(data)
        except (json.JSONDecodeError, KeyError):
            pass
    return page_dir.name, False


def _visual_position(data: dict) -> dict:
    position = data.get("position", {})
    return position if isinstance(position, dict) else {}


def _artifact_rel_path(report_path: Path, file_path: Path) -> str:
    try:
        return file_path.relative_to(report_path).as_posix()
    except ValueError:
        return file_path.name


def _invalid_json_artifact_kind(rel_path: str) -> str:
    if rel_path == "definition.pbir":
        return "Report Definition"
    if rel_path == "definition/reportExtensions.json":
        return "Report Extension"
    if rel_path.endswith("/visual.json"):
        return "Visual"
    if rel_path.endswith("/page.json"):
        return "Page"
    if rel_path.endswith(".bookmark.json"):
        return "Bookmark"
    if rel_path == "definition/report.json":
        return "Report"
    if rel_path == "report.json":
        return "PBIR-Legacy Report"
    return "Definition JSON"


def scan_invalid_report_json(report_path: Path) -> list[ReportIssue]:
    rpt_name = report_display_name(report_path)
    definition_dir = report_path / "definition"
    issues: list[ReportIssue] = []
    json_files = []
    if definition_dir.exists():
        json_files.extend(sorted(definition_dir.rglob("*.json")))
    for report_file in (report_path / "definition.pbir", report_path / "report.json"):
        if report_file.exists():
            json_files.append(report_file)

    for json_file in json_files:
        try:
            json.loads(json_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            rel_path = _artifact_rel_path(report_path, json_file)
            parts = rel_path.split("/")
            page = ""
            visual_id = ""
            if "pages" in parts:
                page_idx = parts.index("pages")
                if page_idx + 1 < len(parts):
                    page_dir = report_path / "definition" / "pages" / parts[page_idx + 1]
                    page = _get_page_name(page_dir)
                if "visuals" in parts:
                    visual_idx = parts.index("visuals")
                    if visual_idx + 1 < len(parts):
                        visual_id = parts[visual_idx + 1]
            issues.append(ReportIssue(
                severity="error",
                issue_type="invalid_report_json",
                report=rpt_name,
                page=page,
                visual_id=visual_id,
                artifact_kind=_invalid_json_artifact_kind(rel_path),
                artifact_path=rel_path,
                message=f"Could not parse PBIR JSON file: {exc}",
            ))
        except OSError as exc:
            rel_path = _artifact_rel_path(report_path, json_file)
            issues.append(ReportIssue(
                severity="warning",
                issue_type="invalid_report_json",
                report=rpt_name,
                artifact_kind=_invalid_json_artifact_kind(rel_path),
                artifact_path=rel_path,
                message=f"Could not read PBIR JSON file: {exc}",
            ))

    return issues


def _page_name_for_section(report_path: Path, section_id: str) -> str:
    page_dir = report_path / "definition" / "pages" / section_id
    if page_dir.exists():
        return _get_page_name(page_dir)
    return section_id


def _bookmark_visual_ref_parts(path: str) -> tuple[str, str] | None:
    m = re.search(r"explorationState\.sections\.([^.]+)\.visualContainers\.([^.]+)\.", path)
    if not m:
        return None
    return m.group(1), m.group(2)


def _bookmark_projection_entry_prefix(path: str) -> str:
    m = re.search(r"(.*\.singleVisual\.projections\.[^.]+\.\[\d+\])", path)
    return m.group(1) if m else path


def _report_issue_from_usage(
    usage: UsageRef,
    *,
    issue_type: str,
    severity: str,
    message: str,
    suggestions: list[dict] | None = None,
) -> ReportIssue:
    return ReportIssue(
        severity=severity,
        issue_type=issue_type,
        report=usage.report,
        page=usage.page,
        visual_type=usage.visual_type,
        visual_title=usage.visual_title,
        visual_id=usage.visual_id,
        context=usage.context,
        table=usage.table,
        name=usage.name,
        ref_type=usage.ref_type,
        artifact_kind=usage.artifact_kind,
        artifact_path=usage.artifact_path,
        source_path=usage.source_path,
        selector_value=usage.selector_value,
        stale_kind=usage.stale_kind,
        visual_hidden=usage.visual_hidden,
        page_hidden=usage.page_hidden,
        visual_x=usage.visual_x,
        visual_y=usage.visual_y,
        visual_width=usage.visual_width,
        visual_height=usage.visual_height,
        message=message,
        suggestions=suggestions or [],
    )


def _report_issue_dedupe_key(issue: ReportIssue) -> tuple:
    if issue.issue_type == "invalid_report_json":
        return (issue.issue_type, issue.report, issue.artifact_path)
    return (
        issue.issue_type,
        issue.report,
        issue.page,
        issue.visual_id,
        issue.table.casefold(),
        issue.name.casefold(),
        issue.ref_type.casefold(),
        issue.artifact_path,
        issue.source_path,
        issue.selector_value,
        issue.stale_kind,
        issue.message,
    )


def _dedupe_report_issues(issues: list[ReportIssue]) -> list[ReportIssue]:
    deduped: dict[tuple, ReportIssue] = {}
    for issue in issues:
        deduped.setdefault(_report_issue_dedupe_key(issue), issue)
    return list(deduped.values())


def _normalize_suggestion_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _suggestion_confidence(score: float) -> str:
    if score >= 0.92:
        return "High"
    if score >= 0.78:
        return "Medium"
    return "Low"


def _item_matches_ref_type(item: ModelItem, ref_type: str) -> bool:
    if ref_type == "Measure":
        return item.item_type == "Measure"
    if ref_type == "Column":
        return item.item_type in ("Column", "Calculated Column")
    if ref_type == "HierarchyLevel":
        return item.item_type in ("Column", "Calculated Column")
    return True


def _fuzzy_suggestions(usage: UsageRef, items: list[ModelItem]) -> list[dict]:
    target_name = _normalize_suggestion_text(usage.name)
    target_table = _normalize_suggestion_text(usage.table)
    if not target_name:
        return []

    candidates = []
    for item in items:
        if not _item_matches_ref_type(item, usage.ref_type):
            continue
        item_name = _normalize_suggestion_text(item.name)
        item_table = _normalize_suggestion_text(item.table)
        if not item_name:
            continue
        if (
            target_name[0] != item_name[0]
            and target_name not in item_name
            and item_name not in target_name
            and target_table != item_table
        ):
            continue

        name_score = SequenceMatcher(None, target_name, item_name).ratio()
        table_score = SequenceMatcher(None, target_table, item_table).ratio() if target_table and item_table else 0.0
        score = (name_score * 0.75) + (table_score * 0.25)
        if target_name == item_name:
            score = max(score, 0.93)
        if score < 0.64:
            continue

        candidates.append({
            "table": item.table,
            "name": item.name,
            "type": item.item_type,
            "score": round(score, 3),
            "confidence": _suggestion_confidence(score),
            "action": "suggest_replace",
        })

    candidates.sort(key=lambda s: (-s["score"], s["table"].casefold(), s["name"].casefold()))
    return candidates[:5]


def _classify_missing_usage(
    usage: UsageRef,
    *,
    model_table_names: set[str],
    report_entity_names: set[str],
    measure_keys: set[tuple[str, str]],
    column_keys: set[tuple[str, str]],
    hierarchy_keys: set[tuple[str, str]],
) -> str | None:
    table_key = usage.table.casefold()
    nkey = normalize_key(usage.table, usage.name)
    table_known = table_key in model_table_names or table_key in report_entity_names

    if not table_known:
        return "missing_table"
    if usage.ref_type == "Measure" and nkey not in measure_keys:
        return "missing_report_measure" if table_key in report_entity_names and table_key not in model_table_names else "missing_measure"
    if usage.ref_type == "Column" and nkey not in column_keys:
        return "missing_column"
    if usage.ref_type == "HierarchyLevel" and nkey not in hierarchy_keys:
        return "missing_column"
    if usage.ref_type == "metadata" and nkey not in measure_keys and nkey not in column_keys:
        return "missing_column"
    return None


def _missing_issue_message(issue_type: str, usage: UsageRef) -> str:
    ref_label = f"{usage.table}[{usage.name}]" if usage.table and usage.name else usage.name or usage.table
    if issue_type == "inactive_visual_filter_reference":
        return f"Visual filter pane contains an inactive all-values filter card for {ref_label}; remove the stale card or reconnect it to an existing field."
    if issue_type == "stale_formatting_rule":
        return f"Visual formatting rule uses old field {ref_label}; clean the stale formatting rule or recreate it from an existing field."
    if issue_type == "missing_table":
        return f"Report references table/entity '{usage.table}', but it was not found in the selected semantic model or report metadata."
    if issue_type == "missing_measure":
        return f"Report references measure {ref_label}, but that measure was not found."
    if issue_type == "missing_report_measure":
        return f"Report references report-level measure {ref_label}, but that report measure definition was not found."
    if issue_type == "missing_column":
        return f"Report references {ref_label}, but that field was not found."
    return f"Report reference {ref_label} could not be resolved."


def build_report_issues(
    *,
    direct_usages: list[UsageRef],
    stale_usages: list[UsageRef],
    invalid_json_issues: list[ReportIssue],
    all_items: list[ModelItem],
    model_items: list[ModelItem],
    hierarchies: list[HierarchyInfo],
    warnings: list[AnalyzerWarning],
) -> list[ReportIssue]:
    model_table_names = {item.table.casefold() for item in model_items}
    report_entity_names = {item.table.casefold() for item in all_items if item.source_kind == "report"}
    measure_keys = {normalize_key(*item.key) for item in all_items if item.item_type == "Measure"}
    column_keys = {normalize_key(*item.key) for item in all_items if item.item_type in ("Column", "Calculated Column")}
    hierarchy_keys = {normalize_key(info.table, info.name) for info in hierarchies}

    issues: list[ReportIssue] = list(invalid_json_issues)
    suggestion_cache: dict[tuple[str, str, str], list[dict]] = {}
    for usage in direct_usages:
        issue_type = _classify_missing_usage(
            usage,
            model_table_names=model_table_names,
            report_entity_names=report_entity_names,
            measure_keys=measure_keys,
            column_keys=column_keys,
            hierarchy_keys=hierarchy_keys,
        )
        if not issue_type:
            continue
        severity = "error"
        if usage.stale_kind == "inactive_visual_filter_reference":
            issue_type = "inactive_visual_filter_reference"
            severity = "warning"
        suggestion_key = (usage.table.casefold(), usage.name.casefold(), usage.ref_type.casefold())
        if suggestion_key not in suggestion_cache:
            suggestion_cache[suggestion_key] = _fuzzy_suggestions(usage, all_items)
        issues.append(_report_issue_from_usage(
            usage,
            issue_type=issue_type,
            severity=severity,
            message=_missing_issue_message(issue_type, usage),
            suggestions=suggestion_cache[suggestion_key] if severity == "error" else [],
        ))

    for usage in stale_usages:
        if usage.stale_kind == "inactive_visual_filter_reference":
            missing_type = _classify_missing_usage(
                usage,
                model_table_names=model_table_names,
                report_entity_names=report_entity_names,
                measure_keys=measure_keys,
                column_keys=column_keys,
                hierarchy_keys=hierarchy_keys,
            )
            if not missing_type:
                continue
            issues.append(_report_issue_from_usage(
                usage,
                issue_type="inactive_visual_filter_reference",
                severity="warning",
                message=_missing_issue_message("inactive_visual_filter_reference", usage),
                suggestions=[],
            ))
            continue

        if usage.stale_kind == "bookmark_projection_entry":
            issue_type = "stale_bookmark_projection"
        elif usage.stale_kind == "formatting_rule_reference":
            issue_type = "stale_formatting_rule"
        else:
            issue_type = "stale_visual_selector"
        label = f"{usage.table}[{usage.name}]" if usage.table and usage.name else usage.selector_value
        message = (
            _missing_issue_message("stale_formatting_rule", usage)
            if issue_type == "stale_formatting_rule"
            else f"Stale PBIR metadata still points to {label}, but it is not part of the live visual query."
        )
        issues.append(_report_issue_from_usage(
            usage,
            issue_type=issue_type,
            severity="warning",
            message=message,
            suggestions=[{"action": "clean_stale", "confidence": "High"}],
        ))

    report_usage_by_table: dict[str, list[UsageRef]] = defaultdict(list)
    for usage in direct_usages:
        report_usage_by_table[usage.table.casefold()].append(usage)
    for warning in warnings:
        if warning.code not in ("UNRESOLVED_NAMEOF_TARGET", "AMBIGUOUS_NAMEOF_TARGET") or not warning.table:
            continue
        for usage in report_usage_by_table.get(warning.table.casefold(), []):
            issues.append(_report_issue_from_usage(
                usage,
                issue_type="field_parameter_warning",
                severity="warning",
                message=warning.message,
                suggestions=[],
            ))

    return _dedupe_report_issues(issues)


# ── Report Scanning ───────────────────────────────────────────────────────────


def scan_report_visuals(
    report_path: Path,
    invalid_json_issues: list[ReportIssue] | None = None,
) -> tuple[list[UsageRef], list[UsageRef], dict]:
    """Returns (live_usages, stale_usages, visual_meta) where visual_meta maps visual_id -> info."""
    rpt_name = report_display_name(report_path)
    pages_dir = report_path / "definition" / "pages"
    usages = []
    stale_usages = []
    visual_meta = {}

    if not pages_dir.exists():
        return usages, stale_usages, visual_meta

    for page_dir in sorted(pages_dir.iterdir()):
        if not page_dir.is_dir():
            continue

        page_name, page_hidden = _get_page_info(page_dir)
        visuals_dir = page_dir / "visuals"
        if not visuals_dir.exists():
            continue

        for visual_dir in sorted(visuals_dir.iterdir()):
            if not visual_dir.is_dir():
                continue

            visual_json = visual_dir / "visual.json"
            if not visual_json.exists():
                continue

            try:
                data = json.loads(visual_json.read_text(encoding="utf-8"))
            except json.JSONDecodeError as exc:
                if invalid_json_issues is not None:
                    invalid_json_issues.append(ReportIssue(
                        severity="error",
                        issue_type="invalid_report_json",
                        report=rpt_name,
                        page=page_name,
                        visual_id=visual_dir.name,
                        artifact_kind="Visual",
                        artifact_path=_artifact_rel_path(report_path, visual_json),
                        message=f"Could not parse PBIR JSON file: {exc}",
                    ))
                continue

            visual_type, visual_title = _get_visual_info(data)
            visual_id = visual_dir.name
            visual_hidden = data.get("isHidden") is True or data.get("hidden") is True
            position = _visual_position(data)
            visual_refs_local = []
            live_query_refs = _extract_query_refs(data.get("visual", {}).get("query", {}).get("queryState", {}))
            field_parameter_buckets = _extract_field_parameter_buckets(data.get("visual", {}).get("query", {}).get("queryState", {}))
            inactive_filter_prefixes = _inactive_filter_field_prefixes(data)

            refs = _find_json_refs(data)
            formatting_rule_entries = {
                prefix
                for ref in refs
                if _determine_context(ref["path"]) == "Aggregation"
                and ref["path"].startswith("visual.objects.")
                for prefix in (_formatting_rule_entry_prefix(ref["path"]),)
                if prefix
            }
            stale_entry_selectors: dict[str, str] = {}
            for ref in refs:
                raw_value = ref.get("raw_value", "")
                selector_prefix = _selector_entry_prefix(ref["path"])
                if (
                    ref["ref_type"] == "metadata"
                    and _is_selector_metadata_path(ref["path"])
                    and raw_value
                    and raw_value not in live_query_refs
                    and selector_prefix not in formatting_rule_entries
                ):
                    stale_entry_selectors[selector_prefix] = raw_value
            for ref in refs:
                if ref["ref_type"] == "metadata" and _selector_entry_prefix(ref["path"]) in formatting_rule_entries:
                    continue
                context = _determine_context(ref["path"])
                entry_selector_value = next(
                    (selector_value for prefix, selector_value in stale_entry_selectors.items() if ref["path"].startswith(prefix)),
                    "",
                )
                formatting_selector_prefix = (
                    _formatting_selector_entry_prefix(ref["path"])
                    if context == "Formatting" and ".selector." in ref["path"] and not _is_selector_metadata_path(ref["path"])
                    else ""
                )
                inactive_filter_prefix = next(
                    (prefix for prefix in inactive_filter_prefixes if ref["path"].startswith(prefix)),
                    "",
                )
                is_formatting_rule_reference = context == "Aggregation" and ref["path"].startswith("visual.objects.")
                is_stale_selector = bool(entry_selector_value)
                is_stale_formatting_selector = bool(formatting_selector_prefix)
                is_inactive_filter_reference = bool(inactive_filter_prefix)
                u = UsageRef(
                    table=ref["table"],
                    name=ref["name"],
                    ref_type=ref["ref_type"],
                    report=rpt_name,
                    page=page_name,
                    visual_type=visual_type,
                    visual_title=visual_title,
                    visual_id=visual_id,
                    context=context,
                    source_path=ref["path"],
                    artifact_kind="Visual",
                    artifact_path=_artifact_rel_path(report_path, visual_json),
                    selector_value=entry_selector_value or format_item_ref((ref["table"], ref["name"])),
                    is_stale=is_stale_selector or is_stale_formatting_selector or is_inactive_filter_reference or is_formatting_rule_reference,
                    stale_kind=(
                        "visual_formatting_selector_entry" if is_stale_formatting_selector
                        else "inactive_visual_filter_reference" if inactive_filter_prefix
                        else "formatting_rule_reference" if is_formatting_rule_reference
                        else ""
                    ),
                    visual_hidden=visual_hidden,
                    page_hidden=page_hidden,
                    visual_x=position.get("x"),
                    visual_y=position.get("y"),
                    visual_width=position.get("width"),
                    visual_height=position.get("height"),
                )
                if is_stale_selector or is_stale_formatting_selector or is_inactive_filter_reference or is_formatting_rule_reference:
                    u.context = (
                        "Inactive Filter" if is_inactive_filter_reference
                        else "Stale Formatting Rule" if is_formatting_rule_reference
                        else "Stale Formatting"
                    )
                    stale_usages.append(u)
                else:
                    usages.append(u)
                    visual_refs_local.append((ref["table"], ref["name"]))

            visual_meta[visual_id] = {
                "type": visual_type,
                "title": visual_title,
                "page": page_name,
                "hidden": visual_hidden,
                "page_hidden": page_hidden,
                "x": position.get("x"),
                "y": position.get("y"),
                "width": position.get("width"),
                "height": position.get("height"),
                "refs": visual_refs_local,
                "field_parameter_buckets": field_parameter_buckets,
            }

    return usages, stale_usages, visual_meta


def scan_visual_interactions(
    report_path: Path,
    visual_meta: dict,
    invalid_json_issues: list[ReportIssue] | None = None,
) -> list[UsageRef]:
    rpt_name = report_display_name(report_path)
    pages_dir = report_path / "definition" / "pages"
    usages = []

    if not pages_dir.exists():
        return usages

    for page_dir in sorted(pages_dir.iterdir()):
        if not page_dir.is_dir():
            continue

        page_json = page_dir / "page.json"
        if not page_json.exists():
            continue

        try:
            data = json.loads(page_json.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            if invalid_json_issues is not None:
                invalid_json_issues.append(ReportIssue(
                    severity="error",
                    issue_type="invalid_report_json",
                    report=rpt_name,
                    page=page_dir.name,
                    artifact_kind="Page",
                    artifact_path=_artifact_rel_path(report_path, page_json),
                    message=f"Could not parse PBIR JSON file: {exc}",
                ))
            continue

        page_name = data.get("displayName", page_dir.name)
        page_hidden = _page_is_hidden(data)

        # ── Visual interactions (slicer targets) ──
        interactions = data.get("visualInteractions", [])

        for inter in interactions:
            if inter.get("type") != "DataFilter":
                continue

            source_id = inter.get("source", "")
            target_id = inter.get("target", "")
            source_meta = visual_meta.get(source_id, {})
            target_meta = visual_meta.get(target_id, {})

            if source_meta.get("type") == "slicer":
                target_label = target_meta.get("type", "unknown")
                if target_meta.get("title"):
                    target_label += f' "{target_meta["title"]}"'

                for table, name in source_meta.get("refs", []):
                    usages.append(UsageRef(
                        table=table,
                        name=name,
                        ref_type="Column",
                        report=rpt_name,
                        page=page_name,
                        visual_type=f"slicer \u2192 {target_label}",
                        visual_title=source_meta.get("title", ""),
                        visual_id=source_id,
                        context="Slicer Interaction",
                        source_path="visualInteractions",
                        artifact_kind="Page",
                        artifact_path=_artifact_rel_path(report_path, page_json),
                        visual_hidden=bool(source_meta.get("hidden", False)),
                        page_hidden=bool(source_meta.get("page_hidden", page_hidden)),
                        visual_x=source_meta.get("x"),
                        visual_y=source_meta.get("y"),
                        visual_width=source_meta.get("width"),
                        visual_height=source_meta.get("height"),
                    ))

        # ── Page-level filters ──
        page_filter_groups = (
            ("Page Filter", "Page Filter", data.get("filters", []), ["filters"]),
            (
                "Filters pane",
                "Filters pane (page)",
                data.get("filterConfig", {}).get("filters", []) if isinstance(data.get("filterConfig"), dict) else [],
                ["filterConfig", "filters"],
            ),
        )
        for visual_type, context, page_filters, path_prefix in page_filter_groups:
            if not isinstance(page_filters, list):
                continue
            for idx, filt in enumerate(page_filters):
                if not isinstance(filt, dict):
                    continue
                refs = _filter_pane_refs(filt, path_prefix + [f"[{idx}]"])
                for ref in refs:
                    usages.append(UsageRef(
                        table=ref["table"],
                        name=ref["name"],
                        ref_type=ref["ref_type"],
                        report=rpt_name,
                        page=page_name,
                        visual_type=visual_type,
                        visual_title="",
                        visual_id="",
                        context=context,
                        source_path=ref["path"],
                        artifact_kind="Page",
                        artifact_path=_artifact_rel_path(report_path, page_json),
                        page_hidden=page_hidden,
                    ))

        # ── Drillthrough target fields ──
        drillthrough = data.get("drillthrough", data.get("drillthroughFilters", []))
        if isinstance(drillthrough, list):
            for dt in drillthrough:
                if not isinstance(dt, dict):
                    continue
                refs = _find_json_refs(dt)
                for ref in refs:
                    usages.append(UsageRef(
                        table=ref["table"],
                        name=ref["name"],
                        ref_type=ref["ref_type"],
                        report=rpt_name,
                        page=page_name,
                        visual_type="Drillthrough",
                        visual_title="",
                        visual_id="",
                        context="Drillthrough",
                        source_path=ref["path"],
                        artifact_kind="Page",
                        artifact_path=_artifact_rel_path(report_path, page_json),
                        page_hidden=page_hidden,
                    ))
        elif isinstance(drillthrough, dict):
            refs = _find_json_refs(drillthrough)
            for ref in refs:
                usages.append(UsageRef(
                    table=ref["table"],
                    name=ref["name"],
                    ref_type=ref["ref_type"],
                    report=rpt_name,
                    page=page_name,
                    visual_type="Drillthrough",
                    visual_title="",
                    visual_id="",
                    context="Drillthrough",
                    source_path=ref["path"],
                    artifact_kind="Page",
                    artifact_path=_artifact_rel_path(report_path, page_json),
                    page_hidden=page_hidden,
                ))

    return usages


def scan_report_filters(
    report_path: Path,
    invalid_json_issues: list[ReportIssue] | None = None,
) -> list[UsageRef]:
    rpt_name = report_display_name(report_path)
    report_json = report_path / "definition" / "report.json"
    usages = []

    if not report_json.exists():
        return usages

    try:
        data = json.loads(report_json.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        if invalid_json_issues is not None:
            invalid_json_issues.append(ReportIssue(
                severity="error",
                issue_type="invalid_report_json",
                report=rpt_name,
                artifact_kind="Report",
                artifact_path=_artifact_rel_path(report_path, report_json),
                message=f"Could not parse PBIR JSON file: {exc}",
            ))
        return usages

    filter_config = data.get("filterConfig", {})
    report_filters = filter_config.get("filters", []) if isinstance(filter_config, dict) else []
    if not isinstance(report_filters, list):
        return usages

    for idx, filt in enumerate(report_filters):
        if not isinstance(filt, dict):
            continue
        refs = _filter_pane_refs(filt, ["filterConfig", "filters", f"[{idx}]"])
        for ref in refs:
            usages.append(UsageRef(
                table=ref["table"],
                name=ref["name"],
                ref_type=ref["ref_type"],
                report=rpt_name,
                page="",
                visual_type="Filters pane",
                visual_title="All pages",
                visual_id="",
                context="Filters pane (all pages)",
                source_path=ref["path"],
                artifact_kind="Report",
                artifact_path=_artifact_rel_path(report_path, report_json),
            ))

    return usages


def scan_bookmarks(
    report_path: Path,
    visual_meta: dict | None = None,
    invalid_json_issues: list[ReportIssue] | None = None,
) -> tuple[list[UsageRef], list[UsageRef]]:
    rpt_name = report_display_name(report_path)
    bookmarks_dir = report_path / "definition" / "bookmarks"
    usages = []
    stale_usages = []
    visual_meta = visual_meta or {}

    if not bookmarks_dir.exists():
        return usages, stale_usages

    for bm_file in sorted(bookmarks_dir.glob("*.bookmark.json")):
        try:
            data = json.loads(bm_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            if invalid_json_issues is not None:
                invalid_json_issues.append(ReportIssue(
                    severity="error",
                    issue_type="invalid_report_json",
                    report=rpt_name,
                    visual_type="Bookmark",
                    visual_title=bm_file.stem,
                    visual_id=bm_file.stem,
                    artifact_kind="Bookmark",
                    artifact_path=_artifact_rel_path(report_path, bm_file),
                    message=f"Could not parse PBIR JSON file: {exc}",
                ))
            continue

        bm_name = data.get("displayName", bm_file.stem)
        stale_entry_prefixes: dict[str, dict] = {}
        sections = data.get("explorationState", {}).get("sections", {})
        if isinstance(sections, dict):
            for section_id, section in sections.items():
                if not isinstance(section, dict):
                    continue
                visual_containers = section.get("visualContainers", {})
                if not isinstance(visual_containers, dict):
                    continue
                for visual_id, container in visual_containers.items():
                    if not isinstance(container, dict):
                        continue
                    visual_info = visual_meta.get(visual_id, {})
                    if not visual_info and invalid_json_issues is not None:
                        page_name = _page_name_for_section(report_path, section_id)
                        source_path = ".".join([
                            "explorationState",
                            "sections",
                            section_id,
                            "visualContainers",
                            visual_id,
                        ])
                        invalid_json_issues.append(ReportIssue(
                            severity="warning",
                            issue_type="orphan_bookmark_visual_state",
                            report=rpt_name,
                            page=page_name,
                            visual_type="Bookmark",
                            visual_title=bm_name,
                            visual_id=visual_id,
                            context=f"Bookmark: {bm_name}" if bm_name else "Bookmark",
                            artifact_kind="Bookmark",
                            artifact_path=_artifact_rel_path(report_path, bm_file),
                            source_path=source_path,
                            message=f"Bookmark '{bm_name}' stores visual state for '{visual_id}', but that visual was not found on the report page.",
                            suggestions=[],
                        ))
                    single_visual = container.get("singleVisual", {})
                    if not isinstance(single_visual, dict):
                        continue
                    bookmark_params = single_visual.get("parameters", {})
                    bookmark_proj = single_visual.get("projections", {})
                    field_parameter_buckets = set(visual_info.get("field_parameter_buckets", set()))
                    if isinstance(bookmark_params, dict):
                        field_parameter_buckets |= {bucket for bucket, values in bookmark_params.items() if isinstance(values, list) and values}
                    for bucket in field_parameter_buckets:
                        entries = bookmark_proj.get(bucket, [])
                        if not isinstance(entries, list):
                            continue
                        for idx, entry in enumerate(entries):
                            refs_in_entry = _find_json_refs(
                                entry,
                                [
                                    "explorationState",
                                    "sections",
                                    section_id,
                                    "visualContainers",
                                    visual_id,
                                    "singleVisual",
                                    "projections",
                                    bucket,
                                    f"[{idx}]",
                                ],
                            )
                            if not refs_in_entry:
                                continue
                            stale_entry_prefixes[
                                ".".join([
                                    "explorationState",
                                    "sections",
                                    section_id,
                                    "visualContainers",
                                    visual_id,
                                    "singleVisual",
                                    "projections",
                                    bucket,
                                    f"[{idx}]",
                                ])
                            ] = {
                                "bucket": bucket,
                                "page": visual_info.get("page") or _page_name_for_section(report_path, section_id),
                                "visual_type": visual_info.get("type") or single_visual.get("visualType", "Bookmark"),
                                "visual_title": visual_info.get("title", ""),
                                "visual_id": visual_id,
                                "hidden": bool(visual_info.get("hidden", False)),
                                "page_hidden": bool(visual_info.get("page_hidden", False)),
                                "x": visual_info.get("x"),
                                "y": visual_info.get("y"),
                                "width": visual_info.get("width"),
                                "height": visual_info.get("height"),
                            }
        refs = _find_json_refs(data)

        for ref in refs:
            ref_location = _bookmark_visual_ref_parts(ref["path"])
            if ref_location:
                section_id, visual_id = ref_location
                visual_info = visual_meta.get(visual_id, {})
                page_name = visual_info.get("page") or _page_name_for_section(report_path, section_id)
                visual_type = visual_info.get("type") or "Bookmark"
                visual_title = visual_info.get("title", "")
                visual_hidden = bool(visual_info.get("hidden", False))
                page_hidden = bool(visual_info.get("page_hidden", False))
                visual_x = visual_info.get("x")
                visual_y = visual_info.get("y")
                visual_width = visual_info.get("width")
                visual_height = visual_info.get("height")
            else:
                page_name = ""
                visual_type = "Bookmark"
                visual_title = bm_name
                visual_id = bm_file.stem
                visual_hidden = False
                page_hidden = False
                visual_x = None
                visual_y = None
                visual_width = None
                visual_height = None

            matched_stale = next(
                (info for prefix, info in stale_entry_prefixes.items() if ref["path"].startswith(prefix)),
                None,
            )
            if matched_stale:
                visual_hidden = bool(matched_stale.get("hidden", visual_hidden))
                page_hidden = bool(matched_stale.get("page_hidden", page_hidden))
                visual_x = matched_stale.get("x", visual_x)
                visual_y = matched_stale.get("y", visual_y)
                visual_width = matched_stale.get("width", visual_width)
                visual_height = matched_stale.get("height", visual_height)
            bookmark_context = f"Bookmark: {bm_name}" if bm_name else "Bookmark"
            usage_ref = UsageRef(
                table=ref["table"],
                name=ref["name"],
                ref_type=ref["ref_type"],
                report=rpt_name,
                page=page_name,
                visual_type=visual_type,
                visual_title=visual_title or bm_name,
                visual_id=visual_id,
                context=bookmark_context,
                source_path=ref["path"],
                artifact_kind="Bookmark",
                artifact_path=_artifact_rel_path(report_path, bm_file),
                is_stale=bool(matched_stale),
                stale_kind="bookmark_projection_entry" if matched_stale else "",
                selector_value=matched_stale["bucket"] if matched_stale else "",
                visual_hidden=visual_hidden,
                page_hidden=page_hidden,
                visual_x=visual_x,
                visual_y=visual_y,
                visual_width=visual_width,
                visual_height=visual_height,
            )
            if matched_stale:
                usage_ref.context = f"Stale Bookmark: {bm_name}" if bm_name else "Stale Bookmark"
                stale_usages.append(usage_ref)
            else:
                usages.append(usage_ref)

    return usages, stale_usages


def scan_additional_definition_json(
    report_path: Path,
    excluded_files: set[Path],
    invalid_json_issues: list[ReportIssue] | None = None,
) -> list[UsageRef]:
    """Fallback scan for model refs in report JSON not covered by visual/page/bookmark scanners."""
    rpt_name = report_display_name(report_path)
    definition_dir = report_path / "definition"
    usages = []

    if not definition_dir.exists():
        return usages

    for json_file in sorted(definition_dir.rglob("*.json")):
        if json_file in excluded_files:
            continue

        try:
            data = json.loads(json_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            if invalid_json_issues is not None:
                invalid_json_issues.append(ReportIssue(
                    severity="error",
                    issue_type="invalid_report_json",
                    report=rpt_name,
                    artifact_kind=_invalid_json_artifact_kind(_artifact_rel_path(report_path, json_file)),
                    artifact_path=_artifact_rel_path(report_path, json_file),
                    message=f"Could not parse PBIR JSON file: {exc}",
                ))
            continue

        refs = _find_json_refs(data)
        if not refs:
            continue

        for ref in refs:
            usages.append(UsageRef(
                table=ref["table"],
                name=ref["name"],
                ref_type=ref["ref_type"],
                report=rpt_name,
                page="",
                visual_type="Definition JSON",
                visual_title="",
                visual_id="",
                context="Definition Metadata",
                source_path=ref["path"],
                artifact_kind="Definition JSON",
                artifact_path=_artifact_rel_path(report_path, json_file),
            ))

    return usages


# ── DAX Dependency Analysis ──────────────────────────────────────────────────


def _split_dax_comments(dax_body: str) -> tuple[str, str]:
    """Return (active_code, comments) from a DAX expression."""
    active_parts = []
    comment_parts = []
    i = 0

    while i < len(dax_body):
        if dax_body.startswith("//", i):
            end = dax_body.find("\n", i)
            if end == -1:
                comment_parts.append(dax_body[i:])
                active_parts.append(" " * (len(dax_body) - i))
                break
            comment_parts.append(dax_body[i:end])
            active_parts.append(" " * (end - i))
            i = end
            continue
        if dax_body.startswith("/*", i):
            end = dax_body.find("*/", i + 2)
            if end == -1:
                comment_parts.append(dax_body[i:])
                active_parts.append(" " * (len(dax_body) - i))
                break
            end += 2
            comment_parts.append(dax_body[i:end])
            active_parts.append(" " * (end - i))
            i = end
            continue

        active_parts.append(dax_body[i])
        i += 1

    return "".join(active_parts), "\n".join(comment_parts)


def _extract_dax_qualified_refs_from_text(text: str) -> set[tuple[str, str]]:
    return {(table, name) for table, name, _, _ in _scan_dax_qualified_refs(text)}


def _extract_dax_unqualified_refs_from_text(text: str) -> set[str]:
    cleaned_chars = list(text)
    for _, _, start, end in _scan_dax_qualified_refs(text):
        for idx in range(start, end):
            cleaned_chars[idx] = " "
    cleaned = "".join(cleaned_chars)

    refs = set()
    i = 0
    while i < len(cleaned):
        bracketed = _read_dax_bracketed_name(cleaned, i)
        if not bracketed:
            i += 1
            continue

        name, end = bracketed
        if not name.startswith("@"):
            refs.add(name)
        i = end
    return refs


def _extract_dax_table_refs_from_text(text: str) -> set[str]:
    refs = set()
    i = 0
    while i < len(text):
        bracketed = _read_dax_bracketed_name(text, i)
        if bracketed:
            _, i = bracketed
            continue

        parsed = read_single_quoted_name(text, i)
        if not parsed:
            i += 1
            continue

        table_name, end = parsed
        while end < len(text) and text[end].isspace():
            end += 1
        if end < len(text) and text[end] == "[":
            i = end
            continue
        refs.add(table_name)
        i = end
    return refs


def _scan_dax_qualified_refs(text: str) -> list[tuple[str, str, int, int]]:
    refs: list[tuple[str, str, int, int]] = []
    i = 0
    while i < len(text):
        if text[i] == "'":
            parsed = read_single_quoted_name(text, i)
            if not parsed:
                i += 1
                continue

            table_name, table_end = parsed
            bracket_start = _skip_dax_whitespace(text, table_end)
            bracketed = _read_dax_bracketed_name(text, bracket_start)
            if bracketed:
                object_name, object_end = bracketed
                refs.append((table_name, object_name, i, object_end))
                i = object_end
                continue
            i = table_end
            continue

        identifier = _read_dax_unquoted_table_name(text, i)
        if identifier:
            table_name, table_end = identifier
            bracket_start = _skip_dax_whitespace(text, table_end)
            bracketed = _read_dax_bracketed_name(text, bracket_start)
            if bracketed:
                object_name, object_end = bracketed
                refs.append((table_name, object_name, i, object_end))
                i = object_end
                continue
            i = table_end
            continue

        i += 1
    return refs


def _read_dax_unquoted_table_name(text: str, start: int) -> tuple[str, int] | None:
    if start > 0 and (text[start - 1].isalnum() or text[start - 1] in ("_", "'")):
        return None
    match = re.match(r"[A-Za-z_]\w*", text[start:])
    if not match:
        return None
    return match.group(0), start + len(match.group(0))


def _read_dax_bracketed_name(text: str, start: int) -> tuple[str, int] | None:
    if start >= len(text) or text[start] != "[":
        return None

    chars: list[str] = []
    i = start + 1
    while i < len(text):
        ch = text[i]
        if ch == "]":
            if i + 1 < len(text) and text[i + 1] == "]":
                chars.append("]")
                i += 2
                continue
            return "".join(chars), i + 1
        chars.append(ch)
        i += 1
    return None


def _skip_dax_whitespace(text: str, start: int) -> int:
    i = start
    while i < len(text) and text[i].isspace():
        i += 1
    return i


def _extract_dax_qualified_refs(dax_body: str) -> set[tuple[str, str]]:
    """Extract qualified refs: 'Table'[Name] and Table[Name]."""
    active_dax, _ = _split_dax_comments(dax_body)
    return _extract_dax_qualified_refs_from_text(active_dax)


def _extract_dax_unqualified_refs(dax_body: str) -> set[str]:
    """Extract unqualified [Name] references."""
    active_dax, _ = _split_dax_comments(dax_body)
    return _extract_dax_unqualified_refs_from_text(active_dax)


def _extract_dax_table_refs(dax_body: str) -> set[str]:
    """Extract bare table refs: e.g. COUNTROWS('Table')."""
    active_dax, _ = _split_dax_comments(dax_body)
    return _extract_dax_table_refs_from_text(active_dax)


def extract_dax_commented_refs(dax_body: str) -> set[str]:
    """Extract item-like refs found only in DAX comments for diagnostics."""
    _, comments = _split_dax_comments(dax_body)
    refs = {f"{table}[{name}]" for table, name in _extract_dax_qualified_refs_from_text(comments)}
    refs |= {f"[{name}]" for name in _extract_dax_unqualified_refs_from_text(comments)}
    refs |= {f"{table}" for table in _extract_dax_table_refs_from_text(comments)}
    return refs


def find_broken_dax_references(items: list[ModelItem]) -> dict[tuple[str, str], list[dict[str, str]]]:
    """Return item_key -> structured unresolved DAX refs for measures/calc columns."""
    existing_keys = {normalize_key(*item.key) for item in items}
    same_table_column_keys = {
        normalize_key(item.table, item.name)
        for item in items
        if item.item_type in ("Column", "Calculated Column")
    }
    measure_names = {item.name.casefold() for item in items if item.item_type == "Measure"}
    table_names = {item.table.casefold() for item in items}
    broken_refs: dict[tuple[str, str], list[dict[str, str]]] = {}

    for item in items:
        if item.item_type not in ("Measure", "Calculated Column") or not item.dax_body:
            continue

        missing: dict[tuple[str, str], dict[str, str]] = {}
        for tbl, name in item.explicit_measure_refs:
            if normalize_key(tbl, name) not in existing_keys:
                ref_text = format_item_ref((tbl, name))
                missing[("explicit-measure", ref_text.casefold())] = {
                    "kind": "measure",
                    "ref": ref_text,
                    "message": f"Missing measure: {ref_text}",
                }

        for tbl, name in _extract_dax_qualified_refs(item.dax_body):
            if normalize_key(tbl, name) not in existing_keys:
                ref_text = format_item_ref((tbl, name))
                table_exists = tbl.casefold() in table_names
                missing[("column", ref_text.casefold())] = {
                    "kind": "column",
                    "ref": ref_text,
                    "message": (
                        f"Missing column: {ref_text}" if table_exists
                        else f"Missing column (table not found): {ref_text}"
                    ),
                }

        for name in _extract_dax_unqualified_refs(item.dax_body):
            if item.item_type == "Calculated Column":
                same_table_key = normalize_key(item.table, name)
                if same_table_key not in same_table_column_keys and name.casefold() not in measure_names:
                    ref_text = f"[{name}]"
                    missing[("unqualified", ref_text.casefold())] = {
                        "kind": "unqualified",
                        "ref": ref_text,
                        "message": f"Missing same-table column or measure: {ref_text}",
                    }
            elif name.casefold() not in measure_names:
                ref_text = f"[{name}]"
                missing[("measure", ref_text.casefold())] = {
                    "kind": "measure",
                    "ref": ref_text,
                    "message": f"Missing measure: {ref_text}",
                }

        for table_name in _extract_dax_table_refs(item.dax_body):
            if table_name.casefold() not in table_names:
                missing[("table", table_name.casefold())] = {
                    "kind": "table",
                    "ref": table_name,
                    "message": f"Missing table: {table_name}",
                }

        if missing:
            broken_refs[item.key] = sorted(missing.values(), key=lambda x: x["ref"].casefold())

    return broken_refs


def build_dax_dependency_graph(items: list[ModelItem]) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Build item_key -> {referenced measure keys} graph."""
    measure_keys = {item.key for item in items if item.item_type == "Measure"}
    measure_name_index: dict[str, set[tuple[str, str]]] = defaultdict(set)
    measure_key_index = {normalize_key(*k): k for k in measure_keys}
    for key in measure_keys:
        measure_name_index[key[1].casefold()].add(key)

    deps: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for item in items:
        if not item.dax_body or item.item_type not in ("Measure", "Calculated Column"):
            continue

        refs: set[tuple[str, str]] = set()
        for tbl, name in item.explicit_measure_refs:
            key = measure_key_index.get(normalize_key(tbl, name))
            if key and key != item.key:
                refs.add(key)

        for tbl, name in _extract_dax_qualified_refs(item.dax_body):
            key = measure_key_index.get(normalize_key(tbl, name))
            if key and key != item.key:
                refs.add(key)

        for name in _extract_dax_unqualified_refs(item.dax_body):
            refs |= {key for key in measure_name_index.get(name.casefold(), set()) if key != item.key}

        deps[item.key] = refs

    return deps


def is_used_status(status: str) -> bool:
    return status.startswith("USED") or status.startswith("INDIRECT")


def build_dax_column_deps(items: list[ModelItem]) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Build item_key -> {(table, column)} graph for column refs in DAX."""
    column_keys = {
        item.key for item in items
        if item.item_type in ("Column", "Calculated Column")
    }
    column_key_index = {normalize_key(*k): k for k in column_keys}

    deps: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for item in items:
        if not item.dax_body or item.item_type not in ("Measure", "Calculated Column"):
            continue

        refs: set[tuple[str, str]] = set()
        for tbl, name in _extract_dax_qualified_refs(item.dax_body):
            key = column_key_index.get(normalize_key(tbl, name))
            if key and key != item.key:
                refs.add(key)

        # In calculated columns, [Name] often refers to another same-table column.
        if item.item_type == "Calculated Column":
            for name in _extract_dax_unqualified_refs(item.dax_body):
                key = column_key_index.get(normalize_key(item.table, name))
                if key and key != item.key:
                    refs.add(key)

        deps[item.key] = refs

    return deps


def build_dax_table_deps(items: list[ModelItem]) -> dict[tuple[str, str], set[str]]:
    """Build item_key -> {referenced table names via bare table refs in DAX} graph."""
    table_name_index = {item.table.casefold(): item.table for item in items}

    deps: dict[tuple[str, str], set[str]] = {}
    for item in items:
        if not item.dax_body or item.item_type not in ("Measure", "Calculated Column"):
            continue

        refs: set[str] = set()
        for table_name in _extract_dax_table_refs(item.dax_body):
            resolved = table_name_index.get(table_name.casefold())
            if resolved:
                refs.add(resolved)

        deps[item.key] = refs

    return deps


def resolve_indirect_measures(
    all_items: list[ModelItem],
    directly_used_measures: set[tuple[str, str]],
    dax_deps: dict[tuple[str, str], set[tuple[str, str]]],
) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Transitive closure: find all measures needed by directly-used ones.
    Returns {indirect_measure_key: {measures that depend on it}}."""
    # Forward pass: find ALL measures reachable from directly-used
    all_needed = set(directly_used_measures)
    changed = True
    while changed:
        changed = False
        for measure_key in list(all_needed):
            if measure_key in dax_deps:
                for dep in dax_deps[measure_key]:
                    if dep not in all_needed:
                        all_needed.add(dep)
                        changed = True

    # Build "via" mapping for indirect ones
    indirectly_used = {}
    for item in all_items:
        if item.item_type == "Measure" and item.key in all_needed and item.key not in directly_used_measures:
            via = set()
            for dep_key, refs in dax_deps.items():
                if item.key in refs and dep_key in all_needed:
                    via.add(dep_key)
            indirectly_used[item.key] = via

    return indirectly_used


def resolve_indirect_columns(
    all_items: list[ModelItem],
    directly_used_keys: set[tuple[str, str]],
    relationship_keys: set[tuple[str, str]],
    rls_keys: set[tuple[str, str]],
    all_needed_measures: set[tuple[str, str]],
    dax_col_deps: dict[tuple[str, str], set[tuple[str, str]]],
) -> dict[tuple[str, str], set[str]]:
    """Find columns only used via DAX of needed measures.
    Returns {(table, col): {measure names that reference it}}."""
    already_used = directly_used_keys | relationship_keys | rls_keys
    indirect_cols = {}

    for measure_key in all_needed_measures:
        for col_key in dax_col_deps.get(measure_key, set()):
            col_nkey = normalize_key(*col_key)
            if col_nkey not in already_used:
                if col_nkey not in indirect_cols:
                    indirect_cols[col_nkey] = set()
                indirect_cols[col_nkey].add(format_item_ref(measure_key))

    return indirect_cols


# ── Main Analysis ─────────────────────────────────────────────────────────────


def _normalize_cardinality(value: str) -> str:
    raw = (value or "").strip().lower()
    if raw in ("one", "1"):
        return "one"
    if raw in ("many", "*"):
        return "many"
    return ""


def _classify_table_role(table: str, relationships: list[RelationshipInfo]) -> tuple[str, str]:
    active_relationships = [rel for rel in relationships if rel.is_active]
    if not active_relationships:
        if relationships:
            return "isolated", "Only inactive relationships were found for this table."
        return "isolated", "No relationships were found for this table."

    sides = []
    unknown_count = 0
    for rel in active_relationships:
        if rel.from_table.casefold() == table.casefold():
            own_cardinality = _normalize_cardinality(rel.from_cardinality)
            other_cardinality = _normalize_cardinality(rel.to_cardinality)
        else:
            own_cardinality = _normalize_cardinality(rel.to_cardinality)
            other_cardinality = _normalize_cardinality(rel.from_cardinality)

        if own_cardinality == "one" and other_cardinality == "many":
            sides.append("one")
        elif own_cardinality == "many" and other_cardinality == "one":
            sides.append("many")
        else:
            unknown_count += 1

    if sides and len(set(sides)) == 1 and unknown_count == 0:
        if sides[0] == "one":
            return "dimension-like", "All active relationships put this table on the one side of a 1:* relationship."
        return "fact-like", "All active relationships put this table on the many side of a 1:* relationship."

    if sides or unknown_count:
        if "one" in sides and "many" in sides:
            return "mixed", "This table appears on both the one and many sides of active relationships."
        return "unknown", "Active relationships exist, but their role pattern is not a clean 1:* shape."

    return "isolated", "No active relationships were found for this table."


def build_table_summaries(
    results: list[dict],
    all_usages: list[UsageRef],
    relationship_details: list[RelationshipInfo],
    dax_column_deps: dict[tuple[str, str], set[tuple[str, str]]],
    dax_table_deps: dict[tuple[str, str], set[str]],
    field_parameter_issues: dict[str, list[str]] | None = None,
) -> list[dict]:
    rows_by_table: dict[str, list[dict]] = defaultdict(list)
    usage_by_table: dict[str, list[UsageRef]] = defaultdict(list)
    relationships_by_table: dict[str, list[RelationshipInfo]] = defaultdict(list)

    for row in results:
        rows_by_table[row["item"].table].append(row)

    for usage in all_usages:
        usage_by_table[usage.table].append(usage)

    for rel in relationship_details:
        relationships_by_table[rel.from_table].append(rel)
        if rel.to_table.casefold() != rel.from_table.casefold():
            relationships_by_table[rel.to_table].append(rel)

    tables = sorted(rows_by_table.keys(), key=str.casefold)
    summaries = []

    field_parameter_issues = field_parameter_issues or {}

    for table in tables:
        rows = rows_by_table[table]
        usages = usage_by_table.get(table, [])
        relationships = relationships_by_table.get(table, [])
        reports = sorted({u.report for u in usages if u.report}, key=str.casefold)
        pages = sorted({u.page for u in usages if u.page}, key=str.casefold)
        related_tables = sorted({
            rel.to_table if rel.from_table.casefold() == table.casefold() else rel.from_table
            for rel in relationships
        }, key=str.casefold)

        relationship_only_columns = []
        single_column_measures = []
        items_in_table = []
        hidden_count = 0
        used_count = 0
        direct_report_measure_count = 0
        direct_report_column_count = 0
        measure_count = 0
        column_count = 0
        calc_column_count = 0

        for row in sorted(rows, key=lambda r: (r["item"].item_type, r["item"].name.casefold())):
            item = row["item"]
            status = row["status"]
            has_direct_usage = bool(row.get("has_direct_usage"))
            if item.is_hidden:
                hidden_count += 1
            if is_used_status(status):
                used_count += 1
            if item.item_type == "Measure":
                measure_count += 1
                if has_direct_usage:
                    direct_report_measure_count += 1
                column_deps = sorted(dax_column_deps.get(item.key, set()))
                if len(column_deps) == 1:
                    dep_table, dep_column = column_deps[0]
                    single_column_measures.append({
                        "measure": format_item_ref(item.key),
                        "column": format_item_ref((dep_table, dep_column)),
                    })
            elif item.item_type == "Calculated Column":
                calc_column_count += 1
                column_count += 1
                if has_direct_usage:
                    direct_report_column_count += 1
            else:
                column_count += 1
                if has_direct_usage:
                    direct_report_column_count += 1

            if item.item_type in ("Column", "Calculated Column") and status == "USED (Relationship)":
                relationship_only_columns.append(format_item_ref(item.key))

            items_in_table.append({
                "name": item.name,
                "ref": format_item_ref(item.key),
                "type": item.item_type,
                "status": status,
                "removal_risk": row.get("removal_risk", "") or None,
                "review_triggers": row.get("review_triggers", []),
                "broken_dax_refs": row.get("broken_dax_refs", []),
                "broken_dax_ref_details": row.get("broken_dax_ref_details", []),
                "usage_count": len(row["usages"]),
            })

        active_relationships = sum(1 for rel in relationships if rel.is_active)
        inactive_relationships = len(relationships) - active_relationships
        one_to_many = 0
        many_to_one = 0
        one_to_one = 0
        many_to_many = 0
        relationship_items = []

        for rel in relationships:
            from_cardinality = _normalize_cardinality(rel.from_cardinality)
            to_cardinality = _normalize_cardinality(rel.to_cardinality)
            if from_cardinality == "one" and to_cardinality == "many":
                one_to_many += 1
            elif from_cardinality == "many" and to_cardinality == "one":
                many_to_one += 1
            elif from_cardinality == "one" and to_cardinality == "one":
                one_to_one += 1
            elif from_cardinality == "many" and to_cardinality == "many":
                many_to_many += 1

            if rel.from_table.casefold() == table.casefold():
                local_column = rel.from_column
                other_table = rel.to_table
                other_column = rel.to_column
                local_cardinality = from_cardinality
                other_cardinality = to_cardinality
            else:
                local_column = rel.to_column
                other_table = rel.from_table
                other_column = rel.from_column
                local_cardinality = to_cardinality
                other_cardinality = from_cardinality

            if local_cardinality == "one" and other_cardinality == "many":
                role = "one-side"
            elif local_cardinality == "many" and other_cardinality == "one":
                role = "many-side"
            else:
                role = "unknown"

            cardinality_label = ""
            if from_cardinality and to_cardinality:
                cardinality_label = ("1" if from_cardinality == "one" else "*") + ":" + ("1" if to_cardinality == "one" else "*")

            relationship_items.append({
                "name": rel.name,
                "local_column": format_item_ref((table, local_column)),
                "other_table": other_table,
                "other_column": format_item_ref((other_table, other_column)),
                "cardinality": cardinality_label,
                "is_active": rel.is_active,
                "role": role,
            })

        role_label, role_reason = _classify_table_role(table, relationships)
        signals = []
        if role_label == "dimension-like":
            signals.append("Looks like a dimension table based on active 1:* relationships.")
        elif role_label == "fact-like":
            signals.append("Looks like a fact table based on active 1:* relationships.")
        elif role_label == "mixed":
            signals.append("Participates on both sides of active relationships.")

        if relationship_only_columns:
            count = len(relationship_only_columns)
            noun = "column" if count == 1 else "columns"
            signals.append(f"{count} {noun} are only used for relationships.")

        table_fp_issues = field_parameter_issues.get(table, [])
        if table_fp_issues:
            count = len(table_fp_issues)
            noun = "target" if count == 1 else "targets"
            signals.insert(0, f"Broken field parameter: {count} unresolved NAMEOF {noun}.")

        external_dax_dependents = sorted({
            format_item_ref(row["item"].key)
            for row in results
            if row["item"].table.casefold() != table.casefold()
            and table in dax_table_deps.get(row["item"].key, set())
        }, key=str.casefold)
        if external_dax_dependents:
            count = len(external_dax_dependents)
            noun = "item" if count == 1 else "items"
            verb = "depends" if count == 1 else "depend"
            signals.append(f"{count} DAX {noun} outside this table {verb} on it.")

        if single_column_measures:
            count = len(single_column_measures)
            noun = "measure" if count == 1 else "measures"
            signals.append(f"{count} {noun} depend on exactly one column.")

        if not reports:
            signals.append("No direct report references were found for this table.")

        summaries.append({
            "name": table,
            "role_label": role_label,
            "role_reason": role_reason,
            "item_count": len(rows),
            "measure_count": measure_count,
            "column_count": column_count,
            "calculated_column_count": calc_column_count,
            "direct_report_measure_count": direct_report_measure_count,
            "direct_report_column_count": direct_report_column_count,
            "used_item_count": used_count,
            "unused_item_count": len(rows) - used_count,
            "hidden_item_count": hidden_count,
            "usage_ref_count": sum(len(row["usages"]) for row in rows),
            "report_count": len(reports),
            "reports": reports,
            "page_count": len(pages),
            "pages": pages,
            "relationship_count": len(relationships),
            "active_relationship_count": active_relationships,
            "inactive_relationship_count": inactive_relationships,
            "one_to_many_count": one_to_many,
            "many_to_one_count": many_to_one,
            "one_to_one_count": one_to_one,
            "many_to_many_count": many_to_many,
            "related_tables": related_tables,
            "relationship_only_columns": relationship_only_columns,
            "single_column_measures": single_column_measures,
            "external_dax_dependents": external_dax_dependents,
            "relationships": relationship_items,
            "signals": signals,
            "field_parameter_issues": table_fp_issues,
            "items": items_in_table,
        })

    return summaries


def analyze(
    workspace: Path,
    model_filters: list[str] | None = None,
    report_filters: list[str] | None = None,
    model_paths: list[Path] | None = None,
    report_paths: list[Path] | None = None,
    model_search_roots: list[Path] | None = None,
    report_search_roots: list[Path] | None = None,
) -> dict:
    if model_paths is not None:
        models = _unique_sorted_paths(model_paths)
    else:
        model_roots = model_search_roots or [workspace]
        models = discover_models(model_roots)
        models = filter_models(models, model_filters)

    if report_paths is not None:
        reports = _unique_sorted_paths(report_paths)
    else:
        report_roots = report_search_roots or [workspace]
        reports = discover_reports(report_roots)
        reports = filter_reports(reports, report_filters)

    if not models:
        roots = model_search_roots or [workspace]
        roots_display = ", ".join(str(p) for p in roots)
        print(f"Error: No matching *.SemanticModel found under: {roots_display}", file=sys.stderr)
        sys.exit(1)
    if not reports:
        roots = report_search_roots or [workspace]
        roots_display = ", ".join(str(p) for p in roots)
        print(f"Error: No matching *.Report found under: {roots_display}", file=sys.stderr)
        sys.exit(1)

    # ── Parse model ──
    all_items = []
    all_relationship_cols = set()
    all_relationship_details = []
    all_rls_refs = []
    all_hierarchies = []
    all_field_parameters = []
    all_unsupported_metadata_refs = []
    parsed_model_items = []
    warnings: list[AnalyzerWarning] = []

    for model_path in models:
        model_items = parse_model_items(model_path)
        parsed_model_items.extend(model_items)
        all_items.extend(model_items)
        relationship_details = parse_relationship_details(model_path)
        all_relationship_details.extend(relationship_details)
        all_relationship_cols |= {
            (rel.from_table, rel.from_column) for rel in relationship_details
        } | {
            (rel.to_table, rel.to_column) for rel in relationship_details
        }
        all_rls_refs.extend(parse_rls_roles(model_path))
        all_hierarchies.extend(parse_hierarchies(model_path))
        all_unsupported_metadata_refs.extend(parse_unsupported_metadata_refs(model_path))
        all_field_parameters.extend(
            resolve_field_parameter_targets(
                parse_field_parameters(model_path, warnings),
                model_items,
                model_path.name,
                warnings,
            )
        )

    for report_path in reports:
        all_items.extend(parse_report_extension_measures(report_path, warnings))

    all_items = _dedupe_items_with_warnings(all_items, warnings)

    # ── Scan reports ──
    all_visual_usages = []
    all_stale_visual_usages = []
    all_interaction_usages = []
    all_report_filter_usages = []
    all_bookmark_usages = []
    all_stale_bookmark_usages = []
    all_definition_meta_usages = []
    all_report_issues = []
    all_invalid_json_issues = []

    for report_path in reports:
        all_report_issues.extend(scan_invalid_report_json(report_path))
        all_report_issues.extend(scan_report_extension_issues(report_path))
        visual_usages, stale_visual_usages, visual_meta = scan_report_visuals(report_path, all_invalid_json_issues)
        all_visual_usages.extend(visual_usages)
        all_stale_visual_usages.extend(stale_visual_usages)
        all_interaction_usages.extend(scan_visual_interactions(report_path, visual_meta, all_invalid_json_issues))
        all_report_filter_usages.extend(scan_report_filters(report_path, all_invalid_json_issues))
        bookmark_usages, stale_bookmark_usages = scan_bookmarks(report_path, visual_meta, all_invalid_json_issues)
        all_bookmark_usages.extend(bookmark_usages)
        all_stale_bookmark_usages.extend(stale_bookmark_usages)

        excluded_files = set(report_path.glob("definition/pages/*/visuals/*/visual.json"))
        excluded_files |= set(report_path.glob("definition/pages/*/page.json"))
        excluded_files.add(report_path / "definition" / "report.json")
        excluded_files |= set(report_path.glob("definition/bookmarks/*.bookmark.json"))
        all_definition_meta_usages.extend(
            scan_additional_definition_json(report_path, excluded_files, all_invalid_json_issues)
        )

    direct_usages = (
        all_visual_usages
        + all_interaction_usages
        + all_report_filter_usages
        + all_bookmark_usages
        + all_definition_meta_usages
    )

    # ── Build indices ──
    direct_usage_index: dict[tuple[str, str], list[UsageRef]] = defaultdict(list)
    for u in direct_usages:
        direct_usage_index[normalize_key(u.table, u.name)].append(u)

    stale_usage_index: dict[tuple[str, str], list[UsageRef]] = defaultdict(list)
    for u in all_stale_visual_usages + all_stale_bookmark_usages:
        stale_usage_index[normalize_key(u.table, u.name)].append(u)

    synthetic_field_parameter_usages, field_parameter_targets = promote_field_parameter_usages(
        direct_usages,
        all_field_parameters,
    )

    all_usages = direct_usages + synthetic_field_parameter_usages

    report_issues = build_report_issues(
        direct_usages=direct_usages,
        stale_usages=all_stale_visual_usages + all_stale_bookmark_usages,
        invalid_json_issues=all_report_issues + all_invalid_json_issues,
        all_items=all_items,
        model_items=parsed_model_items,
        hierarchies=all_hierarchies,
        warnings=warnings,
    )

    usage_index: dict[tuple[str, str], list[UsageRef]] = defaultdict(list)
    for u in all_usages:
        usage_index[normalize_key(u.table, u.name)].append(u)

    relationship_keys = {normalize_key(t, c) for t, c in all_relationship_cols}
    rls_keys = {normalize_key(r[1], r[2]) for r in all_rls_refs}
    rls_role_map: dict[tuple[str, str], list[str]] = defaultdict(list)
    for role, tbl, col in all_rls_refs:
        rls_role_map[normalize_key(tbl, col)].append(role)

    # ── DAX dependencies ──
    dax_deps = build_dax_dependency_graph(all_items)
    dax_col_deps = build_dax_column_deps(all_items)
    dax_table_deps = build_dax_table_deps(all_items)
    broken_dax_refs = find_broken_dax_references(all_items)

    directly_used_measures = set()
    directly_used_keys = set()
    for item in all_items:
        nkey = normalize_key(*item.key)
        if nkey in usage_index:
            directly_used_keys.add(nkey)
            if item.item_type == "Measure":
                directly_used_measures.add(item.key)

    # Also count relationship/RLS as directly used for transitive analysis
    for item in all_items:
        nkey = normalize_key(*item.key)
        if nkey in relationship_keys or nkey in rls_keys:
            directly_used_keys.add(nkey)

    indirect_measures = resolve_indirect_measures(all_items, directly_used_measures, dax_deps)
    all_needed_measures = directly_used_measures | set(indirect_measures.keys())

    indirect_columns = resolve_indirect_columns(
        all_items, directly_used_keys, relationship_keys, rls_keys,
        all_needed_measures, dax_col_deps,
    )

    # ── Hierarchy index ──
    # Map hierarchy (table, name) -> [column names]
    hierarchy_map: dict[tuple[str, str], list[str]] = {}
    for h in all_hierarchies:
        hierarchy_map[(h.table, h.name)] = h.columns

    # Find which hierarchies are referenced in reports (via HierarchyLevel refs)
    used_hierarchies: set[tuple[str, str]] = set()
    for u in all_usages:
        if u.ref_type == "HierarchyLevel":
            used_hierarchies.add(normalize_key(u.table, u.name))

    # Build hierarchy-column usage: columns backing used hierarchies
    hierarchy_col_keys: dict[tuple[str, str], str] = {}  # nkey -> hierarchy name
    for (tbl, hier_name), columns in hierarchy_map.items():
        hier_nkey = normalize_key(tbl, hier_name)
        if hier_nkey in used_hierarchies:
            for col in columns:
                col_nkey = normalize_key(tbl, col)
                hierarchy_col_keys[col_nkey] = hier_name

    # ── sortByColumn index ──
    sort_by_map: dict[tuple[str, str], str] = {}  # target nkey -> source column name
    for item in all_items:
        if item.sort_by_column and item.item_type in ("Column", "Calculated Column"):
            target_nkey = normalize_key(item.table, item.sort_by_column)
            sort_by_map[target_nkey] = item.name

    # ── isKey index ──
    key_col_keys = {normalize_key(*item.key) for item in all_items if item.is_key}

    # ── Unsupported metadata index ──
    unsupported_metadata_by_key: dict[tuple[str, str], list[UnsupportedMetadataRef]] = defaultdict(list)
    for ref in all_unsupported_metadata_refs:
        for key in ref.item_keys:
            unsupported_metadata_by_key[normalize_key(*key)].append(ref)

    # ── Classify each item ──
    results = []
    for item in all_items:
        nkey = normalize_key(*item.key)
        usages = usage_index.get(nkey, [])
        stale_usages = stale_usage_index.get(nkey, [])
        has_direct_usage = nkey in direct_usage_index
        field_parameter_tables = field_parameter_targets.get(nkey, set())
        is_relationship = nkey in relationship_keys
        is_rls = nkey in rls_keys
        is_indirect_measure = item.key in indirect_measures and item.item_type == "Measure"
        is_indirect_column = nkey in indirect_columns and item.item_type in ("Column", "Calculated Column")
        is_key_col = nkey in key_col_keys
        is_hierarchy_col = nkey in hierarchy_col_keys
        is_sort_target = nkey in sort_by_map

        if item.key in broken_dax_refs:
            count = len(broken_dax_refs[item.key])
            noun = "ref" if count == 1 else "refs"
            status = f"BROKEN ({count} missing {noun})"
        elif has_direct_usage:
            status = "USED"
        elif field_parameter_tables:
            parameter_tables = ", ".join(sorted(field_parameter_tables, key=str.casefold))
            status = f"USED (Field Parameter: {parameter_tables})"
        elif is_relationship:
            status = "USED (Relationship)"
        elif is_rls:
            roles = rls_role_map.get(nkey, [])
            status = f"USED (RLS: {', '.join(roles)})"
        elif is_key_col:
            status = "USED (Key Column)"
        elif is_hierarchy_col:
            status = f"USED (Hierarchy: {hierarchy_col_keys[nkey]})"
        elif is_sort_target:
            # Only mark as sort-used if the column being sorted is itself used
            source_nkey = normalize_key(item.table, sort_by_map[nkey])
            source_used = source_nkey in usage_index or source_nkey in relationship_keys
            if source_used:
                status = f"USED (Sort Column for: {sort_by_map[nkey]})"
            elif is_indirect_measure:
                via = ", ".join(sorted(format_item_ref(k) for k in indirect_measures[item.key]))
                status = f"INDIRECT (via: {via})"
            elif is_indirect_column:
                via = ", ".join(sorted(indirect_columns[nkey]))
                status = f"INDIRECT (via: {via})"
            else:
                status = "NOT USED"
        elif is_indirect_measure:
            via = ", ".join(sorted(format_item_ref(k) for k in indirect_measures[item.key]))
            status = f"INDIRECT (via: {via})"
        elif is_indirect_column:
            via = ", ".join(sorted(indirect_columns[nkey]))
            status = f"INDIRECT (via: {via})"
        else:
            status = "NOT USED"

        # ── Removal risk ──
        review_triggers: list[str] = []
        if item.key in broken_dax_refs:
            removal_risk = ""
            review_triggers.extend([detail["message"] for detail in broken_dax_refs[item.key]])
        elif status != "NOT USED":
            removal_risk = ""
        elif item.is_inferred:
            removal_risk = "Do not remove"
        else:
            # Check if any other item's DAX references this item
            has_dax_dependents = False
            for dep_key, refs in dax_deps.items():
                if item.key in refs:
                    has_dax_dependents = True
                    break
            if not has_dax_dependents:
                for dep_key, refs in dax_col_deps.items():
                    if item.key in refs:
                        has_dax_dependents = True
                        break
            if item.is_hidden:
                review_triggers.append("Item is hidden")
            if item.is_key:
                review_triggers.append("Item is marked as a key")
            review_triggers.extend(
                _unsupported_metadata_review_trigger(item, ref)
                for ref in unsupported_metadata_by_key.get(nkey, [])
            )
            if review_triggers:
                removal_risk = "Review"
            elif has_dax_dependents:
                removal_risk = "Caution"
            else:
                removal_risk = "Safe"

        results.append({
            "item": item,
            "status": status,
            "usages": usages,
            "stale_usages": stale_usages,
            "has_direct_usage": has_direct_usage,
            "removal_risk": removal_risk,
            "review_triggers": review_triggers,
            "broken_dax_refs": [detail["ref"] for detail in broken_dax_refs.get(item.key, [])],
            "broken_dax_ref_details": broken_dax_refs.get(item.key, []),
        })

    # ── Table-level summary ──
    field_parameter_issues_by_table: dict[str, list[str]] = defaultdict(list)
    for warning in warnings:
        if warning.code in ("UNRESOLVED_NAMEOF_TARGET", "AMBIGUOUS_NAMEOF_TARGET") and warning.table:
            field_parameter_issues_by_table[warning.table].append(warning.message)

    table_summaries = build_table_summaries(
        results,
        all_usages,
        all_relationship_details,
        dax_col_deps,
        dax_table_deps,
        {table: sorted(set(messages), key=str.casefold) for table, messages in field_parameter_issues_by_table.items()},
    )

    table_stats: dict[str, dict] = defaultdict(lambda: {
        "total": 0, "used": 0, "unused": 0,
        "measures": 0, "measures_used": 0,
        "columns": 0, "columns_used": 0,
    })
    for r in results:
        tbl = r["item"].table
        is_measure = r["item"].item_type == "Measure"
        is_used = is_used_status(r["status"])
        table_stats[tbl]["total"] += 1
        if is_used:
            table_stats[tbl]["used"] += 1
        else:
            table_stats[tbl]["unused"] += 1
        if is_measure:
            table_stats[tbl]["measures"] += 1
            if is_used:
                table_stats[tbl]["measures_used"] += 1
        else:
            table_stats[tbl]["columns"] += 1
            if is_used:
                table_stats[tbl]["columns_used"] += 1

    # ── Summary stats ──
    summary = {
        "total_measures": sum(1 for i in all_items if i.item_type == "Measure"),
        "total_model_measures": sum(
            1 for i in all_items
            if i.item_type == "Measure" and i.source_kind == "model"
        ),
        "total_report_measures": sum(
            1 for i in all_items
            if i.item_type == "Measure" and i.source_kind == "report"
        ),
        "total_columns": sum(1 for i in all_items if i.item_type in ("Column", "Calculated Column")),
        "total_calc_columns": sum(1 for i in all_items if i.item_type == "Calculated Column"),
        "used_in_visuals": sum(
            1 for r in results
            if r["status"] == "USED" or r["status"].startswith("USED (Field Parameter:")
        ),
        "used_relationship": sum(1 for r in results if "Relationship" in r["status"]),
        "used_rls": sum(1 for r in results if "RLS" in r["status"]),
        "used_key_column": sum(1 for r in results if "Key Column" in r["status"]),
        "used_hierarchy": sum(1 for r in results if "Hierarchy" in r["status"]),
        "used_sort_column": sum(1 for r in results if "Sort Column" in r["status"]),
        "indirect": sum(1 for r in results if r["status"].startswith("INDIRECT")),
        "broken": sum(1 for r in results if r["status"].startswith("BROKEN")),
        "not_used": sum(1 for r in results if r["status"] == "NOT USED"),
        "total_usage_refs": len(all_usages),
        "models": [m.name for m in models],
        "reports": [report_display_name(r) for r in reports],
        "tables": dict(table_stats),
    }

    return {
        "items": results,
        "summary": summary,
        "table_summaries": table_summaries,
        "warnings": [_serialize_warning(w) for w in warnings],
        "report_issues": [_serialize_report_issue(issue) for issue in report_issues],
    }


# ── Output Formatters ─────────────────────────────────────────────────────────


def _escape_pipe(s: str) -> str:
    return s.replace("|", "\\|")


def _format_warnings_section(results: dict) -> list[str]:
    warnings = results.get("warnings", [])
    if not warnings:
        return []

    lines = ["## Warnings\n"]
    for warning in warnings:
        detail_parts = []
        if warning.get("code"):
            detail_parts.append(f"[{warning['code']}]")
        if warning.get("model"):
            detail_parts.append(f"model={warning['model']}")
        if warning.get("table"):
            detail_parts.append(f"table={warning['table']}")
        if warning.get("source_file"):
            detail_parts.append(f"file={warning['source_file']}")

        detail = f" ({'; '.join(detail_parts)})" if detail_parts else ""
        lines.append(f"- {warning['message']}{detail}")
    lines.append("")
    return lines


def format_full(results: dict) -> str:
    lines = []
    s = results["summary"]

    lines.append("# Semantic Model Usage Analysis\n")
    lines.append(f"**Models**: {', '.join(s['models'])}")
    lines.append(f"**Reports**: {', '.join(s['reports'])}\n")
    lines.extend(_format_warnings_section(results))

    lines.append("## Summary\n")
    lines.append("| Metric | Count |")
    lines.append("|--------|------:|")
    lines.append(f"| Total measures | {s['total_measures']} |")
    lines.append(f"| Total columns | {s['total_columns']} |")
    lines.append(f"| Calculated columns | {s['total_calc_columns']} |")
    lines.append(f"| Used in visuals | {s['used_in_visuals']} |")
    lines.append(f"| Used in relationships only | {s['used_relationship']} |")
    lines.append(f"| Used in RLS only | {s['used_rls']} |")
    lines.append(f"| Used as key column | {s['used_key_column']} |")
    lines.append(f"| Used in hierarchy | {s['used_hierarchy']} |")
    lines.append(f"| Used as sort column | {s['used_sort_column']} |")
    lines.append(f"| Indirectly used (via DAX) | {s['indirect']} |")
    lines.append(f"| Broken DAX items | {s.get('broken', 0)} |")
    lines.append(f"| **Not used** | **{s['not_used']}** |")
    lines.append(f"| Total usage references | {s['total_usage_refs']} |")
    lines.append("")

    # ── Table-level summary ──
    table_stats = s.get("tables", {})
    if table_stats:
        lines.append("## Table Summary\n")
        lines.append("| Table | Total | Used | Unused | % Used |")
        lines.append("|-------|------:|-----:|-------:|-------:|")
        for tbl in sorted(table_stats.keys(), key=str.casefold):
            ts = table_stats[tbl]
            total = ts["total"]
            used = ts["used"]
            unused = ts["unused"]
            pct = round(used / total * 100) if total else 0
            marker = " \u2717" if pct == 0 else ""
            lines.append(f"| {_escape_pipe(tbl)} | {total} | {used} | {unused} | {pct}%{marker} |")
        lines.append("")

    lines.append("## Detailed Usage\n")
    lines.append("| Type | Table | Name | Hidden | Display Folder | Report | Page | Visual | Context | Status | Risk |")
    lines.append("|------|-------|------|--------|----------------|--------|------|--------|---------|--------|------|")

    def sort_key(r):
        st = r["status"]
        if st.startswith("BROKEN"):
            order = 0
        elif st == "NOT USED":
            order = 1
        elif st.startswith("INDIRECT"):
            order = 2
        else:
            order = 3
        return (order, r["item"].table, r["item"].name)

    for r in sorted(results["items"], key=sort_key):
        item = r["item"]
        usages = r["usages"]
        status = r["status"]
        hidden = "\u2713" if item.is_hidden else "\u2014"
        display_folder = _escape_pipe(item.display_folder) or "\u2014"
        risk = r.get("removal_risk", "")

        if usages:
            for u in usages:
                title_str = f' "{_escape_pipe(u.visual_title)}"' if u.visual_title else ""
                lines.append(
                    f"| {item.item_type} | {_escape_pipe(item.table)} | {_escape_pipe(item.name)} | "
                    f"{hidden} | {display_folder} | {_escape_pipe(u.report)} | "
                    f"{_escape_pipe(u.page)} | {_escape_pipe(u.visual_type)}{title_str} | "
                    f"{_escape_pipe(u.context)} | {_escape_pipe(status)} | {risk} |"
                )
        else:
            lines.append(
                f"| {item.item_type} | {_escape_pipe(item.table)} | {_escape_pipe(item.name)} | "
                f"{hidden} | {display_folder} | \u2014 | \u2014 | \u2014 | \u2014 | "
                f"{_escape_pipe(status)} | {risk} |"
            )

    return "\n".join(lines)


def format_unused(results: dict) -> str:
    lines = []
    s = results["summary"]
    total = s["total_measures"] + s["total_columns"]

    lines.append("# Unused Measures & Columns\n")
    lines.append(f"**Models**: {', '.join(s['models'])} | **Reports**: {', '.join(s['reports'])}\n")
    lines.extend(_format_warnings_section(results))
    lines.append(f"**{s['not_used']}** unused out of **{total}** total items\n")

    unused = [r for r in results["items"] if r["status"] == "NOT USED"]

    if not unused:
        lines.append("All measures and columns are used. \u2714")
        return "\n".join(lines)

    # ── Table-level summary (unused tables only) ──
    table_stats = s.get("tables", {})
    fully_unused = [t for t, ts in table_stats.items() if ts["used"] == 0]
    if fully_unused:
        lines.append("## Fully Unused Tables\n")
        for tbl in sorted(fully_unused, key=str.casefold):
            lines.append(f"- **{tbl}** ({table_stats[tbl]['total']} items)")
        lines.append("")

    lines.append("## Unused Items\n")
    lines.append("| Type | Table | Name | Hidden | Display Folder | Risk |")
    lines.append("|------|-------|------|--------|----------------|------|")

    risk_order = {"Safe": 0, "Review": 1, "Caution": 2, "Do not remove": 3}

    for r in sorted(unused, key=lambda x: (
        risk_order.get(x.get("removal_risk", ""), 9),
        x["item"].item_type,
        x["item"].table,
        x["item"].name,
    )):
        item = r["item"]
        hidden = "\u2713" if item.is_hidden else "\u2014"
        display_folder = _escape_pipe(item.display_folder) or "\u2014"
        risk = r.get("removal_risk", "")
        lines.append(
            f"| {item.item_type} | {_escape_pipe(item.table)} | "
            f"{_escape_pipe(item.name)} | {hidden} | {display_folder} | {risk} |"
        )

    return "\n".join(lines)


def format_json_output(results: dict) -> str:
    output = {
        "summary": results["summary"],
        "tables": results.get("table_summaries", []),
        "warnings": results.get("warnings", []),
        "reportIssues": results.get("report_issues", []),
        "items": [],
    }
    for r in results["items"]:
        output["items"].append({
            "type": r["item"].item_type,
            "table": r["item"].table,
            "name": r["item"].name,
            "sourceKind": r["item"].source_kind,
            "sourceArtifact": r["item"].source_artifact or None,
            "sourceFile": r["item"].source_file or None,
            "displayFolder": r["item"].display_folder,
            "formatString": r["item"].format_string or None,
            "isHidden": r["item"].is_hidden,
            "isKey": r["item"].is_key,
            "isInferred": r["item"].is_inferred,
            "sortByColumn": r["item"].sort_by_column or None,
            "status": r["status"],
            "removalRisk": r.get("removal_risk", "") or None,
                "reviewTriggers": r.get("review_triggers", []),
                "brokenDaxRefs": r.get("broken_dax_refs", []),
                "brokenDaxRefDetails": r.get("broken_dax_ref_details", []),
                "usages": [
                {
                    "report": u.report,
                    "page": u.page,
                    "visualType": u.visual_type,
                    "visualTitle": u.visual_title,
                    "context": u.context,
                }
                for u in r["usages"]
            ],
        })
    return json.dumps(output, indent=2, ensure_ascii=False)


def format_xlsx(results: dict, output_path: str, announce: bool = True) -> None:
    """Write results to an Excel workbook with Summary, Details, and Unused sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl is required for xlsx output. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"

    def _auto_width(ws, min_width=10, max_width=50):
        for col in ws.columns:
            length = min_width
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = max(length, min(len(str(cell.value)) + 2, max_width))
            ws.column_dimensions[col_letter].width = length

    s = results["summary"]
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    label_font = Font(bold=True)
    light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    warn_font = Font(bold=True, color="FF0000")

    def _section_banner(ws, row, text, cols=2):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = section_fill
            cell.font = section_font
            cell.border = thin_border
        ws.cell(row=row, column=1, value=text)

    def _kv_row(ws, row, label, value, bold_value=False):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=value)
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="right")
        if bold_value:
            c2.font = Font(bold=True)

    # ── Sheet 1: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "4472C4"

    # Title
    row = 1
    ws_sum.cell(row=row, column=1, value="Semantic Model Usage Analysis").font = Font(bold=True, size=16)
    ws_sum.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    # Model / Report info
    row = 3
    _section_banner(ws_sum, row, "Source", 2)
    row += 1
    _kv_row(ws_sum, row, "Models", ", ".join(s["models"]))
    row += 1
    _kv_row(ws_sum, row, "Reports", ", ".join(s["reports"]))

    # ── Overall Metrics ──
    row += 2
    _section_banner(ws_sum, row, "Overall Metrics", 2)
    total_items = s["total_measures"] + s["total_columns"]
    total_used = total_items - s["not_used"] - s.get("broken", 0)
    overall_metrics = [
        ("Total items", total_items),
        ("Total used", total_used),
        ("Broken DAX items", s.get("broken", 0)),
        ("Total unused", s["not_used"]),
        ("Usage references", s["total_usage_refs"]),
    ]
    for label, val in overall_metrics:
        row += 1
        _kv_row(ws_sum, row, label, val, bold_value=(label == "Total unused"))

    # ── Measures ──
    row += 2
    _section_banner(ws_sum, row, "Measures", 2)
    measures_used = sum(
        1 for r in results["items"]
        if r["item"].item_type == "Measure" and is_used_status(r["status"])
    )
    measures_unused = s["total_measures"] - measures_used
    measure_metrics = [
        ("Total measures", s["total_measures"]),
        ("Used", measures_used),
        ("Unused", measures_unused),
        ("Indirectly used (via DAX)", s["indirect"]),
    ]
    for label, val in measure_metrics:
        row += 1
        _kv_row(ws_sum, row, label, val, bold_value=(label == "Unused"))

    # ── Columns ──
    row += 2
    _section_banner(ws_sum, row, "Columns", 2)
    columns_used = sum(
        1 for r in results["items"]
        if r["item"].item_type in ("Column", "Calculated Column") and is_used_status(r["status"])
    )
    columns_unused = s["total_columns"] - columns_used
    column_metrics = [
        ("Total columns", s["total_columns"]),
        ("Calculated columns", s["total_calc_columns"]),
        ("Used", columns_used),
        ("Unused", columns_unused),
        ("Used in visuals", s["used_in_visuals"]),
        ("Used in relationships only", s["used_relationship"]),
        ("Used in RLS only", s["used_rls"]),
        ("Used as key column", s["used_key_column"]),
        ("Used in hierarchy", s["used_hierarchy"]),
        ("Used as sort column", s["used_sort_column"]),
    ]
    for label, val in column_metrics:
        row += 1
        _kv_row(ws_sum, row, label, val, bold_value=(label == "Unused"))

    # ── Table Summary ──
    table_stats = s.get("tables", {})
    if table_stats:
        row += 2
        tbl_headers = ["Table", "Measures", "Measures Used", "Columns", "Columns Used",
                        "Total", "Used", "Unused", "% Used"]
        _section_banner(ws_sum, row, "Table Summary", len(tbl_headers))
        row += 1
        for i, h in enumerate(tbl_headers, 1):
            cell = ws_sum.cell(row=row, column=i, value=h)
            cell.font = label_font
            cell.fill = light_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        for tbl in sorted(table_stats.keys(), key=str.casefold):
            row += 1
            ts = table_stats[tbl]
            total = ts["total"]
            used = ts["used"]
            unused = ts["unused"]
            pct = round(used / total * 100) if total else 0
            values = [
                tbl,
                ts["measures"], ts["measures_used"],
                ts["columns"], ts["columns_used"],
                total, used, unused, f"{pct}%",
            ]
            for i, val in enumerate(values, 1):
                cell = ws_sum.cell(row=row, column=i, value=val)
                cell.border = thin_border
                if i >= 2:
                    cell.alignment = Alignment(horizontal="center")
            if pct == 0:
                ws_sum.cell(row=row, column=9).font = warn_font

    # Column widths
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 45
    for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws_sum.column_dimensions[col_letter].width = 15

    # ── Sheet 2: Details (one row per item, deduplicated) ──
    ws_detail = wb.create_sheet("Details")
    detail_headers = ["Type", "Table", "Name", "Hidden", "Display Folder",
                      "Status", "Removal Risk", "Review Triggers", "Pages Used",
                      "Visual Types", "Contexts"]
    _write_header(ws_detail, detail_headers)

    risk_order = {"Safe": 0, "Review": 1, "Caution": 2, "Do not remove": 3}
    not_used_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    caution_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    sorted_items = sorted(results["items"], key=lambda x: (
        risk_order.get(x.get("removal_risk", ""), 9),
        x["item"].item_type,
        x["item"].table,
        x["item"].name,
    ))

    for row_idx, r in enumerate(sorted_items, 2):
        item = r["item"]
        usages = r["usages"]
        pages = sorted({u.page for u in usages if u.page and u.page != "\u2014"})
        visual_types = sorted({u.visual_type for u in usages if u.visual_type and u.visual_type != "\u2014"})
        contexts = sorted({u.context for u in usages if u.context and u.context != "\u2014"})

        row_data = [
            item.item_type,
            item.table,
            item.name,
            "Yes" if item.is_hidden else "No",
            item.display_folder or "",
            r["status"],
            r.get("removal_risk", ""),
            "; ".join(r.get("review_triggers", [])),
            ", ".join(pages),
            ", ".join(visual_types),
            ", ".join(contexts),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_detail.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
        if r["status"] == "NOT USED":
            risk = r.get("removal_risk", "")
            fill = caution_fill if risk in ("Caution", "Do not remove") else not_used_fill
            for col in range(1, len(row_data) + 1):
                ws_detail.cell(row=row_idx, column=col).fill = fill

    _auto_width(ws_detail)

    # ── Sheet 3: All References (one row per usage ref) ──
    ws_refs = wb.create_sheet("All References")
    ref_headers = ["Type", "Table", "Name", "Hidden", "Display Folder",
                   "Report", "Page", "Visual Type", "Visual Title", "Context",
                   "Status", "Removal Risk", "Review Triggers"]
    _write_header(ws_refs, ref_headers)

    row_idx = 2
    for r in sorted_items:
        item = r["item"]
        hidden = "Yes" if item.is_hidden else "No"
        status = r["status"]
        risk = r.get("removal_risk", "")
        review_triggers = "; ".join(r.get("review_triggers", []))
        if r["usages"]:
            for u in r["usages"]:
                row_data = [
                    item.item_type, item.table, item.name, hidden,
                    item.display_folder or "",
                    u.report, u.page, u.visual_type, u.visual_title or "",
                    u.context, status, risk, review_triggers,
                ]
                for col, val in enumerate(row_data, 1):
                    ws_refs.cell(row=row_idx, column=col, value=val).border = thin_border
                row_idx += 1
        else:
            row_data = [
                item.item_type, item.table, item.name, hidden,
                item.display_folder or "",
                "", "", "", "", "", status, risk, review_triggers,
            ]
            for col, val in enumerate(row_data, 1):
                ws_refs.cell(row=row_idx, column=col, value=val).border = thin_border
            if status == "NOT USED":
                fill = caution_fill if risk in ("Caution", "Do not remove") else not_used_fill
                for col in range(1, len(row_data) + 1):
                    ws_refs.cell(row=row_idx, column=col).fill = fill
            row_idx += 1

    _auto_width(ws_refs)

    wb.save(output_path)
    if announce:
        print(f"Excel report saved to: {output_path}")


def create_xlsx_bytes(results: dict) -> bytes:
    """Render the Excel report to bytes for HTTP downloads or other in-memory use."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = Path(tmp_dir) / "analysis.xlsx"
        format_xlsx(results, str(output_path), announce=False)
        return output_path.read_bytes()


# ── Entry Point ───────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Analyze one TMDL semantic model against one or more PBIR reports",
    )
    parser.add_argument(
        "workspace",
        nargs="?",
        default=".",
        help="Base search path (default: .). Supports non-standard layouts via recursive discovery.",
    )
    parser.add_argument(
        "--format",
        choices=["full", "unused", "json", "xlsx"],
        default="full",
        help="Output format: full (detailed matrix), unused (only unused items), json, xlsx (Excel)",
    )
    parser.add_argument(
        "-o", "--output",
        help="Output file path. Required for xlsx format. For other formats, writes to file instead of stdout.",
    )
    parser.add_argument(
        "--models-path",
        nargs="+",
        help="One or more paths to search for the single .SemanticModel folder to analyze",
    )
    parser.add_argument(
        "--reports-path",
        nargs="+",
        help="One or more paths to search for .Report folders (or direct .Report paths)",
    )
    parser.add_argument(
        "--model",
        nargs="+",
        help="Filter semantic models by name (substring match, case-insensitive)",
    )
    parser.add_argument(
        "--report",
        nargs="+",
        help="Filter reports by name (substring match, case-insensitive)",
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Interactive selector for models and reports",
    )
    args = parser.parse_args()

    workspace = Path(args.workspace).resolve()
    if not workspace.exists() and not args.models_path and not args.reports_path:
        print(f"Error: workspace path does not exist: {workspace}", file=sys.stderr)
        sys.exit(1)

    model_roots = [Path(p).resolve() for p in args.models_path] if args.models_path else [workspace]
    report_roots = [Path(p).resolve() for p in args.reports_path] if args.reports_path else [workspace]

    for p in model_roots + report_roots:
        if not p.exists():
            print(f"Error: search path does not exist: {p}", file=sys.stderr)
            sys.exit(1)

    models = filter_models(discover_models(model_roots), args.model)
    reports = filter_reports(discover_reports(report_roots), args.report)

    if args.interactive:
        if not sys.stdin.isatty():
            print("Error: --interactive requires a TTY terminal.", file=sys.stderr)
            sys.exit(1)
        models = select_paths_interactively(
            "models",
            models,
            lambda p: f"{p.name} ({p})",
        )
        reports = select_paths_interactively(
            "reports",
            reports,
            lambda p: f"{report_display_name(p)} ({p})",
        )

    if len(models) != 1:
        if not models:
            print("Error: No matching semantic model found.", file=sys.stderr)
        else:
            found = ", ".join(str(m) for m in models)
            print(
                "Error: exactly one semantic model must be selected. "
                f"Found {len(models)}: {found}",
                file=sys.stderr,
            )
        sys.exit(1)

    results = analyze(
        workspace,
        model_paths=models,
        report_paths=reports,
        model_search_roots=model_roots,
        report_search_roots=report_roots,
    )

    if args.format == "xlsx":
        output_path = args.output
        if not output_path:
            model_name = models[0].name.replace(".SemanticModel", "") if models else "model"
            output_path = f"{model_name}_usage_analysis.xlsx"
        format_xlsx(results, output_path)
    else:
        if args.format == "full":
            text = format_full(results)
        elif args.format == "unused":
            text = format_unused(results)
        elif args.format == "json":
            text = format_json_output(results)
        if args.output:
            Path(args.output).write_text(text, encoding="utf-8")
            print(f"Report saved to: {args.output}")
        else:
            print(text)


if __name__ == "__main__":
    main()
