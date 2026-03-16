#!/usr/bin/env python3
"""
Semantic Model Compare Engine (TMDL vs TMDL).

Structured model-to-model comparison with pluggable sources.
MVP source implementation is local filesystem TMDL.
"""

from __future__ import annotations

import json
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from . import analyzer


LEVEL_ORDER = ["model", "relationship", "table", "measure", "column"]
CHANGE_ORDER = ["added", "removed", "changed"]


@dataclass
class CompareWarning:
    code: str
    severity: str
    message: str
    side: str | None = None
    source_file: str | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "severity": self.severity,
            "message": self.message,
            "side": self.side,
            "source_file": self.source_file,
        }


@dataclass
class MeasureSnapshot:
    table: str
    name: str
    properties: dict[str, Any]

    def as_payload(self) -> dict[str, Any]:
        return {
            "table": self.table,
            "name": self.name,
            "properties": dict(sorted(self.properties.items())),
        }


@dataclass
class ColumnSnapshot:
    table: str
    name: str
    properties: dict[str, Any]

    def as_payload(self) -> dict[str, Any]:
        return {
            "table": self.table,
            "name": self.name,
            "properties": dict(sorted(self.properties.items())),
        }


@dataclass
class TableSnapshot:
    name: str
    properties: dict[str, Any]
    measures: dict[str, MeasureSnapshot] = field(default_factory=dict)
    columns: dict[str, ColumnSnapshot] = field(default_factory=dict)

    def as_payload(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "properties": dict(sorted(self.properties.items())),
            "measureNames": sorted(self.measures.keys()),
            "columnNames": sorted(self.columns.keys()),
        }


@dataclass
class RelationshipSnapshot:
    name: str
    from_table: str
    from_column: str
    to_table: str
    to_column: str
    properties: dict[str, Any]

    @property
    def key(self) -> tuple[str, str, str, str, str]:
        return (self.name, self.from_table, self.from_column, self.to_table, self.to_column)

    def as_payload(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "fromTable": self.from_table,
            "fromColumn": self.from_column,
            "toTable": self.to_table,
            "toColumn": self.to_column,
            "properties": dict(sorted(self.properties.items())),
        }


@dataclass
class ModelSnapshot:
    model_name: str
    model_properties: dict[str, Any]
    tables: dict[str, TableSnapshot]
    relationships: dict[tuple[str, str, str, str, str], RelationshipSnapshot]

    def model_payload(self) -> dict[str, Any]:
        return {
            "name": self.model_name,
            "properties": dict(sorted(self.model_properties.items())),
        }


class SemanticModelSource(ABC):
    """Adapter contract for loading semantic models from different backends."""

    @abstractmethod
    def load_snapshot(self, model_path: Path) -> tuple[ModelSnapshot, list[CompareWarning]]:
        raise NotImplementedError


class LocalTmdlSource(SemanticModelSource):
    """Local filesystem TMDL implementation of SemanticModelSource."""

    def load_snapshot(self, model_path: Path) -> tuple[ModelSnapshot, list[CompareWarning]]:
        warnings: list[CompareWarning] = []

        definition_dir = model_path / "definition"
        if not definition_dir.exists():
            warnings.append(CompareWarning(
                code="MISSING_DEFINITION_DIR",
                severity="warning",
                message=f"Definition folder was not found for model: {model_path}",
                source_file=str(model_path),
            ))

        model_name = model_path.name.replace(".SemanticModel", "")
        model_properties = self._parse_model_properties(model_path, warnings)
        tables = self._parse_tables(model_path, warnings)
        relationships = self._parse_relationships(model_path, warnings)

        return (
            ModelSnapshot(
                model_name=model_name,
                model_properties=model_properties,
                tables=tables,
                relationships=relationships,
            ),
            warnings,
        )

    def _parse_model_properties(
        self,
        model_path: Path,
        warnings: list[CompareWarning],
    ) -> dict[str, Any]:
        model_file = model_path / "definition" / "model.tmdl"
        if not model_file.exists():
            warnings.append(CompareWarning(
                code="MISSING_MODEL_FILE",
                severity="warning",
                message=f"model.tmdl was not found under {model_path}",
                source_file=str(model_file),
            ))
            return {}

        try:
            lines = model_file.read_text(encoding="utf-8").splitlines()
        except Exception as exc:
            warnings.append(CompareWarning(
                code="MODEL_FILE_READ_FAILED",
                severity="warning",
                message=f"Failed to read {model_file}: {exc}",
                source_file=str(model_file),
            ))
            return {}

        properties: dict[str, Any] = {}
        in_model_block = False
        for line in lines:
            if not line.startswith("\t") and line.startswith("model "):
                in_model_block = True
                decl_name = line[6:].strip().strip("'\"")
                if decl_name:
                    properties["declarationName"] = decl_name
                continue

            if not in_model_block:
                continue
            if not line.startswith("\t") and line.strip():
                break
            if line.startswith("\t\t"):
                continue
            if not line.startswith("\t"):
                continue

            stripped = line.strip()
            if not stripped or stripped.startswith("//"):
                continue
            if (
                stripped.startswith("ref table ")
                or stripped.startswith("annotation ")
                or stripped.startswith("culture ")
                or stripped.startswith("role ")
                or stripped.startswith("perspective ")
                or stripped.startswith("tablePermission ")
            ):
                continue
            parsed = _parse_scalar_property(stripped)
            if parsed:
                key, value = parsed
                properties[key] = value

        return dict(sorted(properties.items()))

    def _parse_tables(
        self,
        model_path: Path,
        warnings: list[CompareWarning],
    ) -> dict[str, TableSnapshot]:
        tables_dir = model_path / "definition" / "tables"
        if not tables_dir.exists():
            warnings.append(CompareWarning(
                code="MISSING_TABLES_DIR",
                severity="warning",
                message=f"tables folder was not found for model: {model_path}",
                source_file=str(tables_dir),
            ))
            return {}

        tables: dict[str, TableSnapshot] = {}
        for filepath in sorted(tables_dir.glob("*.tmdl")):
            table = self._parse_table_file(filepath, warnings)
            if table:
                tables[table.name] = table
        return dict(sorted(tables.items(), key=lambda item: item[0].casefold()))

    def _parse_table_file(
        self,
        filepath: Path,
        warnings: list[CompareWarning],
    ) -> TableSnapshot | None:
        try:
            lines = filepath.read_text(encoding="utf-8").splitlines()
        except Exception as exc:
            warnings.append(CompareWarning(
                code="TABLE_FILE_READ_FAILED",
                severity="warning",
                message=f"Failed to read table file {filepath.name}: {exc}",
                source_file=str(filepath),
            ))
            return None

        table_name: str | None = None
        table_props: dict[str, Any] = {}
        measures: dict[str, MeasureSnapshot] = {}
        columns: dict[str, ColumnSnapshot] = {}

        i = 0
        while i < len(lines):
            line = lines[i]

            if not line.startswith("\t") and line.startswith("table "):
                table_name = line[6:].strip().strip("'\"")
                i += 1
                continue

            if not table_name:
                i += 1
                continue

            if line.startswith("\tmeasure "):
                measure, i = _parse_measure_block(lines, i, table_name)
                measures[measure.name] = measure
                continue

            if line.startswith("\tcolumn "):
                column, i = _parse_column_block(lines, i, table_name)
                columns[column.name] = column
                continue

            if line.startswith("\t") and not line.startswith("\t\t"):
                stripped = line.strip()
                if _is_table_sub_block(stripped):
                    i += 1
                    continue
                parsed = _parse_scalar_property(stripped)
                if parsed:
                    key, value = parsed
                    table_props[key] = value

            i += 1

        if not table_name:
            warnings.append(CompareWarning(
                code="TABLE_DECLARATION_NOT_FOUND",
                severity="warning",
                message=f"No table declaration found in {filepath.name}",
                source_file=str(filepath),
            ))
            return None

        return TableSnapshot(
            name=table_name,
            properties=dict(sorted(table_props.items())),
            measures=dict(sorted(measures.items(), key=lambda item: item[0].casefold())),
            columns=dict(sorted(columns.items(), key=lambda item: item[0].casefold())),
        )

    def _parse_relationships(
        self,
        model_path: Path,
        warnings: list[CompareWarning],
    ) -> dict[tuple[str, str, str, str, str], RelationshipSnapshot]:
        rel_file = model_path / "definition" / "relationships.tmdl"
        if not rel_file.exists():
            return {}

        relationships: dict[tuple[str, str, str, str, str], RelationshipSnapshot] = {}
        try:
            for rel in analyzer.parse_relationship_details(model_path):
                snapshot = RelationshipSnapshot(
                    name=rel.name,
                    from_table=rel.from_table,
                    from_column=rel.from_column,
                    to_table=rel.to_table,
                    to_column=rel.to_column,
                    properties={
                        "fromCardinality": rel.from_cardinality,
                        "toCardinality": rel.to_cardinality,
                        "isActive": rel.is_active,
                    },
                )
                relationships[snapshot.key] = snapshot
        except Exception as exc:
            warnings.append(CompareWarning(
                code="RELATIONSHIP_PARSE_FAILED",
                severity="warning",
                message=f"Failed to parse relationships for {model_path}: {exc}",
                source_file=str(rel_file),
            ))
        return dict(sorted(relationships.items(), key=lambda item: _relationship_sort_key(item[0])))


def _relationship_sort_key(key: tuple[str, str, str, str, str]) -> tuple[str, str, str, str, str]:
    return tuple(part.casefold() for part in key)


def _is_table_sub_block(stripped: str) -> bool:
    return (
        stripped.startswith("partition ")
        or stripped.startswith("hierarchy ")
        or stripped.startswith("annotation ")
        or stripped.startswith("column ")
        or stripped.startswith("measure ")
        or stripped.startswith("calculationGroup ")
        or stripped.startswith("detailRowsDefinition")
        or stripped.startswith("changedProperty ")
    )


def _strip_quotes(value: str) -> str:
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
        return value[1:-1]
    return value


def _normalize_scalar(value: str) -> Any:
    raw = _strip_quotes(value.strip())
    lower = raw.casefold()
    if lower in ("true", "false"):
        return lower == "true"
    if re.fullmatch(r"[+-]?\d+", raw):
        try:
            return int(raw)
        except ValueError:
            return raw
    return raw


def _canonical_property_key(key: str) -> str:
    compact = key.strip()
    lower = compact.casefold()
    return {
        "hidden": "isHidden",
        "ishidden": "isHidden",
        "displayfolder": "displayFolder",
        "iskey": "isKey",
        "isnameinferred": "isNameInferred",
        "isdatatypeinferred": "isDataTypeInferred",
        "sortbycolumn": "sortByColumn",
        "fromcardinality": "fromCardinality",
        "tocardinality": "toCardinality",
        "isactive": "isActive",
    }.get(lower, compact)


def _parse_scalar_property(stripped_line: str) -> tuple[str, Any] | None:
    if not stripped_line:
        return None
    if stripped_line.startswith("//"):
        return None

    if ":" in stripped_line:
        key, value = stripped_line.split(":", 1)
        key = key.strip()
        if not key or " " in key:
            return None
        return (_canonical_property_key(key), _normalize_scalar(value))

    if "=" in stripped_line:
        key, value = stripped_line.split("=", 1)
        key = key.strip()
        if not key or " " in key:
            return None
        return (_canonical_property_key(key), _normalize_scalar(value))

    if " " in stripped_line:
        return None
    return (_canonical_property_key(stripped_line), True)


def _parse_measure_block(
    lines: list[str],
    start_idx: int,
    table_name: str,
) -> tuple[MeasureSnapshot, int]:
    line = lines[start_idx]
    declaration = line[len("\tmeasure "):].strip()
    if "=" in declaration:
        raw_name, first_expr = declaration.split("=", 1)
    else:
        raw_name, first_expr = declaration, ""
    name = _strip_quotes(raw_name.strip())

    props: dict[str, Any] = {}
    expr_lines = [first_expr.strip()] if first_expr.strip() else []
    i = start_idx + 1

    while i < len(lines):
        inner = lines[i]
        if inner.startswith("\t") and not inner.startswith("\t\t"):
            break
        if not inner.startswith("\t") and inner.strip():
            break
        if not inner.strip():
            i += 1
            continue

        if inner.startswith("\t\t\t"):
            expr_lines.append(inner.strip())
            i += 1
            continue

        if inner.startswith("\t\t"):
            stripped = inner.strip()
            if stripped.startswith("expression") and "=" in stripped:
                expr_lines.append(stripped.split("=", 1)[1].strip())
            elif stripped.startswith("formatStringDefinition") and "=" in stripped:
                expr_lines.append(stripped.split("=", 1)[1].strip())
            else:
                parsed = _parse_scalar_property(stripped)
                if parsed:
                    key, value = parsed
                    props[key] = value
            i += 1
            continue

        i += 1

    normalized_expression = _normalize_expression("\n".join(expr_lines))
    if normalized_expression:
        props["expression"] = normalized_expression

    return (
        MeasureSnapshot(
            table=table_name,
            name=name,
            properties=dict(sorted(props.items())),
        ),
        i,
    )


def _parse_column_block(
    lines: list[str],
    start_idx: int,
    table_name: str,
) -> tuple[ColumnSnapshot, int]:
    line = lines[start_idx]
    declaration = line[len("\tcolumn "):].strip()
    name = _strip_quotes(declaration)

    props: dict[str, Any] = {}
    expr_lines: list[str] = []
    i = start_idx + 1

    while i < len(lines):
        inner = lines[i]
        if inner.startswith("\t") and not inner.startswith("\t\t"):
            break
        if not inner.startswith("\t") and inner.strip():
            break
        if not inner.strip():
            i += 1
            continue

        if inner.startswith("\t\t\t"):
            expr_lines.append(inner.strip())
            i += 1
            continue

        if inner.startswith("\t\t"):
            stripped = inner.strip()
            if stripped.startswith("expression") and "=" in stripped:
                expr_lines.append(stripped.split("=", 1)[1].strip())
            else:
                parsed = _parse_scalar_property(stripped)
                if parsed:
                    key, value = parsed
                    props[key] = value
            i += 1
            continue

        i += 1

    normalized_expression = _normalize_expression("\n".join(expr_lines))
    if normalized_expression:
        props["expression"] = normalized_expression
        props["itemType"] = "Calculated Column"
    else:
        props["itemType"] = "Column"

    return (
        ColumnSnapshot(
            table=table_name,
            name=name,
            properties=dict(sorted(props.items())),
        ),
        i,
    )


def _normalize_expression(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized: list[str] = []
    for line in text.split("\n"):
        stripped = line.strip()
        if stripped == "```":
            continue
        if stripped.startswith("//"):
            continue
        normalized.append(re.sub(r"\s+$", "", stripped))

    while normalized and not normalized[0]:
        normalized.pop(0)
    while normalized and not normalized[-1]:
        normalized.pop()
    return "\n".join(normalized)


def compare_models(
    baseline_model_path: Path,
    candidate_model_path: Path,
    source: SemanticModelSource | None = None,
) -> dict[str, Any]:
    source = source or LocalTmdlSource()

    baseline_snapshot, baseline_warnings = source.load_snapshot(baseline_model_path)
    candidate_snapshot, candidate_warnings = source.load_snapshot(candidate_model_path)

    warnings = [w.as_dict() for w in baseline_warnings] + [w.as_dict() for w in candidate_warnings]

    changes: list[dict[str, Any]] = []
    changes.extend(_diff_model_level(baseline_snapshot, candidate_snapshot))
    changes.extend(_diff_relationships(baseline_snapshot, candidate_snapshot))
    changes.extend(_diff_tables(baseline_snapshot, candidate_snapshot))

    changes.sort(key=_change_sort_key)

    summary = _build_summary(baseline_model_path, candidate_model_path, changes)
    return {
        "summary": summary,
        "changes": changes,
        "warnings": warnings,
    }


def _change_sort_key(change: dict[str, Any]) -> tuple[int, int, str, str]:
    level_index = LEVEL_ORDER.index(change["level"]) if change["level"] in LEVEL_ORDER else len(LEVEL_ORDER)
    change_index = CHANGE_ORDER.index(change["changeType"]) if change["changeType"] in CHANGE_ORDER else len(
        CHANGE_ORDER
    )
    table = (change.get("table") or "").casefold()
    name = (change.get("name") or "").casefold()
    return (level_index, change_index, table, name)


def _build_summary(
    baseline_model_path: Path,
    candidate_model_path: Path,
    changes: list[dict[str, Any]],
) -> dict[str, Any]:
    level_counts = {
        level: {"added": 0, "removed": 0, "changed": 0}
        for level in LEVEL_ORDER
    }
    totals = {"added": 0, "removed": 0, "changed": 0}

    for change in changes:
        level = change["level"]
        change_type = change["changeType"]
        if level in level_counts and change_type in level_counts[level]:
            level_counts[level][change_type] += 1
        if change_type in totals:
            totals[change_type] += 1

    return {
        "baselineModel": baseline_model_path.name,
        "candidateModel": candidate_model_path.name,
        "totalChanges": len(changes),
        "countsByChangeType": totals,
        "levels": level_counts,
    }


def _new_change(
    *,
    level: str,
    change_type: str,
    object_id: str,
    table: str | None,
    name: str,
    property_changes: list[dict[str, Any]],
    before: dict[str, Any] | None,
    after: dict[str, Any] | None,
) -> dict[str, Any]:
    return {
        "level": level,
        "changeType": change_type,
        "objectId": object_id,
        "table": table,
        "name": name,
        "propertyChanges": property_changes,
        "before": before,
        "after": after,
    }


def _diff_properties(before: dict[str, Any], after: dict[str, Any]) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for key in sorted(set(before.keys()) | set(after.keys()), key=str.casefold):
        before_value = before.get(key)
        after_value = after.get(key)
        if before_value != after_value:
            changes.append({
                "property": key,
                "before": before_value,
                "after": after_value,
            })
    return changes

def _diff_model_level(
    baseline: ModelSnapshot,
    candidate: ModelSnapshot,
) -> list[dict[str, Any]]:
    property_changes = _diff_properties(baseline.model_properties, candidate.model_properties)
    if not property_changes:
        return []

    return [
        _new_change(
            level="model",
            change_type="changed",
            object_id="model::semantic-model",
            table=None,
            name="Semantic Model",
            property_changes=property_changes,
            before=baseline.model_payload(),
            after=candidate.model_payload(),
        )
    ]


def _diff_relationships(
    baseline: ModelSnapshot,
    candidate: ModelSnapshot,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []

    baseline_keys = set(baseline.relationships.keys())
    candidate_keys = set(candidate.relationships.keys())

    for key in sorted(candidate_keys - baseline_keys, key=_relationship_sort_key):
        rel = candidate.relationships[key]
        object_id = (
            f"relationship::{rel.name}|{rel.from_table}[{rel.from_column}]|"
            f"{rel.to_table}[{rel.to_column}]"
        )
        changes.append(_new_change(
            level="relationship",
            change_type="added",
            object_id=object_id,
            table=None,
            name=rel.name or f"{rel.from_table}->{rel.to_table}",
            property_changes=[],
            before=None,
            after=rel.as_payload(),
        ))

    for key in sorted(baseline_keys - candidate_keys, key=_relationship_sort_key):
        rel = baseline.relationships[key]
        object_id = (
            f"relationship::{rel.name}|{rel.from_table}[{rel.from_column}]|"
            f"{rel.to_table}[{rel.to_column}]"
        )
        changes.append(_new_change(
            level="relationship",
            change_type="removed",
            object_id=object_id,
            table=None,
            name=rel.name or f"{rel.from_table}->{rel.to_table}",
            property_changes=[],
            before=rel.as_payload(),
            after=None,
        ))

    for key in sorted(baseline_keys & candidate_keys, key=_relationship_sort_key):
        before_rel = baseline.relationships[key]
        after_rel = candidate.relationships[key]
        property_changes = _diff_properties(before_rel.properties, after_rel.properties)
        if not property_changes:
            continue
        object_id = (
            f"relationship::{before_rel.name}|{before_rel.from_table}[{before_rel.from_column}]|"
            f"{before_rel.to_table}[{before_rel.to_column}]"
        )
        changes.append(_new_change(
            level="relationship",
            change_type="changed",
            object_id=object_id,
            table=None,
            name=before_rel.name or f"{before_rel.from_table}->{before_rel.to_table}",
            property_changes=property_changes,
            before=before_rel.as_payload(),
            after=after_rel.as_payload(),
        ))

    return changes


def _diff_tables(
    baseline: ModelSnapshot,
    candidate: ModelSnapshot,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []

    baseline_tables = baseline.tables
    candidate_tables = candidate.tables
    baseline_names = set(baseline_tables.keys())
    candidate_names = set(candidate_tables.keys())

    for table_name in sorted(candidate_names - baseline_names, key=str.casefold):
        table = candidate_tables[table_name]
        changes.append(_new_change(
            level="table",
            change_type="added",
            object_id=f"table::{table.name}",
            table=table.name,
            name=table.name,
            property_changes=[],
            before=None,
            after=table.as_payload(),
        ))

    for table_name in sorted(baseline_names - candidate_names, key=str.casefold):
        table = baseline_tables[table_name]
        changes.append(_new_change(
            level="table",
            change_type="removed",
            object_id=f"table::{table.name}",
            table=table.name,
            name=table.name,
            property_changes=[],
            before=table.as_payload(),
            after=None,
        ))

    for table_name in sorted(baseline_names & candidate_names, key=str.casefold):
        before_table = baseline_tables[table_name]
        after_table = candidate_tables[table_name]

        table_property_changes = _diff_properties(before_table.properties, after_table.properties)
        if table_property_changes:
            changes.append(_new_change(
                level="table",
                change_type="changed",
                object_id=f"table::{before_table.name}",
                table=before_table.name,
                name=before_table.name,
                property_changes=table_property_changes,
                before=before_table.as_payload(),
                after=after_table.as_payload(),
            ))

        changes.extend(_diff_measures(before_table, after_table))
        changes.extend(_diff_columns(before_table, after_table))

    return changes


def _diff_measures(
    baseline_table: TableSnapshot,
    candidate_table: TableSnapshot,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    baseline_measures = baseline_table.measures
    candidate_measures = candidate_table.measures
    baseline_names = set(baseline_measures.keys())
    candidate_names = set(candidate_measures.keys())

    for name in sorted(candidate_names - baseline_names, key=str.casefold):
        measure = candidate_measures[name]
        changes.append(_new_change(
            level="measure",
            change_type="added",
            object_id=f"measure::{measure.table}::{measure.name}",
            table=measure.table,
            name=measure.name,
            property_changes=[],
            before=None,
            after=measure.as_payload(),
        ))

    for name in sorted(baseline_names - candidate_names, key=str.casefold):
        measure = baseline_measures[name]
        changes.append(_new_change(
            level="measure",
            change_type="removed",
            object_id=f"measure::{measure.table}::{measure.name}",
            table=measure.table,
            name=measure.name,
            property_changes=[],
            before=measure.as_payload(),
            after=None,
        ))

    for name in sorted(baseline_names & candidate_names, key=str.casefold):
        before = baseline_measures[name]
        after = candidate_measures[name]
        property_changes = _diff_properties(before.properties, after.properties)
        if property_changes:
            changes.append(_new_change(
                level="measure",
                change_type="changed",
                object_id=f"measure::{before.table}::{before.name}",
                table=before.table,
                name=before.name,
                property_changes=property_changes,
                before=before.as_payload(),
                after=after.as_payload(),
            ))

    return changes


def _diff_columns(
    baseline_table: TableSnapshot,
    candidate_table: TableSnapshot,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    baseline_columns = baseline_table.columns
    candidate_columns = candidate_table.columns
    baseline_names = set(baseline_columns.keys())
    candidate_names = set(candidate_columns.keys())

    for name in sorted(candidate_names - baseline_names, key=str.casefold):
        column = candidate_columns[name]
        changes.append(_new_change(
            level="column",
            change_type="added",
            object_id=f"column::{column.table}::{column.name}",
            table=column.table,
            name=column.name,
            property_changes=[],
            before=None,
            after=column.as_payload(),
        ))

    for name in sorted(baseline_names - candidate_names, key=str.casefold):
        column = baseline_columns[name]
        changes.append(_new_change(
            level="column",
            change_type="removed",
            object_id=f"column::{column.table}::{column.name}",
            table=column.table,
            name=column.name,
            property_changes=[],
            before=column.as_payload(),
            after=None,
        ))

    for name in sorted(baseline_names & candidate_names, key=str.casefold):
        before = baseline_columns[name]
        after = candidate_columns[name]
        property_changes = _diff_properties(before.properties, after.properties)
        if property_changes:
            changes.append(_new_change(
                level="column",
                change_type="changed",
                object_id=f"column::{before.table}::{before.name}",
                table=before.table,
                name=before.name,
                property_changes=property_changes,
                before=before.as_payload(),
                after=after.as_payload(),
            ))

    return changes


def format_compare_json_output(results: dict[str, Any]) -> str:
    return json.dumps(results, ensure_ascii=False, indent=2)


def create_compare_xlsx_bytes(results: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary = results.get("summary", {})
    ws_summary["A1"] = "Semantic Model Compare"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Baseline"
    ws_summary["B3"] = summary.get("baselineModel", "")
    ws_summary["A4"] = "Candidate"
    ws_summary["B4"] = summary.get("candidateModel", "")
    ws_summary["A5"] = "Total Changes"
    ws_summary["B5"] = summary.get("totalChanges", 0)

    ws_summary["A7"] = "Level"
    ws_summary["B7"] = "Added"
    ws_summary["C7"] = "Removed"
    ws_summary["D7"] = "Changed"
    for col in ("A", "B", "C", "D"):
        ws_summary[f"{col}7"].font = Font(bold=True)

    row = 8
    level_counts = summary.get("levels", {})
    for level in LEVEL_ORDER:
        counts = level_counts.get(level, {})
        ws_summary.cell(row=row, column=1, value=level)
        ws_summary.cell(row=row, column=2, value=counts.get("added", 0))
        ws_summary.cell(row=row, column=3, value=counts.get("removed", 0))
        ws_summary.cell(row=row, column=4, value=counts.get("changed", 0))
        row += 1

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 12
    ws_summary.column_dimensions["D"].width = 12

    changes = results.get("changes", [])
    for level in LEVEL_ORDER:
        ws = wb.create_sheet(title=level.capitalize())
        headers = ["Change Type", "Table", "Name", "Property", "Before", "After"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = Font(bold=True)

        row_idx = 2
        level_changes = [c for c in changes if c.get("level") == level]
        for change in level_changes:
            table_name = change.get("table") or ""
            name = change.get("name") or ""
            property_changes = change.get("propertyChanges") or []
            if property_changes:
                for prop_change in property_changes:
                    ws.cell(row=row_idx, column=1, value=change.get("changeType"))
                    ws.cell(row=row_idx, column=2, value=table_name)
                    ws.cell(row=row_idx, column=3, value=name)
                    ws.cell(row=row_idx, column=4, value=prop_change.get("property", ""))
                    ws.cell(row=row_idx, column=5, value=_stringify_value(prop_change.get("before")))
                    ws.cell(row=row_idx, column=6, value=_stringify_value(prop_change.get("after")))
                    row_idx += 1
            else:
                ws.cell(row=row_idx, column=1, value=change.get("changeType"))
                ws.cell(row=row_idx, column=2, value=table_name)
                ws.cell(row=row_idx, column=3, value=name)
                ws.cell(row=row_idx, column=4, value="")
                ws.cell(row=row_idx, column=5, value=_stringify_value(change.get("before")))
                ws.cell(row=row_idx, column=6, value=_stringify_value(change.get("after")))
                row_idx += 1

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 48
        ws.column_dimensions["F"].width = 48

    import io

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)

