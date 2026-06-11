import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from semantic_model_cleaner import analyzer


FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"
PUBLIC_DEMO_DIR = (
    Path(__file__).resolve().parents[1] / "src" / "semantic_model_cleaner" / "demo_workspace"
)
PRODUCT_QA_WORKSPACE_DIR = (
    Path(__file__).resolve().parents[1] / "examples" / "product-qa-workspace"
)


def _analyze_fixture(name: str) -> dict:
    return analyzer.analyze((FIXTURES_DIR / name).resolve())


def _find_item(results: dict, table: str, name: str, item_type: str | None = None) -> dict:
    matches = [
        row for row in results["items"]
        if row["item"].table == table and row["item"].name == name
        and (item_type is None or row["item"].item_type == item_type)
    ]
    assert matches, f"Item not found: {table}[{name}]"
    if item_type is None:
        assert len(matches) == 1, f"Expected one match for {table}[{name}], found {len(matches)}"
    return matches[0]


def _write_basic_sales_model(model: Path) -> None:
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn Amount\n"
        "\tcolumn Region\n"
        "\tmeasure Revenue = SUM(Sales[Amount])\n"
        "\tmeasure 'Revenue Total' = SUM(Sales[Amount])\n",
        encoding="utf-8",
    )


def _write_page(report: Path, page_id: str = "Page1", display_name: str = "Overview") -> Path:
    page_dir = report / "definition" / "pages" / page_id
    page_dir.mkdir(parents=True)
    (page_dir / "page.json").write_text(json.dumps({"displayName": display_name}, indent=2), encoding="utf-8")
    return page_dir


def test_field_parameter_targets_are_marked_used():
    results = _analyze_fixture("field_parameter_used")

    revenue = _find_item(results, "Sales", "Revenue", "Measure")
    margin = _find_item(results, "Sales", "Margin %", "Measure")

    assert revenue["status"] == "USED (Field Parameter: Metric Parameter)"
    assert margin["status"] == "USED (Field Parameter: Metric Parameter)"
    assert [u.context for u in revenue["usages"]] == ["Field Parameter"]
    assert [u.context for u in margin["usages"]] == ["Field Parameter"]


def test_public_demo_workspace_analyzes_cleanly():
    results = analyzer.analyze(PUBLIC_DEMO_DIR.resolve())

    revenue = _find_item(results, "Sales", "Revenue", "Measure")
    margin = _find_item(results, "Sales", "Margin %", "Measure")

    assert results["summary"]["models"] == ["TestModel.SemanticModel"]
    assert results["summary"]["reports"] == ["TestReport"]
    assert revenue["status"] == "USED (Field Parameter: Metric Parameter)"
    assert margin["status"] == "USED (Field Parameter: Metric Parameter)"
    assert results["warnings"] == []


def test_product_qa_workspace_exercises_trust_workflows():
    results = analyzer.analyze(PRODUCT_QA_WORKSPACE_DIR.resolve())

    stale_margin = _find_item(results, "Sales", "Stale Margin", "Measure")
    perspective_revenue = _find_item(results, "Sales", "Perspective Revenue", "Measure")
    store_code = _find_item(results, "Store", "Store Code", "Column")
    cleanup_note = _find_item(results, "Sales", "Cleanup Note", "Column")
    report_margin = _find_item(results, "Report Metrics", "Report Margin", "Measure")

    assert results["summary"]["models"] == ["ProductQA.SemanticModel"]
    assert results["summary"]["reports"] == ["Executive"]
    assert cleanup_note["status"] == "NOT USED"
    assert cleanup_note["removal_risk"] == "Safe"
    assert perspective_revenue["status"] == "NOT USED"
    assert perspective_revenue["removal_risk"] == "Review"
    assert any(
        trigger.startswith("Unsupported Metadata: Perspectives")
        for trigger in perspective_revenue["review_triggers"]
    )
    assert store_code["status"] == "USED (RLS: Store Role)"
    assert report_margin["item"].source_kind == "report"

    stale_issue_types = {issue["issueType"] for issue in results["report_issues"]}
    assert {
        "missing_measure",
        "missing_column",
        "stale_visual_selector",
        "stale_formatting_rule",
    } <= stale_issue_types
    assert stale_margin["status"] == "NOT USED"
    assert stale_margin["removal_risk"] == "Safe"
    assert len(stale_margin["stale_usages"]) >= 2


def test_split_tmdl_item_records_source_file(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)
    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn Amount\n"
        "\t\tdataType: decimal\n",
        encoding="utf-8",
    )
    split_file = tables_dir / "Sales.Measures.tmdl"
    split_file.write_text(
        "table Sales\n"
        "\tmeasure Revenue = SUM(Sales[Amount])\n"
        "\tmeasure Cost = 1\n",
        encoding="utf-8",
    )

    results = analyzer.analyze(workspace.resolve())

    revenue = _find_item(results, "Sales", "Revenue", "Measure")
    payload = json.loads(analyzer.format_json_output(results))
    payload_revenue = next(item for item in payload["items"] if item["table"] == "Sales" and item["name"] == "Revenue")
    assert revenue["item"].source_file == str(split_file)
    assert payload_revenue["sourceFile"] == str(split_file)


def test_source_root_discovery_ignores_internal_fixture_and_hidden_reports(tmp_path):
    source_root = tmp_path / "Semantic-Model-Cleaner"
    (source_root / "src" / "semantic_model_cleaner").mkdir(parents=True)
    (source_root / "pyproject.toml").write_text("[project]\nname = 'semantic-model-cleaner'\n", encoding="utf-8")

    internal_report = source_root / "tests" / "fixtures" / "sample" / "Reports" / "TestReport.Report"
    example_report = source_root / "examples" / "product-qa-workspace" / "Reports" / "TestReport.Report"
    demo_report = (
        source_root / "src" / "semantic_model_cleaner" / "demo_workspace" / "Reports" / "TestReport.Report"
    )
    hidden_report = source_root / ".claude" / "worktrees" / "old" / "Reports" / "TestReport.Report"
    real_report = source_root / "Workspace" / "Reports" / "RealReport.Report"
    for report in (internal_report, example_report, demo_report, hidden_report, real_report):
        report.mkdir(parents=True)

    reports = analyzer.discover_reports([source_root])

    assert reports == [real_report.resolve()]


def test_direct_artifact_discovery_still_allows_explicit_fixture_paths(tmp_path):
    source_root = tmp_path / "Semantic-Model-Cleaner"
    (source_root / "src" / "semantic_model_cleaner").mkdir(parents=True)
    (source_root / "pyproject.toml").write_text("[project]\nname = 'semantic-model-cleaner'\n", encoding="utf-8")
    explicit_report = source_root / "tests" / "fixtures" / "sample" / "Reports" / "TestReport.Report"
    explicit_report.mkdir(parents=True)

    reports = analyzer.discover_reports([explicit_report])

    assert reports == [explicit_report.resolve()]


def test_unused_field_parameter_table_does_not_promote_targets():
    results = _analyze_fixture("field_parameter_unused")

    revenue = _find_item(results, "Sales", "Revenue", "Measure")
    margin = _find_item(results, "Sales", "Margin %", "Measure")

    assert revenue["status"] == "USED"
    assert margin["status"] == "NOT USED"


def test_ambiguous_nameof_target_emits_warning():
    results = _analyze_fixture("nameof_ambiguous_target")

    warnings = results["warnings"]
    assert any(w["code"] == "AMBIGUOUS_NAMEOF_TARGET" for w in warnings)

    measure = _find_item(results, "Sales", "Revenue", "Measure")
    column = _find_item(results, "Sales", "Revenue", "Column")
    assert measure["status"] == "NOT USED"
    assert column["status"] == "NOT USED"


def test_json_output_includes_warnings():
    results = _analyze_fixture("nameof_ambiguous_target")

    payload = json.loads(analyzer.format_json_output(results))

    assert "warnings" in payload
    assert isinstance(payload["warnings"], list)
    assert "reportIssues" in payload
    assert isinstance(payload["reportIssues"], list)


def test_report_issues_include_missing_visual_fields_filters_and_suggestions(tmp_path):
    model = tmp_path / "Models" / "Sales.SemanticModel"
    report = tmp_path / "Reports" / "Executive.Report"
    _write_basic_sales_model(model)
    page_dir = _write_page(report)
    visual_dir = page_dir / "visuals" / "Visual1"
    visual_dir.mkdir(parents=True)
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "isHidden": True,
                "position": {"x": 120, "y": 240, "width": 320, "height": 180},
                "visual": {
                    "visualType": "tableEx",
                    "query": {
                        "queryState": {
                            "Values": {
                                "projections": [
                                    {
                                        "field": {
                                            "Measure": {
                                                "Property": "Revenue Totl",
                                                "Expression": {"SourceRef": {"Entity": "Sales"}},
                                            }
                                        },
                                        "queryRef": "Sales.Revenue Totl",
                                    },
                                    {
                                        "field": {
                                            "Column": {
                                                "Property": "Region Missing",
                                                "Expression": {"SourceRef": {"Entity": "Sales"}},
                                            }
                                        },
                                        "queryRef": "Sales.Region Missing",
                                    },
                                ]
                            }
                        }
                    },
                    "filterConfig": {
                        "filters": [
                            {
                                "field": {
                                    "Column": {
                                        "Property": "Inactive Only",
                                        "Expression": {"SourceRef": {"Entity": "Legacy Sales"}},
                                    }
                                }
                            },
                            {
                                "field": {
                                    "Column": {
                                        "Property": "Amount",
                                        "Expression": {"SourceRef": {"Entity": "Legacy Sales"}},
                                    }
                                },
                                "filter": {
                                    "From": [{"Name": "l", "Entity": "Legacy Sales", "Type": 0}],
                                    "Where": [
                                        {
                                            "Condition": {
                                                "In": {
                                                    "Expressions": [
                                                        {
                                                            "Column": {
                                                                "Expression": {"SourceRef": {"Source": "l"}},
                                                                "Property": "Amount",
                                                            }
                                                        }
                                                    ]
                                                }
                                            }
                                        }
                                    ],
                                },
                            }
                        ]
                    },
                    "objects": {
                        "values": [
                            {
                                "properties": {
                                    "fontColor": {
                                        "solid": {
                                            "color": {
                                                "expr": {
                                                    "FillRule": {
                                                        "Input": {
                                                            "Aggregation": {
                                                                "Expression": {
                                                                    "Column": {
                                                                        "Expression": {"SourceRef": {"Entity": "Legacy Sales"}},
                                                                        "Property": "Old Color Driver",
                                                                    }
                                                                },
                                                                "Function": 0,
                                                            }
                                                        },
                                                        "FillRule": {
                                                            "linearGradient2": {
                                                                "min": {"color": {"Literal": {"Value": "'#c1272d'"}}},
                                                                "max": {"color": {"Literal": {"Value": "'#7faf23'"}}},
                                                            }
                                                        },
                                                    }
                                                }
                                            }
                                        }
                                    }
                                },
                                "selector": {
                                    "data": [{"dataViewWildcard": {"matchingOption": 1}}],
                                    "metadata": "Sum(Legacy Sales.Old Color Driver)",
                                },
                            }
                        ],
                        "dataPoint": [
                            {
                                "properties": {"fill": {"solid": {"color": {"expr": {"Literal": {"Value": "'#228B22'"}}}}}},
                                "selector": {
                                    "data": [
                                        {
                                            "scopeId": {
                                                "Comparison": {
                                                    "Left": {
                                                        "Column": {
                                                            "Property": "Old Format Selector",
                                                            "Expression": {"SourceRef": {"Entity": "Legacy Sales"}},
                                                        }
                                                    },
                                                    "Right": {"Literal": {"Value": "'Old'"}},
                                                }
                                            }
                                        }
                                    ]
                                },
                            }
                        ]
                    },
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(tmp_path.resolve())
    issue_types = {issue["issueType"] for issue in results["report_issues"]}

    assert {"missing_measure", "missing_column", "missing_table"} <= issue_types
    inactive_issue = next(issue for issue in results["report_issues"] if issue["name"] == "Inactive Only")
    assert inactive_issue["issueType"] == "inactive_visual_filter_reference"
    assert inactive_issue["severity"] == "warning"
    assert inactive_issue["context"] == "Inactive Filter"
    assert inactive_issue["sourcePath"].endswith("visual.filterConfig.filters.[0].field.Column")
    old_format_selector = next(issue for issue in results["report_issues"] if issue["name"] == "Old Format Selector")
    assert old_format_selector["issueType"] == "stale_visual_selector"
    assert old_format_selector["severity"] == "warning"
    assert old_format_selector["context"] == "Stale Formatting"
    assert old_format_selector["staleKind"] == "visual_formatting_selector_entry"
    assert old_format_selector["suggestions"][0]["action"] == "clean_stale"
    stale_formatting_rule = next(issue for issue in results["report_issues"] if issue["name"] == "Old Color Driver")
    assert stale_formatting_rule["issueType"] == "stale_formatting_rule"
    assert stale_formatting_rule["severity"] == "warning"
    assert stale_formatting_rule["context"] == "Stale Formatting Rule"
    assert stale_formatting_rule["staleKind"] == "formatting_rule_reference"
    assert stale_formatting_rule["suggestions"][0]["action"] == "clean_stale"
    assert "selector.metadata" not in stale_formatting_rule["sourcePath"]
    missing_measure = next(issue for issue in results["report_issues"] if issue["issueType"] == "missing_measure")
    assert missing_measure["report"] == "Executive"
    assert missing_measure["page"] == "Overview"
    assert missing_measure["visualId"] == "Visual1"
    assert missing_measure["context"] == "Values"
    assert missing_measure["visualHidden"] is True
    assert missing_measure["pageHidden"] is False
    assert missing_measure["visualX"] == 120
    assert missing_measure["visualY"] == 240
    assert missing_measure["suggestions"][0]["table"] == "Sales"
    assert missing_measure["suggestions"][0]["name"] == "Revenue Total"
    assert missing_measure["suggestions"][0]["action"] == "suggest_replace"

    revenue_total = _find_item(results, "Sales", "Revenue Total", "Measure")
    assert revenue_total["status"] == "NOT USED"


def test_report_issue_visual_labels_distinguish_table_visuals(tmp_path):
    model = tmp_path / "Models" / "Sales.SemanticModel"
    report = tmp_path / "Reports" / "Executive.Report"
    _write_basic_sales_model(model)
    page_dir = _write_page(report)

    first_visual = page_dir / "visuals" / "WhatsNewVisual"
    second_visual = page_dir / "visuals" / "AlertsVisual"
    titled_visual = page_dir / "visuals" / "TitledVisual"
    first_visual.mkdir(parents=True)
    second_visual.mkdir(parents=True)
    titled_visual.mkdir(parents=True)

    def write_visual(path: Path, table: str, column: str, title_text: str = "") -> None:
        visual = {
            "visual": {
                "visualType": "tableEx",
                "query": {
                    "queryState": {
                        "Values": {
                            "projections": [
                                {
                                    "field": {
                                        "Column": {
                                            "Property": column,
                                            "Expression": {"SourceRef": {"Entity": table}},
                                        }
                                    },
                                    "queryRef": f"{table}.{column}",
                                }
                            ]
                        }
                    }
                },
            }
        }
        if title_text:
            visual["visual"]["visualContainerObjects"] = {
                "title": [
                    {
                        "properties": {
                            "text": {"expr": {"Literal": {"Value": f"'{title_text}'"}}},
                            "show": {"expr": {"Literal": {"Value": "false"}}},
                        }
                    }
                ]
            }
        path.write_text(json.dumps(visual, indent=2), encoding="utf-8")

    write_visual(first_visual / "visual.json", "WhatsNew", "Release Date")
    write_visual(second_visual / "visual.json", "Alerts", "Alert Text")
    write_visual(titled_visual / "visual.json", "MissingTitleTable", "Name", "Release Notes")

    results = analyzer.analyze(tmp_path.resolve())
    issues_by_visual = {issue["visualId"]: issue for issue in results["report_issues"]}

    assert issues_by_visual["WhatsNewVisual"]["visualTitle"] == ""
    assert issues_by_visual["AlertsVisual"]["visualTitle"] == ""
    assert issues_by_visual["TitledVisual"]["visualTitle"] == "Release Notes"


def test_report_issues_include_page_bookmark_definition_and_invalid_json(tmp_path):
    model = tmp_path / "Models" / "Sales.SemanticModel"
    report = tmp_path / "Reports" / "Executive.Report"
    _write_basic_sales_model(model)
    page_dir = _write_page(report)
    visual_dir = page_dir / "visuals" / "Visual1"
    visual_dir.mkdir(parents=True)
    bookmarks_dir = report / "definition" / "bookmarks"
    bookmarks_dir.mkdir(parents=True)
    (visual_dir / "visual.json").write_text(
        json.dumps({"visual": {"visualType": "card", "query": {"queryState": {}}}}, indent=2),
        encoding="utf-8",
    )
    (page_dir / "page.json").write_text(
        json.dumps(
            {
                "displayName": "Overview",
                "visibility": "HiddenInViewMode",
                "filters": [
                    {
                        "field": {
                            "Column": {
                                "Property": "Missing Page Filter",
                                "Expression": {"SourceRef": {"Entity": "Sales"}},
                            }
                        }
                    }
                ],
                "filterConfig": {
                    "filters": [
                        {
                            "field": {
                                "Column": {
                                    "Property": "Amount",
                                    "Expression": {"SourceRef": {"Entity": "Sales"}},
                                }
                            },
                            "filter": {
                                "From": [{"Name": "s", "Entity": "Sales", "Type": 0}],
                                "Where": [
                                    {
                                        "Condition": {
                                            "In": {
                                                "Expressions": [
                                                    {
                                                        "Column": {
                                                            "Expression": {"SourceRef": {"Source": "s"}},
                                                            "Property": "Missing Pane Filter",
                                                        }
                                                    }
                                                ]
                                            }
                                        }
                                    }
                                ],
                            },
                        }
                    ]
                },
                "drillthrough": [
                    {
                        "field": {
                            "Column": {
                                "Property": "Missing Drill",
                                "Expression": {"SourceRef": {"Entity": "Sales"}},
                            }
                        }
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (report / "definition" / "report.json").write_text(
        json.dumps(
            {
                "filterConfig": {
                    "filters": [
                        {
                            "field": {
                                "Column": {
                                    "Property": "Missing All Pages Filter",
                                    "Expression": {"SourceRef": {"Entity": "Sales"}},
                                }
                            }
                        }
                    ]
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (bookmarks_dir / "Broken.bookmark.json").write_text(
        json.dumps(
            {
                "displayName": "Broken bookmark",
                "explorationState": {
                    "sections": {
                        "Page1": {
                            "visualContainers": {
                                "Visual1": {
                                    "singleVisual": {
                                        "visualType": "card",
                                        "projections": {
                                            "Values": [
                                                {
                                                    "field": {
                                                        "Measure": {
                                                            "Property": "Missing Bookmark Measure",
                                                            "Expression": {"SourceRef": {"Entity": "Sales"}},
                                                        }
                                                    }
                                                }
                                            ]
                                        },
                                    }
                                },
                                "DeletedVisual": {
                                    "singleVisual": {
                                        "visualType": "card",
                                        "projections": {}
                                    }
                                }
                            }
                        }
                    }
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    custom_dir = report / "definition" / "custom"
    custom_dir.mkdir(parents=True)
    (custom_dir / "extra.json").write_text(
        json.dumps(
            {
                "field": {
                    "Column": {
                        "Property": "Missing Definition",
                        "Expression": {"SourceRef": {"Entity": "Sales"}},
                    }
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (custom_dir / "bad.json").write_text("{ bad json", encoding="utf-8")

    results = analyzer.analyze(tmp_path.resolve())
    issues = results["report_issues"]

    assert any(issue["issueType"] == "missing_column" and issue["context"] == "Page Filter" for issue in issues)
    assert any(
        issue["issueType"] == "missing_column"
        and issue["context"] == "Filters pane (page)"
        and issue["pageHidden"] is True
        and issue["visualHidden"] is False
        for issue in issues
    )
    pane_issue = next(
        issue for issue in issues
        if issue["issueType"] == "missing_column"
        and issue["context"] == "Filters pane (page)"
        and issue["name"] == "Missing Pane Filter"
    )
    assert pane_issue["sourcePath"].endswith("filter.Where.[0].Condition.In.Expressions.[0].Column")
    assert any(
        issue["issueType"] == "missing_column"
        and issue["context"] == "Filters pane (all pages)"
        and issue["artifactKind"] == "Report"
        and issue["artifactPath"] == "definition/report.json"
        for issue in issues
    )
    assert any(issue["issueType"] == "missing_column" and issue["context"] == "Drillthrough" for issue in issues)
    assert any(
        issue["issueType"] == "missing_measure"
        and issue["artifactKind"] == "Bookmark"
        and issue["context"] == "Bookmark: Broken bookmark"
        for issue in issues
    )
    orphan_bookmark = next(issue for issue in issues if issue["issueType"] == "orphan_bookmark_visual_state")
    assert orphan_bookmark["visualId"] == "DeletedVisual"
    assert orphan_bookmark["artifactKind"] == "Bookmark"
    assert orphan_bookmark["sourcePath"].endswith("visualContainers.DeletedVisual")
    assert any(issue["issueType"] == "missing_column" and issue["artifactKind"] == "Definition JSON" for issue in issues)
    assert any(issue["issueType"] == "invalid_report_json" and issue["artifactPath"] == "definition/custom/bad.json" for issue in issues)


def test_report_issues_respect_selected_reports_only(tmp_path):
    model = tmp_path / "Models" / "Sales.SemanticModel"
    selected_report = tmp_path / "Reports" / "Selected.Report"
    unselected_report = tmp_path / "Reports" / "Unselected.Report"
    _write_basic_sales_model(model)
    for report, missing_name in ((selected_report, "Selected Missing"), (unselected_report, "Unselected Missing")):
        page_dir = _write_page(report)
        visual_dir = page_dir / "visuals" / "Visual1"
        visual_dir.mkdir(parents=True)
        (visual_dir / "visual.json").write_text(
            json.dumps(
                {
                    "visual": {
                        "visualType": "card",
                        "query": {
                            "queryState": {
                                "Values": {
                                    "projections": [
                                        {
                                            "field": {
                                                "Measure": {
                                                    "Property": missing_name,
                                                    "Expression": {"SourceRef": {"Entity": "Sales"}},
                                                }
                                            }
                                        }
                                    ]
                                }
                            }
                        },
                    }
                },
                indent=2,
            ),
            encoding="utf-8",
        )

    results = analyzer.analyze(
        tmp_path.resolve(),
        model_paths=[model],
        report_paths=[selected_report],
    )

    assert {issue["report"] for issue in results["report_issues"]} == {"Selected"}
    assert any(issue["name"] == "Selected Missing" for issue in results["report_issues"])
    assert all(issue["name"] != "Unselected Missing" for issue in results["report_issues"])


def test_rls_table_permission_marks_referenced_column_used(tmp_path):
    model = tmp_path / "Models" / "TestModel.SemanticModel"
    report = tmp_path / "Reports" / "TestReport.Report"
    tables_dir = model / "definition" / "tables"
    roles_dir = model / "definition" / "roles"
    tables_dir.mkdir(parents=True)
    roles_dir.mkdir(parents=True)
    (report / "definition").mkdir(parents=True)

    (tables_dir / "Store.tmdl").write_text(
        "table Store\n"
        "\tcolumn 'Store Code'\n"
        "\t\tdataType: int64\n"
        "\tcolumn Name\n"
        "\t\tdataType: string\n",
        encoding="utf-8",
    )
    (roles_dir / "Store Role.tmdl").write_text(
        "role 'Store Role'\n"
        "\tmodelPermission: read\n"
        "\ttablePermission Store = 'Store'[Store Code] IN {1, 10}\n",
        encoding="utf-8",
    )

    results = analyzer.analyze(tmp_path.resolve())

    store_code = _find_item(results, "Store", "Store Code", "Column")
    store_name = _find_item(results, "Store", "Name", "Column")

    assert store_code["status"] == "USED (RLS: Store Role)"
    assert store_name["status"] == "NOT USED"
    assert results["summary"]["used_rls"] == 1


def test_multiline_rls_table_permission_uses_declared_table_for_unqualified_columns(tmp_path):
    model = tmp_path / "Models" / "TestModel.SemanticModel"
    report = tmp_path / "Reports" / "TestReport.Report"
    tables_dir = model / "definition" / "tables"
    roles_dir = model / "definition" / "roles"
    tables_dir.mkdir(parents=True)
    roles_dir.mkdir(parents=True)
    (report / "definition").mkdir(parents=True)

    (tables_dir / "Store.tmdl").write_text(
        "table Store\n"
        "\tcolumn 'Store Code'\n"
        "\t\tdataType: int64\n"
        "\tcolumn Region\n"
        "\t\tdataType: string\n"
        "\tcolumn Name\n"
        "\t\tdataType: string\n",
        encoding="utf-8",
    )
    (roles_dir / "Store Role.tmdl").write_text(
        "role 'Store Role'\n"
        "\tmodelPermission: read\n"
        "\ttablePermission Store =\n"
        "\t\t```\n"
        "\t\t'Store'[Store Code] IN {1, 10}\n"
        "\t\t    && [Region] = \"North\"\n"
        "\t\t```\n",
        encoding="utf-8",
    )

    results = analyzer.analyze(tmp_path.resolve())

    store_code = _find_item(results, "Store", "Store Code", "Column")
    region = _find_item(results, "Store", "Region", "Column")
    store_name = _find_item(results, "Store", "Name", "Column")

    assert store_code["status"] == "USED (RLS: Store Role)"
    assert region["status"] == "USED (RLS: Store Role)"
    assert store_name["status"] == "NOT USED"
    assert results["summary"]["used_rls"] == 2


def test_stale_formatting_selectors_do_not_count_as_live_usage(tmp_path):
    model = tmp_path / "Models" / "TestModel.SemanticModel"
    report = tmp_path / "Reports" / "TestReport.Report"
    tables_dir = model / "definition" / "tables"
    visual_dir = report / "definition" / "pages" / "Page 1" / "visuals" / "visual1"
    tables_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)

    (tables_dir / "_Measures.tmdl").write_text(
        "table _Measures\n"
        "\tmeasure 'PL_LINE Budget' = 1\n"
        "\tmeasure 'PL_LINE LY' = 1\n"
        "\tmeasure 'Revenue LY' = 1\n",
        encoding="utf-8",
    )
    (report / "definition" / "pages" / "Page 1" / "page.json").write_text(
        json.dumps({"displayName": "FINANCE"}, indent=2),
        encoding="utf-8",
    )
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "visual": {
                    "visualType": "lineClusteredColumnComboChart",
                    "query": {
                        "queryState": {
                            "Y": {
                                "projections": [
                                    {
                                        "field": {
                                            "Measure": {
                                                "Property": "PL_LINE Budget",
                                                "Expression": {"SourceRef": {"Entity": "_Measures"}},
                                            }
                                        },
                                        "queryRef": "_Measures.PL_LINE Budget",
                                    }
                                ]
                            }
                        }
                    },
                    "objects": {
                        "labels": [
                            {
                                "properties": {"show": {"expr": {"Literal": {"Value": "true"}}}},
                                "selector": {"metadata": "_Measures.Revenue LY"},
                            },
                            {
                                "properties": {"show": {"expr": {"Literal": {"Value": "true"}}}},
                                "selector": {"metadata": "_Measures.PL_LINE Budget"},
                            },
                        ],
                        "referenceLabel": [
                            {
                                "properties": {
                                    "value": {
                                        "expr": {
                                            "Measure": {
                                                "Property": "Revenue LY",
                                                "Expression": {"SourceRef": {"Entity": "_Measures"}},
                                            }
                                        }
                                    }
                                },
                                "selector": {"metadata": "_Measures.Net Sales - Actuals"},
                            },
                            {
                                "properties": {
                                    "value": {
                                        "expr": {
                                            "Measure": {
                                                "Property": "PL_LINE LY",
                                                "Expression": {"SourceRef": {"Entity": "_Measures"}},
                                            }
                                        }
                                    }
                                },
                                "selector": {"metadata": "_Measures.PL_LINE Budget"},
                            },
                        ],
                    },
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(tmp_path.resolve())

    live_measure = _find_item(results, "_Measures", "PL_LINE Budget", "Measure")
    live_reference_measure = _find_item(results, "_Measures", "PL_LINE LY", "Measure")
    stale_measure = _find_item(results, "_Measures", "Revenue LY", "Measure")

    assert live_measure["status"] == "USED"
    assert len(live_measure["usages"]) == 3
    assert live_measure["stale_usages"] == []

    assert live_reference_measure["status"] == "USED"
    assert len(live_reference_measure["usages"]) == 1
    assert live_reference_measure["stale_usages"] == []

    assert stale_measure["status"] == "NOT USED"
    assert stale_measure["usages"] == []
    assert len(stale_measure["stale_usages"]) == 2
    assert {usage.selector_value for usage in stale_measure["stale_usages"]} == {
        "_Measures.Revenue LY",
        "_Measures.Net Sales - Actuals",
    }
    assert all(usage.context == "Stale Formatting" for usage in stale_measure["stale_usages"])
    stale_issues = [issue for issue in results["report_issues"] if issue["issueType"] == "stale_visual_selector"]
    assert len(stale_issues) >= 2
    assert {
        "_Measures.Revenue LY",
        "_Measures.Net Sales - Actuals",
    } <= {issue["selectorValue"] for issue in stale_issues}
    assert all(issue["suggestions"][0]["action"] == "clean_stale" for issue in stale_issues)


def test_source_ref_alias_counts_as_live_report_usage(tmp_path):
    model = tmp_path / "Models" / "TestModel.SemanticModel"
    report = tmp_path / "Reports" / "TestReport.Report"
    tables_dir = model / "definition" / "tables"
    visual_dir = report / "definition" / "pages" / "Page1" / "visuals" / "Visual1"
    tables_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn Amount\n"
        "\t\tdataType: decimal\n"
        "\tmeasure Revenue = SUM(Sales[Amount])\n",
        encoding="utf-8",
    )
    (report / "definition" / "pages" / "Page1" / "page.json").write_text(
        json.dumps({"displayName": "Overview"}, indent=2),
        encoding="utf-8",
    )
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "visual": {
                    "query": {
                        "SemanticQueryDataShapeCommand": {
                            "Query": {
                                "Version": 2,
                                "From": [{"Name": "s", "Entity": "Sales", "Type": 0}],
                                "Select": [
                                    {
                                        "Measure": {
                                            "Expression": {"SourceRef": {"Source": "s"}},
                                            "Property": "Revenue",
                                        }
                                    }
                                ],
                            }
                        },
                        "queryState": {
                            "Y": {
                                "projections": [
                                    {
                                        "field": {
                                            "Measure": {
                                                "Expression": {"SourceRef": {"Source": "s"}},
                                                "Property": "Revenue",
                                            }
                                        },
                                        "queryRef": "s.Revenue",
                                    }
                                ]
                            }
                        },
                    }
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(tmp_path.resolve())
    revenue = _find_item(results, "Sales", "Revenue", "Measure")

    assert revenue["status"] == "USED"
    assert revenue["usages"][0].source_path.endswith("Measure")


def test_stale_bookmark_projections_do_not_count_as_live_usage(tmp_path):
    model = tmp_path / "Models" / "TestModel.SemanticModel"
    report = tmp_path / "Reports" / "TestReport.Report"
    tables_dir = model / "definition" / "tables"
    visual_dir = report / "definition" / "pages" / "Page1" / "visuals" / "Visual1"
    bookmarks_dir = report / "definition" / "bookmarks"
    tables_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)
    bookmarks_dir.mkdir(parents=True)

    (tables_dir / "_Measures.tmdl").write_text(
        "table _Measures\n"
        "\tmeasure 'PL_LINE Budget' = 1\n"
        "\tmeasure 'PL_LINE LY' = 1\n"
        "\tmeasure 'A&P - LY' = 1\n",
        encoding="utf-8",
    )
    (report / "definition" / "pages" / "Page1" / "page.json").write_text(
        json.dumps({"displayName": "MARKETING"}, indent=2),
        encoding="utf-8",
    )
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "name": "Visual1",
                "visual": {
                    "visualType": "lineClusteredColumnComboChart",
                    "query": {
                        "queryState": {
                            "Y2": {
                                "projections": [
                                    {"queryRef": "_Measures.PL_LINE Budget"},
                                    {"queryRef": "_Measures.PL_LINE LY"},
                                ],
                                "fieldParameters": [
                                    {"parameterExpr": {"Column": {"Expression": {"SourceRef": {"Entity": "FP_KPI_Finance"}}, "Property": "FP_KPI_Finance"}}}
                                ],
                            }
                        }
                    },
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (bookmarks_dir / "bookmark1.bookmark.json").write_text(
        json.dumps(
            {
                "displayName": "A&P Spend",
                "name": "bookmark1",
                "options": {
                    "applyOnlyToTargetVisuals": True,
                    "targetVisualNames": ["Visual1"],
                },
                "explorationState": {
                    "activeSection": "Page1",
                    "sections": {
                        "Page1": {
                            "visualContainers": {
                                "Visual1": {
                                    "singleVisual": {
                                        "visualType": "lineClusteredColumnComboChart",
                                        "projections": {
                                            "Y2": [
                                                {
                                                    "Measure": {
                                                        "Expression": {"SourceRef": {"Entity": "_Measures"}},
                                                        "Property": "A&P - LY",
                                                    }
                                                }
                                            ]
                                        },
                                        "parameters": {
                                            "Y2": [
                                                {
                                                    "expr": {
                                                        "Column": {
                                                            "Expression": {"SourceRef": {"Entity": "FP_KPI_Finance"}},
                                                            "Property": "FP_KPI_Finance",
                                                        }
                                                    }
                                                }
                                            ]
                                        },
                                    }
                                }
                            }
                        }
                    }
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(tmp_path.resolve())
    stale_measure = _find_item(results, "_Measures", "A&P - LY", "Measure")

    assert stale_measure["status"] == "NOT USED"
    assert stale_measure["usages"] == []
    assert len(stale_measure["stale_usages"]) == 1
    assert stale_measure["stale_usages"][0].context == "Stale Bookmark: A&P Spend"
    assert stale_measure["stale_usages"][0].stale_kind == "bookmark_projection_entry"
    assert stale_measure["stale_usages"][0].selector_value == "Y2"
    bookmark_issues = [issue for issue in results["report_issues"] if issue["issueType"] == "stale_bookmark_projection"]
    assert len(bookmark_issues) == 1
    assert bookmark_issues[0]["staleKind"] == "bookmark_projection_entry"
    assert bookmark_issues[0]["selectorValue"] == "Y2"


def test_unused_hidden_column_includes_review_triggers(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "Date.tmdl").write_text(
        "table Date\n"
        "\tcolumn DateKey\n"
        "\t\tisHidden\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())
    date_key = _find_item(results, "Date", "DateKey", "Column")

    assert date_key["status"] == "NOT USED"
    assert date_key["removal_risk"] == "Review"
    assert date_key["review_triggers"] == ["Item is hidden"]

    payload = json.loads(analyzer.format_json_output(results))
    payload_item = next(item for item in payload["items"] if item["table"] == "Date" and item["name"] == "DateKey")
    assert payload_item["reviewTriggers"] == ["Item is hidden"]


def test_unused_item_is_review_when_perspective_may_hide_dependency(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    perspectives_dir = model / "definition" / "perspectives"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    perspectives_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure Revenue = SUM(Sales[Amount])\n"
        "\tcolumn Amount\n",
        encoding="utf-8",
    )
    (perspectives_dir / "Executive.tmdl").write_text(
        "perspective Executive\n"
        "\tperspectiveTable Sales\n"
        "\t\tperspectiveMeasure Revenue\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())
    revenue = _find_item(results, "Sales", "Revenue", "Measure")

    assert revenue["status"] == "NOT USED"
    assert revenue["removal_risk"] == "Review"
    assert revenue["review_triggers"] == [
        (
            "Unsupported Metadata: Perspectives in definition/perspectives/Executive.tmdl "
            "can reference Sales[Revenue]. Hidden dependency: perspective membership may "
            "keep this field available outside scanned report visuals. User harm: deleting "
            "it could break curated perspective views, Excel connections, or downstream "
            "tools that rely on the perspective."
        )
    ]

    payload = json.loads(analyzer.format_json_output(results))
    payload_item = next(item for item in payload["items"] if item["table"] == "Sales" and item["name"] == "Revenue")
    assert payload_item["removalRisk"] == "Review"
    assert payload_item["reviewTriggers"] == revenue["review_triggers"]

    workbook = load_workbook(BytesIO(analyzer.create_xlsx_bytes(results)))
    detail_sheet = workbook["Details"]
    headers = [cell.value for cell in detail_sheet[1]]
    trigger_col = headers.index("Review Triggers") + 1
    revenue_row = next(row for row in detail_sheet.iter_rows(min_row=2) if row[2].value == "Revenue")
    assert revenue_row[trigger_col - 1].value == revenue["review_triggers"][0]


def test_unsupported_metadata_areas_downgrade_safe_items_to_review(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    cultures_dir = model / "definition" / "cultures"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    cultures_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure CalcGroupTarget = 1\n"
        "\tmeasure KpiTarget = 1\n"
        "\tmeasure DetailRowsTarget = 1\n"
        "\tmeasure FormatTarget = 1\n"
        "\tmeasure CoverageTarget = 1\n"
        "\tmeasure SecondaryTarget = 1\n"
        "\tmeasure TranslationTarget = 1\n",
        encoding="utf-8",
    )
    (tables_dir / "Time Intelligence.tmdl").write_text(
        "table 'Time Intelligence'\n"
        "\tcalculationGroup\n"
        "\t\tcalculationItem Current = SELECTEDMEASURE()\n"
        "\t\tcalculationItem Prior = CALCULATE(SELECTEDMEASURE(), 'Sales'[CalcGroupTarget])\n",
        encoding="utf-8",
    )
    (tables_dir / "KpiMetadata.tmdl").write_text(
        "table KpiMetadata\n"
        "\tmeasure KpiCarrier = 1\n"
        "\t\tkpi\n"
        "\t\t\ttargetExpression = 'Sales'[KpiTarget]\n",
        encoding="utf-8",
    )
    (tables_dir / "DetailRowsMetadata.tmdl").write_text(
        "table DetailRowsMetadata\n"
        "\tmeasure DetailRowsCarrier = 1\n"
        "\t\tdetailRowsDefinition = SELECTCOLUMNS(Sales, \"x\", 'Sales'[DetailRowsTarget])\n",
        encoding="utf-8",
    )
    (tables_dir / "FormatMetadata.tmdl").write_text(
        "table FormatMetadata\n"
        "\tmeasure FormatCarrier = 1\n"
        "\t\tformatStringDefinition = 'Sales'[FormatTarget]\n",
        encoding="utf-8",
    )
    (tables_dir / "CoverageMetadata.tmdl").write_text(
        "table CoverageMetadata\n"
        "\tmeasure CoverageCarrier = 1\n"
        "\t\tdataCoverageDefinition = 'Sales'[CoverageTarget] > 0\n",
        encoding="utf-8",
    )
    (tables_dir / "SecondaryMetadata.tmdl").write_text(
        "table SecondaryMetadata\n"
        "\tmeasure SecondaryCarrier = 1\n"
        "\t\tsecondaryExpression = 'Sales'[SecondaryTarget]\n",
        encoding="utf-8",
    )
    (cultures_dir / "en-US.tmdl").write_text(
        "culture en-US\n"
        "\tlinguisticMetadata = 'Sales'[TranslationTarget]\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())

    expectations = {
        "CalcGroupTarget": "Calculation Groups",
        "KpiTarget": "KPI expressions",
        "DetailRowsTarget": "Detail rows",
        "FormatTarget": "Format string definitions",
        "CoverageTarget": "Data coverage definitions",
        "SecondaryTarget": "Secondary expressions",
        "TranslationTarget": "Cultures/translations",
    }
    for measure_name, area in expectations.items():
        item = _find_item(results, "Sales", measure_name, "Measure")
        assert item["status"] == "NOT USED"
        assert item["removal_risk"] == "Review"
        assert any(area in trigger for trigger in item["review_triggers"])
        assert any("Hidden dependency:" in trigger for trigger in item["review_triggers"])
        assert any("User harm:" in trigger for trigger in item["review_triggers"])


def test_ordinary_unused_item_remains_safe_without_unsupported_metadata(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure Revenue = 1\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())
    revenue = _find_item(results, "Sales", "Revenue", "Measure")

    assert revenue["status"] == "NOT USED"
    assert revenue["removal_risk"] == "Safe"
    assert revenue["review_triggers"] == []


def test_unsupported_tmdl_metadata_does_not_review_unrelated_file_refs(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure OrdinaryTarget = 1\n"
        "\tmeasure FormatTarget = 1\n",
        encoding="utf-8",
    )
    (tables_dir / "MetadataCarrier.tmdl").write_text(
        "table MetadataCarrier\n"
        "\tmeasure OrdinaryCarrier = 'Sales'[OrdinaryTarget]\n"
        "\tmeasure FormatCarrier = 1\n"
        "\t\tformatStringDefinition = 'Sales'[FormatTarget]\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())
    ordinary_target = _find_item(results, "Sales", "OrdinaryTarget", "Measure")
    format_target = _find_item(results, "Sales", "FormatTarget", "Measure")

    assert ordinary_target["status"] == "NOT USED"
    assert ordinary_target["removal_risk"] == "Caution"
    assert ordinary_target["review_triggers"] == []
    assert format_target["status"] == "NOT USED"
    assert format_target["removal_risk"] == "Review"
    assert any("Format string definitions" in trigger for trigger in format_target["review_triggers"])


def test_report_extension_measures_are_analyzed_and_promote_model_dependencies(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    page_dir = report / "definition" / "pages" / "Page 1"
    visual_dir = page_dir / "visuals" / "visual1"
    tables_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure 'Base Measure' = 1\n",
        encoding="utf-8",
    )
    (page_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "visual": {
                    "visualType": "card",
                    "query": {
                        "queryState": {
                            "Values": {
                                "projections": [
                                    {
                                        "field": {
                                            "Measure": {
                                                "Property": "Report Revenue",
                                                "Expression": {
                                                    "SourceRef": {
                                                        "Entity": "Sales"
                                                    }
                                                },
                                            }
                                        }
                                    }
                                ]
                            }
                        }
                    },
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (report / "definition" / "reportExtensions.json").write_text(
        json.dumps(
            {
                "$schema": "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/reportExtension/1.0.0/schema.json",
                "name": "extension",
                "entities": [
                    {
                        "name": "Sales",
                        "measures": [
                            {
                                "name": "Report Revenue",
                                "dataType": "Double",
                                "expression": "[Base Measure] + 1",
                                "displayFolder": "Executive",
                                "formatString": "0.0",
                                "references": {
                                    "measures": [
                                        {
                                            "entity": "Sales",
                                            "name": "Base Measure",
                                        }
                                    ]
                                },
                            }
                        ],
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(workspace.resolve())
    report_measure = _find_item(results, "Sales", "Report Revenue", "Measure")
    base_measure = _find_item(results, "Sales", "Base Measure", "Measure")

    assert report_measure["item"].source_kind == "report"
    assert report_measure["item"].source_artifact == "Executive"
    assert report_measure["item"].format_string == "0.0"
    assert report_measure["status"] == "USED"
    assert report_measure["usages"][0].report == "Executive"
    assert base_measure["status"].startswith("INDIRECT")
    assert results["summary"]["total_report_measures"] == 1

    payload = json.loads(analyzer.format_json_output(results))
    payload_item = next(item for item in payload["items"] if item["name"] == "Report Revenue")
    assert payload_item["sourceKind"] == "report"
    assert payload_item["sourceArtifact"] == "Executive"
    assert payload_item["formatString"] == "0.0"


def test_report_extension_measure_schema_gaps_are_report_health_issues(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    report_def_dir = report / "definition"
    tables_dir.mkdir(parents=True)
    report_def_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure Existing = 1\n",
        encoding="utf-8",
    )
    (report_def_dir / "reportExtensions.json").write_text(
        json.dumps(
            {
                "$schema": "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/reportExtension/1.0.0/schema.json",
                "name": "extension",
                "entities": [
                    {
                        "name": "Sales",
                        "measures": [
                            {
                                "name": "Report Revenue",
                                "expression": "[Existing] + 1",
                                "references": {
                                    "unrecognizedReferences": True
                                },
                            },
                            {
                                "dataType": "Double",
                                "expression": "[Existing]",
                            },
                            "not an object",
                        ],
                    },
                    {"measures": [{"name": "No Entity", "dataType": "Double", "expression": "1"}]},
                    "not an entity",
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(workspace.resolve())

    issue_keys = [
        (issue["severity"], issue["issueType"], issue["message"])
        for issue in results["report_issues"]
    ]
    assert issue_keys == [
        (
            "warning",
            "invalid_report_extension_measure",
            "Report extension measure 'Sales[Report Revenue]' is missing required field 'dataType'.",
        ),
        (
            "warning",
            "report_extension_unrecognized_reference",
            "Report extension measure 'Sales[Report Revenue]' contains unrecognized references.",
        ),
        (
            "warning",
            "invalid_report_extension_measure",
            "Report extension measure in entity 'Sales' is missing required field 'name'.",
        ),
        (
            "warning",
            "invalid_report_extension_measure",
            "Report extension measure at entities[0].measures[2] must be an object.",
        ),
        (
            "warning",
            "invalid_report_extension_entity",
            "Report extension entity at entities[1] is missing required field 'name'.",
        ),
        (
            "warning",
            "invalid_report_extension_entity",
            "Report extension entity at entities[2] must be an object.",
        ),
    ]
    assert all(issue["artifactKind"] == "Report Extension" for issue in results["report_issues"])
    assert all(issue["artifactPath"] == "definition/reportExtensions.json" for issue in results["report_issues"])

    payload = json.loads(analyzer.format_json_output(results))
    assert payload["reportIssues"][0]["issueType"] == "invalid_report_extension_measure"


def test_invalid_report_json_is_report_health_issue(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    page_dir = report / "definition" / "pages" / "Page1"
    visual_dir = page_dir / "visuals" / "Visual1"
    tables_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn Amount\n"
        "\t\tdataType: decimal\n",
        encoding="utf-8",
    )
    (page_dir / "page.json").write_text(
        json.dumps({"displayName": "Overview"}),
        encoding="utf-8",
    )
    (visual_dir / "visual.json").write_text("{ bad json", encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())

    assert len(results["report_issues"]) == 1
    issue = results["report_issues"][0]
    assert {
        "severity": issue["severity"],
        "issueType": issue["issueType"],
        "report": issue["report"],
        "page": issue["page"],
        "visualId": issue["visualId"],
        "artifactKind": issue["artifactKind"],
        "artifactPath": issue["artifactPath"],
        "message": issue["message"],
    } == {
        "severity": "error",
        "issueType": "invalid_report_json",
        "report": "Executive",
        "page": "Overview",
        "visualId": "Visual1",
        "artifactKind": "Visual",
        "artifactPath": "definition/pages/Page1/visuals/Visual1/visual.json",
        "message": "Could not parse PBIR JSON file: Expecting property name enclosed in double quotes: line 1 column 3 (char 2)",
    }

    payload = json.loads(analyzer.format_json_output(results))
    assert payload["reportIssues"][0]["issueType"] == "invalid_report_json"


def test_invalid_definition_pbir_is_report_health_issue(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn Amount\n"
        "\t\tdataType: decimal\n",
        encoding="utf-8",
    )
    (report / "definition.pbir").write_text("{ bad json", encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())

    assert results["report_issues"][0]["issueType"] == "invalid_report_json"
    assert results["report_issues"][0]["artifactKind"] == "Report Definition"
    assert results["report_issues"][0]["artifactPath"] == "definition.pbir"


def test_tmsl_model_bim_semantic_model_fails_clearly(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    model.mkdir(parents=True)
    report.mkdir(parents=True)
    (model / "model.bim").write_text("{}", encoding="utf-8")
    (model / "definition.pbism").write_text('{"version":"4.0"}', encoding="utf-8")
    (report / "definition.pbir").write_text('{"version":"4.0"}', encoding="utf-8")

    with pytest.raises(analyzer.UnsupportedSemanticModelError) as exc_info:
        analyzer.analyze(workspace.resolve())

    message = str(exc_info.value)
    assert "TMSL/model.bim" in message
    assert "Convert the Semantic Model to TMDL" in message


def test_dax_dependency_graph_excludes_self_references():
    item = analyzer.ModelItem(
        item_type="Measure",
        table="Sales",
        name="Revenue",
        dax_body="[Revenue] + 1",
    )

    deps = analyzer.build_dax_dependency_graph([item])

    assert deps[item.key] == set()


def test_unused_measure_with_dax_dependents_is_caution(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Test.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)

    (tables_dir / "Measures.tmdl").write_text(
        "table Measures\n"
        "\tmeasure A = 1\n"
        "\tmeasure B = [A] + 1\n",
        encoding="utf-8",
    )
    (report / "definition.pbir").write_text('{"version":"4.0"}', encoding="utf-8")
    (report / "report.json").write_text('{"sections":[]}', encoding="utf-8")

    results = analyzer.analyze(workspace)
    measure_a = _find_item(results, "Measures", "A", "Measure")

    assert measure_a["status"] == "NOT USED"
    assert measure_a["removal_risk"] == "Caution"


def test_unresolved_field_parameter_targets_emit_warnings(tmp_path):
    warnings = []
    info = analyzer.FieldParameterInfo(
        table="FP PNL",
        source_file=tmp_path / "FP PNL.tmdl",
        targets=[("_Measures", "Summary Actuals")],
    )
    items = [
        analyzer.ModelItem(item_type="Measure", table="Sales", name="Revenue"),
    ]

    resolved = analyzer.resolve_field_parameter_targets([info], items, "Model.SemanticModel", warnings)

    assert resolved == [(info, [])]
    assert len(warnings) == 1
    assert warnings[0].code == "UNRESOLVED_NAMEOF_TARGET"
    assert "NAMEOF target _Measures[Summary Actuals] from field parameter table FP PNL does not resolve" in warnings[0].message


def test_commented_dax_refs_are_excluded_from_dependencies_but_exposed():
    target = analyzer.ModelItem(item_type="Measure", table="_Measures", name="Total Other SG&A")
    source = analyzer.ModelItem(
        item_type="Measure",
        table="_Measures",
        name="EBIT - Actuals",
        dax_body="// [Total Other SG&A]\nCALCULATE([Transaction Amount - USD])",
    )
    base = analyzer.ModelItem(item_type="Measure", table="_Measures", name="Transaction Amount - USD")

    deps = analyzer.build_dax_dependency_graph([target, source, base])
    commented = analyzer.extract_dax_commented_refs(source.dax_body)

    assert target.key not in deps[source.key]
    assert base.key in deps[source.key]
    assert "[Total Other SG&A]" in commented


def test_table_summaries_capture_role_patterns_and_single_column_measures(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    pages_dir = report / "definition" / "pages" / "Page 1"
    visual_dir = pages_dir / "visuals" / "visual1"
    tables_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn OrderDateKey\n"
        "\tcolumn Amount\n"
        "\tmeasure Order Date Count = DISTINCTCOUNT(Sales[OrderDateKey])\n",
        encoding="utf-8",
    )
    (tables_dir / "Date.tmdl").write_text(
        "table Date\n"
        "\tcolumn DateKey\n"
        "\t\tisKey\n"
        "\tcolumn CalendarDate\n",
        encoding="utf-8",
    )
    (model / "definition" / "relationships.tmdl").write_text(
        "relationship SalesToDate\n"
        "\tfromColumn: Sales.OrderDateKey\n"
        "\ttoColumn: Date.DateKey\n"
        "\tfromCardinality: many\n"
        "\ttoCardinality: one\n"
        "\tisActive: true\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "visual": {
                    "visualType": "tableEx",
                    "query": {
                        "queryState": {
                            "Values": {
                                "projections": [
                                    {
                                        "field": {
                                            "Measure": {
                                                "Property": "Order Date Count",
                                                "Expression": {"SourceRef": {"Entity": "Sales"}},
                                            }
                                        }
                                    }
                                ]
                            },
                            "Category": {
                                "projections": [
                                    {
                                        "field": {
                                            "Column": {
                                                "Property": "CalendarDate",
                                                "Expression": {"SourceRef": {"Entity": "Date"}},
                                            }
                                        }
                                    }
                                ]
                            },
                        }
                    },
                }
            }
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(workspace.resolve())
    table_summaries = {row["name"]: row for row in results["table_summaries"]}

    assert table_summaries["Date"]["role_label"] == "dimension-like"
    assert table_summaries["Sales"]["role_label"] == "fact-like"
    assert table_summaries["Date"]["relationship_only_columns"] == ["Date[DateKey]"]
    assert table_summaries["Date"]["direct_report_measure_count"] == 0
    assert table_summaries["Date"]["direct_report_column_count"] == 1
    assert table_summaries["Sales"]["direct_report_measure_count"] == 1
    assert table_summaries["Sales"]["direct_report_column_count"] == 0
    assert table_summaries["Sales"]["single_column_measures"] == [
        {"measure": "Sales[Order Date Count]", "column": "Sales[OrderDateKey]"}
    ]


def test_broken_measure_refs_are_flagged(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tmeasure Revenue = [Missing Measure] + SUM('Missing Table'[Amount])\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())
    revenue = _find_item(results, "Sales", "Revenue", "Measure")

    assert revenue["status"] == "BROKEN (2 missing refs)"
    assert revenue["broken_dax_refs"] == ["[Missing Measure]", "Missing Table[Amount]"]
    assert revenue["broken_dax_ref_details"] == [
        {"kind": "measure", "ref": "[Missing Measure]", "message": "Missing measure: [Missing Measure]"},
        {"kind": "column", "ref": "Missing Table[Amount]", "message": "Missing column (table not found): Missing Table[Amount]"},
    ]
    assert results["summary"]["broken"] == 1

    payload = json.loads(analyzer.format_json_output(results))
    payload_item = next(item for item in payload["items"] if item["table"] == "Sales" and item["name"] == "Revenue")
    assert payload_item["brokenDaxRefs"] == ["[Missing Measure]", "Missing Table[Amount]"]
    assert payload_item["brokenDaxRefDetails"] == [
        {"kind": "measure", "ref": "[Missing Measure]", "message": "Missing measure: [Missing Measure]"},
        {"kind": "column", "ref": "Missing Table[Amount]", "message": "Missing column (table not found): Missing Table[Amount]"},
    ]


def test_bare_table_refs_are_flagged_as_broken(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)

    (tables_dir / "Measures.tmdl").write_text(
        "table Measures\n"
        "\tmeasure Row Count = COUNTROWS('Account')\n",
        encoding="utf-8",
    )

    (report / "definition.pbir").write_text('{"version":"4.0"}', encoding="utf-8")
    (report / "report.json").write_text('{"sections":[]}', encoding="utf-8")

    results = analyzer.analyze(workspace)
    row_count = _find_item(results, "Measures", "Row Count", "Measure")

    assert row_count["status"] == "BROKEN (1 missing ref)"
    assert row_count["broken_dax_refs"] == ["Account"]
    assert row_count["broken_dax_ref_details"] == [
        {"kind": "table", "ref": "Account", "message": "Missing table: Account"},
    ]
    assert results["summary"]["broken"] == 1


def test_bare_table_refs_are_exposed_in_table_dependency_signals(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Sales.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)

    (tables_dir / "Sales.tmdl").write_text(
        "table Sales\n"
        "\tcolumn Amount\n",
        encoding="utf-8",
    )
    (tables_dir / "Measures.tmdl").write_text(
        "table Measures\n"
        "\tmeasure Row Count = COUNTROWS('Sales')\n",
        encoding="utf-8",
    )

    (report / "definition.pbir").write_text('{"version":"4.0"}', encoding="utf-8")
    (report / "report.json").write_text('{"sections":[]}', encoding="utf-8")

    results = analyzer.analyze(workspace)
    table_summaries = {row["name"]: row for row in results["table_summaries"]}

    assert table_summaries["Sales"]["external_dax_dependents"] == ["Measures[Row Count]"]
    assert "1 DAX item outside this table depends on it." in table_summaries["Sales"]["signals"]


def test_valid_quoted_column_refs_do_not_trigger_false_broken_table_refs(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Forecast.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)

    (tables_dir / "Date.tmdl").write_text(
        "table Date\n"
        "\tcolumn Month\n"
        "\t\tsummarizeBy: none\n",
        encoding="utf-8",
    )
    (tables_dir / "Measures.tmdl").write_text(
        "table Measures\n"
        "\tmeasure Forecast Amount - USD - March =\n"
        "\t\tCALCULATE(\n"
        "\t\t\t[Forecast Amount - USD - March (view)],\n"
        "\t\t\tKEEPFILTERS('Date'[Month] IN { 4, 5, 6, 7, 8, 9, 10, 11, 12 })\n"
        "\t\t)\n"
        "\t\t+\n"
        "\t\tCALCULATE(\n"
        "\t\t\t[Transaction Amount - USD],\n"
        "\t\t\tKEEPFILTERS('Date'[Month] IN { 1, 2, 3 })\n"
        "\t\t)\n"
        "\tmeasure Forecast Amount - USD - March (view) = 1\n"
        "\tmeasure Transaction Amount - USD = 1\n",
        encoding="utf-8",
    )

    (report / "definition.pbir").write_text('{"version":"4.0"}', encoding="utf-8")
    (report / "report.json").write_text('{"sections":[]}', encoding="utf-8")

    results = analyzer.analyze(workspace)
    row = _find_item(results, "Measures", "Forecast Amount - USD - March", "Measure")

    assert row["status"] != "BROKEN (1 missing ref)"
    assert row["broken_dax_refs"] == []


def test_inline_calculated_columns_are_parsed_and_do_not_false_flag_measures(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Forecast.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    tables_dir.mkdir(parents=True)
    report.mkdir(parents=True)

    (tables_dir / "Date.tmdl").write_text(
        "table Date\n"
        "\tcolumn 'HasActuals USD' = CALCULATE(IF(ISBLANK([Transaction Amount - D365 - USD]), 0, 1), ALLEXCEPT('Date', 'Date'[YearMonth]))\n"
        "\tcolumn YearMonthNumber = 'Date'[Year]*12+'Date'[Month]\n"
        "\tcolumn Year\n"
        "\tcolumn Month\n"
        "\tcolumn YearMonth\n",
        encoding="utf-8",
    )
    (tables_dir / "Measures.tmdl").write_text(
        "table Measures\n"
        "\tmeasure Forecast Amount - USD - September - Dynamic (vena) =\n"
        "\t\tVAR _ForecastCutoff =\n"
        "\t\t\tCALCULATE(\n"
        "\t\t\t\tMAX('Date'[YearMonthNumber]),\n"
        "\t\t\t\t'Date'[HasActuals USD] = 1,\n"
        "\t\t\t\tREMOVEFILTERS('Date')\n"
        "\t\t\t)\n"
        "\t\tRETURN CALCULATE([Base Measure], KEEPFILTERS('Date'[YearMonthNumber] <= _ForecastCutoff))\n"
        "\tmeasure Base Measure = 1\n"
        "\tmeasure Transaction Amount - D365 - USD = 1\n",
        encoding="utf-8",
    )

    (report / "definition.pbir").write_text('{\"version\":\"4.0\"}', encoding="utf-8")
    (report / "report.json").write_text('{\"sections\":[]}', encoding="utf-8")

    parsed = analyzer.parse_model_items(model)
    parsed_refs = {(item.table, item.name, item.item_type) for item in parsed}
    assert ("Date", "HasActuals USD", "Calculated Column") in parsed_refs
    assert ("Date", "YearMonthNumber", "Calculated Column") in parsed_refs

    results = analyzer.analyze(workspace)
    measure = _find_item(results, "Measures", "Forecast Amount - USD - September - Dynamic (vena)", "Measure")
    assert measure["broken_dax_refs"] == []


def test_quoted_tmdl_identifiers_with_apostrophes_analyze_consistently(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Quoted.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    page_dir = report / "definition" / "pages" / "Page 1"
    visual_dir = page_dir / "visuals" / "visual1"
    tables_dir.mkdir(parents=True)
    visual_dir.mkdir(parents=True)

    (tables_dir / "O'Brien.tmdl").write_text(
        "table 'O''Brien'\n"
        "\tcolumn 'Customer''s Id'\n"
        "\tcolumn Amount\n"
        "\tcolumn 'Segment''s Sort'\n"
        "\tcolumn 'Segment''s Name'\n"
        "\t\tsortByColumn: 'Segment''s Sort'\n"
        "\tcolumn 'Geography''s Name'\n"
        "\tmeasure 'Bob''s Revenue' = SUM('O''Brien'[Amount])\n"
        "\thierarchy 'Customer''s Hierarchy'\n"
        "\t\tlevel 'Geography''s Name'\n"
        "\t\t\tcolumn: 'Geography''s Name'\n",
        encoding="utf-8",
    )
    (tables_dir / "Customer.tmdl").write_text(
        "table Customer\n"
        "\tcolumn CustomerId\n",
        encoding="utf-8",
    )
    (tables_dir / "Metrics.tmdl").write_text(
        "table Metrics\n"
        "\tmeasure 'Executive Bob''s Revenue' = 'O''Brien'[Bob's Revenue] + COUNTROWS('O''Brien')\n",
        encoding="utf-8",
    )
    (model / "definition" / "relationships.tmdl").write_text(
        "relationship 'O''Brien Customer'\n"
        "\tfromColumn: 'O''Brien'.'Customer''s Id'\n"
        "\ttoColumn: Customer.CustomerId\n"
        "\tfromCardinality: many\n"
        "\ttoCardinality: one\n",
        encoding="utf-8",
    )
    (page_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")
    (visual_dir / "visual.json").write_text(
        json.dumps(
            {
                "visual": {
                    "visualType": "tableEx",
                    "query": {
                        "queryState": {
                            "Values": {
                                "projections": [
                                    {
                                        "field": {
                                            "Measure": {
                                                "Property": "Executive Bob's Revenue",
                                                "Expression": {"SourceRef": {"Entity": "Metrics"}},
                                            }
                                        }
                                    },
                                    {
                                        "field": {
                                            "Column": {
                                                "Property": "Segment's Name",
                                                "Expression": {"SourceRef": {"Entity": "O'Brien"}},
                                            }
                                        }
                                    },
                                    {
                                        "field": {
                                            "HierarchyLevel": {
                                                "Level": "Geography's Name",
                                                "Expression": {
                                                    "Hierarchy": {
                                                        "Hierarchy": "Customer's Hierarchy",
                                                        "Expression": {"SourceRef": {"Entity": "O'Brien"}},
                                                    }
                                                },
                                            }
                                        }
                                    },
                                ]
                            }
                        }
                    },
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    results = analyzer.analyze(workspace.resolve())

    executive = _find_item(results, "Metrics", "Executive Bob's Revenue", "Measure")
    bob = _find_item(results, "O'Brien", "Bob's Revenue", "Measure")
    amount = _find_item(results, "O'Brien", "Amount", "Column")
    customer_id = _find_item(results, "O'Brien", "Customer's Id", "Column")
    segment_name = _find_item(results, "O'Brien", "Segment's Name", "Column")
    segment_sort = _find_item(results, "O'Brien", "Segment's Sort", "Column")
    geography = _find_item(results, "O'Brien", "Geography's Name", "Column")
    table_summaries = {row["name"]: row for row in results["table_summaries"]}

    assert executive["status"] == "USED"
    assert executive["broken_dax_refs"] == []
    assert bob["status"].startswith("INDIRECT")
    assert amount["status"].startswith("INDIRECT")
    assert customer_id["status"] == "USED (Relationship)"
    assert segment_name["status"] == "USED"
    assert segment_sort["status"] == "USED (Sort Column for: Segment's Name)"
    assert geography["status"] == "USED (Hierarchy: Customer's Hierarchy)"
    assert table_summaries["O'Brien"]["external_dax_dependents"] == [
        "Metrics[Executive Bob's Revenue]"
    ]
    assert results["summary"]["broken"] == 0


def test_field_parameter_tables_warn_on_unresolved_nameof_targets(tmp_path):
    workspace = tmp_path / "Workspace"
    model = workspace / "Models" / "Forecast.SemanticModel"
    report = workspace / "Reports" / "Executive.Report"
    tables_dir = model / "definition" / "tables"
    pages_dir = report / "definition" / "pages" / "Page 1"
    tables_dir.mkdir(parents=True)
    pages_dir.mkdir(parents=True)

    (tables_dir / "_Measures.tmdl").write_text(
        "table _Measures\n"
        "\tmeasure 'Summary Actuals' = 1\n"
        "\tmeasure 'Summary Budget' = 1\n"
        "\tmeasure 'Summary FC2' = 1\n",
        encoding="utf-8",
    )
    (tables_dir / "FP PNL.tmdl").write_text(
        "table 'FP PNL'\n"
        "\tcolumn 'FP PNL'\n"
        "\t\tsummarizeBy: none\n"
        "\tcolumn 'FP PNL Fields'\n"
        "\t\tsummarizeBy: none\n"
        "\tpartition 'FP PNL' = calculated\n"
        "\t\tsource = ```\n"
        "\t\t\t{\n"
        "\t\t\t    (\"Summary Actuals\", NAMEOF('_Measures'[Summary Actuals]), 1, \"MTH\"),\n"
        "\t\t\t    (\"Summary Budget\", NAMEOF('_Measures'[Summary Budget]), 2, \"MTH\"),\n"
        "\t\t\t    //(\"Summary FC1\", NAMEOF('_Measures'[Summary FC1]), 3, \"MTH\"),\n"
        "\t\t\t    (\"Summary Missing Active\", NAMEOF('_Measures'[Summary Missing Active]), 4, \"MTH\"),\n"
        "\t\t\t    (\"Summary Actuals\", NAMEOF('_Measures'[Summary Actuals]), 5, \"BUD\")\n"
        "\t\t\t    // (\"6\", NAMEOF('_Measures'[Summary Missing]), 6, \"BUD\")\n"
        "\t\t\t}\n"
        "\t\t\t```\n",
        encoding="utf-8",
    )
    (pages_dir / "page.json").write_text('{"displayName":"Overview"}', encoding="utf-8")

    results = analyzer.analyze(workspace.resolve())

    unresolved = [w for w in results["warnings"] if w["code"] == "UNRESOLVED_NAMEOF_TARGET"]
    assert len(unresolved) == 1
    assert "_Measures[Summary Missing Active]" in unresolved[0]["message"]
    assert "_Measures[Summary FC1]" not in unresolved[0]["message"]
    assert "_Measures[Summary Missing]" not in unresolved[0]["message"]

    table_summary = next(table for table in results["table_summaries"] if table["name"] == "FP PNL")
    assert table_summary["field_parameter_issues"] == [unresolved[0]["message"]]
    assert table_summary["signals"][0] == "Broken field parameter: 1 unresolved NAMEOF target."
